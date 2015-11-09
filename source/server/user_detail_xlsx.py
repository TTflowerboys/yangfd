# coding: utf-8
from __future__ import unicode_literals
from bson.objectid import ObjectId
from datetime import datetime
from app import f_app
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pytz import timezone
import pytz

f_app.common.memcache_server = ["172.20.101.102:11211"]
f_app.common.mongo_server = "172.20.101.102"


def get_data_directly(user, part, deep=None):
    if user is None:
        return
    user_part = user.get(part, None)
    if (user_part is not None) and (deep is not None):
        return user.get(part).get(deep, None)
    return user_part


def get_data_directly_as_str(user, part, deep=None):
    if user is None:
        return ''
    user_part = user.get(part, None)
    if (user_part is not None) and (deep is not None):
        return user.get(part).get(deep, None)
    if f_app.util.batch_iterable(user_part):
        user_part = '/'.join(user_part)
    if user_part is None:
        return ''
    if isinstance(user_part, datetime):
        loc_t = timezone('Europe/London')
        loc_dt = loc_t.localize(user_part)
        return unicode(loc_dt.strftime('%Y-%m-%d %H:%M:%S %Z%z'))
    return unicode(user_part)


def get_all_enum_value(enum_singlt_type):
    print "enum type " + enum_singlt_type + " loading."
    enum_list_subdic = {}
    for enumitem in f_app.i18n.process_i18n(f_app.enum.get_all(enum_singlt_type)):
        enum_list_subdic.update({enumitem["id"]: enumitem["value"]})
    enum_type_list.update({enum_singlt_type: enum_list_subdic})


def get_data_enum(user, enum_name):
    if user is None:
        return
    if enum_name not in enum_type_list:
        get_all_enum_value(enum_name)
    single = user.get(enum_name, None)
    value_list = []
    if f_app.util.batch_iterable(single):
        for true_single in single:
            if true_single is None:
                continue
            enum_id = true_single.get("id", None)
            value = enum_type_list[enum_name].get(enum_id, None)
            value_list.append(value)
    elif single is not None:
        if single.get("id", None):
            enum_id = unicode(single.get("id", None))
            value = enum_type_list[enum_name].get(enum_id, None)
            if value is not None:
                value_list.append(value)
    if not f_app.util.batch_iterable(value_list):
        value_list = [value_list]
    value_set = set(value_list)
    value_list = list(value_set)
    return unicode('/'.join(value_list))


def get_data_complex(user, target, condition, element):

    '''this func make dict provide get_data_enum to use.
    with user's id and 'condition' to search in the database 'target'
    then gether element in search result, make a new dict return
    '''

    dic = {}
    user_id = user.get("id", None)
    target_database = getattr(f_app, target)
    condition.update({"$or": [{"user_id": ObjectId(user_id)},
                              {"creator_user_id": ObjectId(user_id)}]})
    select_item = target_database.get(target_database.search(condition, per_page=-1))
    element_list = []
    for ticket in select_item:
        element_list.append(ticket.get(element, None))
    dic.update({element: element_list})
    return dic


def get_ticket_newest(user, add_condition={}):
    user_id = user.get("id", None)
    if user_id is None:
        return {}
    condition = ({"type": "rent",
                  "$or": [{"user_id": ObjectId(user_id)},
                          {"creator_user_id": ObjectId(user_id)}]})
    condition.update(add_condition)
    time_list = []
    select_item = f_app.ticket.get(f_app.ticket.search(condition, per_page=-1))
    if select_item is None:
        return {}
    for single in select_item:
        curtime = single.get("time", None)
        time_list.append(curtime)
    if len(time_list) < 1:
        return {}
    time_list.sort()
    for single in select_item:
        curtime = single.get("time", None)
        if curtime == time_list[0]:
            return single
    return {}


def time_period_label(ticket):
    if ticket is None:
        return ''
    time = ""
    period_start = ticket.get("rent_available_time", None)
    period_end = ticket.get("rent_deadline_time", None)
    if period_end is None or period_start is None:
        time = "不明"
    else:
        period = period_end - period_start
        if period.days >= 365:
            time = "longer than 12 months"
        elif 365 > period.days >= 180:
            time = "6 - 12 months"
        elif 180 >= period.days > 90:
            time = "3 - 6 months"
        elif period.days <= 30:
            time = "less than 1 month"
    return time


def get_count(user, target, condition, element, comp):
    dic = get_data_complex(user, target, condition, element)
    if dic.get(element, []) is None:
        return 'a'
    return dic.get(element, []).count(comp)


def get_has_flag(user, target, condition, comp_element, want_value):
    '''this func '''
    dic = get_data_complex(user, target, condition, comp_element)
    return want_value in dic.get(comp_element, None)


def get_max(user, target, condition, comp_element):
    dic = get_data_complex(user, target, condition, comp_element)
    element_list = dic.get(comp_element, None)
    element_list.sort()
    if element_list is not None and len(element_list):
        return element_list[0]
    return ''


def get_correct_col_index(num):
    if num > 26*26:
        return "ZZ"
    if num >= 26:
        return get_correct_col_index(num/26-1)+chr(num-26+65)
    else:
        return chr(num+65)


def format_fit(sheet):
    simsun_font = Font(name="SimSun")
    alignment_fit = Alignment(shrink_to_fit=True)
    for row in sheet.rows:
        for cell in row:
            cell.font = simsun_font
            cell.alignment = alignment_fit
    for num, col in enumerate(sheet.columns):
        lenmax = 0
        for cell in col:
            lencur = 0
            if cell.value is None:
                cell.value = ''
            if isinstance(cell.value, int) or isinstance(cell.value, datetime):
                lencur = len(unicode(cell.value).encode("GBK"))
            elif cell.value is not None:
                lencur = len(cell.value.encode("GBK", "replace"))
            if lencur > lenmax:
                lenmax = lencur
        sheet.column_dimensions[get_correct_col_index(num)].width = lenmax*0.86
        print "col "+get_correct_col_index(num)+" fit."


def get_diff_color(total):
    fill = []
    base_color = 0x999999
    color = 0x0
    if 4 <= total <= 6:
        s = 0x111111
    elif total == 3:
        s = 0x222222
    elif total == 2:
        s = 0x333333
    else:
        return
    for index in range(total):
        color = base_color + s*index
        color_t = '00'+"%x" % color
        fill.append(PatternFill(fill_type='solid', start_color=color_t, end_color=color_t))
    return fill


def be_colorful(sheet, max_segment):
    header_fill = PatternFill(fill_type='solid', start_color='00dddddd', end_color='00dddddd')
    for cell in sheet.rows[0]:
        cell.fill = header_fill
    for col in sheet.columns:
        col_set = set()
        cell_fill = []
        col_list = []
        for num, cell in enumerate(col):
            if num and len(cell.value):
                col_set.add(cell.value)
            if len(col_set) > max_segment:
                break
        if max_segment >= len(col_set) > 1:
            cell_fill = get_diff_color(len(col_set))
            col_list = list(col_set)
            for number, cell in enumerate(col):
                if not number:
                    continue
                for index, value in enumerate(col_list):
                    if cell.value == value:
                        cell.fill = cell_fill[index]


def check_download(user):
    user_id = user.get('id', None)
    credit = f_app.user.credit.get("view_rent_ticket_contact_info", user_id).get("credits", [])
    for single in credit:
        if single.get("tag", None) == "download_ios_app":
            return True
    return False


def get_detail_address(ticket):
    ticket = f_app.i18n.process_i18n(ticket)
    return ' '.join([ticket.get("country", {}).get("code", ''),
                     ticket.get("city", {}).get("name", ''),
                     ticket.get("maponics_neighborhood", {}).get("name", ''),
                     ticket.get("address", ''),
                     ticket.get("zipcode_index", '')])


def get_address(user):
    property_id = get_data_complex(user, "ticket", {"type": "rent"}, "property_id").get("property_id", [])
    if len(property_id) < 1:
        return ''
    elif len(property_id) > 1:
        return 'M'
    else:
        return get_detail_address(f_app.property.get(property_id[0]))


def get_match(ticket):
    match = []
    if "partial_match" in ticket.get("tags", []):
        match.append("部分满足")
    if "perfect_match" in ticket.get("tags", []):
        match.append("完全满足")
    return '/'.join(match)


def get_log_with_id(user, params={}):
    user_id = user.get("id", None)
    if user_id is None:
        return None
    params.update({"id": ObjectId(user_id)})
    select_log = f_app.log.output(f_app.log.search(params, per_page=-1))
    print f_app.util.json_dumps(select_log)
    if select_log is None:
        return []
    return select_log


def get_active_days(logs):
    if logs is None:
        return "0"
    set_list = set()
    for log in logs:
        single_time = log.get("time", None)
        if single_time is None:
            continue
        curtime = datetime(single_time.year, single_time.month, single_time.day)
        set_list.add(curtime)
    return unicode(len(set_list))


enum_type_list = {}

header = ['用户名', '注册时间', '国家', '用户类型', '单独访问次数', '活跃天数',
          'app下载', '房东类型', '有没有草稿', '发布时间', '地区', '房产查看数',
          '房产量', '单间还是整套', '短租长租', '租金', '分享房产', '已出租时间',
          '求租时间', '预算', '地区', '匹配级别', '查看房产次数', '收藏房产次数',
          '查看房东联系方式的次数', '分享房产', '停留时间最多的页面或rental房产',
          '投资意向时间', '投资预算', '期房还是现房', '几居室', '浏览数量',
          '停留时间最多的页面或sales房产', '跳出的页面']

wb = Workbook()
ws = wb.active

ws.append(header)

for number, user in enumerate(f_app.user.get(f_app.user.get_active())):
    ws.append([get_data_directly_as_str(user, "nickname"),
               get_data_directly_as_str(user, "register_time"),
               get_data_directly_as_str(user, "country", "code"),
               get_data_enum(user, "user_type"),
               "",
               get_active_days(get_log_with_id(user)),
               "已下载" if check_download(user) else "未下载",
               get_data_enum(get_data_complex(user, "ticket", {"type": "rent"}, "landlord_type"), "landlord_type"),
               "有" if get_has_flag(user, "ticket", {"type": "rent"}, "status", "draft") else "无",
               unicode(get_max(user, "ticket", {"type": "rent"}, "time")),
               get_address(user),
               "",
               unicode(get_count(user, "ticket", {"type": "rent"}, "type", "rent")),
               get_data_enum(get_data_complex(user, "ticket", {"type": "rent"}, "rent_type"), "rent_type"),
               time_period_label(get_ticket_newest(user)),
               get_ticket_newest(user).get("price", ''),
               "",
               get_data_directly_as_str(get_ticket_newest(user, {"type": "rent", "status": "rent"}), "time"),
               get_data_directly_as_str(get_ticket_newest(user, {"type": "rent_intention", "status": "new"}), "time"),
               get_address(get_ticket_newest(user, {"type": "rent_intention", "status": "new"})),
               get_match(get_ticket_newest(user, {"type": "rent_intention", "status": "new"})),
               unicode(len(get_log_with_id(user, {"type": "route",
                                                  "property_id": {"$exists": True}
                                                  }))),
               "",
               unicode(len(get_log_with_id(user, {"type": "rent_ticket_view_contact_info"}))),
               ])
    print 'user.' + unicode(number) + ' done.'
    if number >= 20:
        break
format_fit(ws)
be_colorful(ws, 6)
wb.save("user_detail.xlsx")
