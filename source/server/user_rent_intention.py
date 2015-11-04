# coding: utf-8
from __future__ import unicode_literals
# from bson.objectid import ObjectId
from datetime import datetime
from datetime import timedelta
from app import f_app
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

f_app.common.memcache_server = ["172.20.101.102:11211"]
f_app.common.mongo_server = "172.20.101.102"


def get_data_directly_as_str(user, part, deep=None):
    if user is None:
        return ''
    user_part = user.get(part, None)
    if (user_part is not None) and (deep is not None):
        return user.get(part).get(deep, None)
    if f_app.util.batch_iterable(user_part):
        user_part = '/'.join(user_part)
    if user_part is None:
        return ''
    return unicode(user_part)


def get_data_enum(user, enum_name):
    if user is None:
        return
    if enum_name not in enum_type_list:
        get_all_enum_value(enum_name)
    single = user.get(enum_name, None)
    value_list = []
    if f_app.util.batch_iterable(single):
        for true_single in single:
            if true_single is None:
                continue
            enum_id = true_single.get("id", None)
            value = enum_type_list[enum_name].get(enum_id, None)
            value_list.append(value)
    elif single is not None:
        if single.get("id", None):
            enum_id = str(single.get("id", None))
            value = enum_type_list[enum_name].get(enum_id, None)
            #print enum_id
            #print enum_type_list[enum_name]
            if value is not None:
                value_list.append(value)
    if not f_app.util.batch_iterable(value_list):
        value_list = [value_list]
    value_set = set(value_list)
    value_list = list(value_set)
    return '/'.join(value_list)


def get_correct_col_index(num):
    if num > 26*26:
        return "ZZ"
    if num >= 26:
        return get_correct_col_index(num/26-1)+chr(num-26+65)
    else:
        return chr(num+65)


def add_link(sheet, target, link=None):
    if target is None:
        return
    if f_app.util.batch_iterable(target):
        pass
    else:
        for index in range(2, len(sheet.rows)+1):
            cell = sheet[target + unicode(index)]
            if len(cell.value):
                if link is None:
                    cell.hyperlink = cell.value
                else:
                    cell.hyperlink = unicode(link)


def get_diff_color(total):
    fill = []
    base_color = 0x999999
    color = 0x0
    if 4 <= total <= 6:
        s = 0x111111
    elif total == 3:
        s = 0x222222
    elif total == 2:
        s = 0x333333
    else:
        return
    for index in range(total):
        color = base_color + s*index
        color_t = '00'+"%x" % color
        fill.append(PatternFill(fill_type='solid', start_color=color_t, end_color=color_t))
    return fill


def be_colorful(sheet, max_segment):
    header_fill = PatternFill(fill_type='solid', start_color='00dddddd', end_color='00dddddd')
    for cell in sheet.rows[0]:
        cell.fill = header_fill
    for col in sheet.columns:
        col_set = set()
        cell_fill = []
        col_list = []
        for num, cell in enumerate(col):
            if num and len(cell.value):
                col_set.add(cell.value)
            if len(col_set) > max_segment:
                break
        if max_segment >= len(col_set) > 1:
            cell_fill = get_diff_color(len(col_set))
            col_list = list(col_set)
            for number, cell in enumerate(col):
                if not number:
                    continue
                for index, value in enumerate(col_list):
                    if cell.value == value:
                        cell.fill = cell_fill[index]


def format_fit(sheet):
    simsun_font = Font(name="SimSun")
    alignment_fit = Alignment(shrink_to_fit=True)
    for row in sheet.rows:
        for cell in row:
            cell.font = simsun_font
            cell.alignment = alignment_fit
    for num, col in enumerate(sheet.columns):
        lenmax = 0
        for cell in col:
            lencur = 0
            if cell.value is None:
                cell.value = ''
            if isinstance(cell.value, int) or isinstance(cell.value, datetime):
                lencur = len(str(cell.value).encode("GBK"))
            elif cell.value is not None:
                lencur = len(cell.value.encode("GBK", "replace"))
            if lencur > lenmax:
                lenmax = lencur
        sheet.column_dimensions[get_correct_col_index(num)].width = lenmax*0.86
        print "col "+get_correct_col_index(num)+" fit."


def get_referer_id(ticket):
    time = ticket.get("time", None)
    diff_time = timedelta(milliseconds=500)
    flag = 0
    if time is None:
        return ''
    for num, single in enumerate(referer_result):
        rst_time = single.get("time", None)
        if rst_time is None:
            continue
        if rst_time - diff_time < time < rst_time + diff_time:
            if flag:
                print "bingo. "+str(rst_time)
            flag = 1
            record = single
    if flag:
        return get_id_in_url(record.get("referer", ''))
    print "cant find when "+str(time)
    return ''


def get_email(ticket):
    user = ticket.get("user", None)
    if user is None:
        return ''
    return user.get("email", '')


def get_wechat(ticket):
    user = ticket.get("user", None)
    if user is None:
        return ''
    return user.get("wechat", '')


def time_period_label(ticket):
    time = ""
    period_start = ticket.get("rent_available_time", None)
    period_end = ticket.get("rent_deadline_time", None)
    if period_end is None or period_start is None:
        time = "不明"
    else:
        period = period_end - period_start
        if period.days >= 365:
            time = "longer than 12 months"
        elif 365 > period.days >= 180:
            time = "6 - 12 months"
        elif 180 >= period.days > 90:
            time = "3 - 6 months"
        elif period.days <= 30:
            time = "less than 1 month"
    return time


def get_landlord_boss_with_ticket_id(landlord_ticket_id):
    if landlord_ticket_id is None:
        return None, None
    try:
        landlord_house = f_app.ticket.get(landlord_ticket_id)
    except:
        return None, None
    else:
        landlord_boss_id = landlord_house.get("user_id", None)
        if landlord_boss_id is None:
            return landlord_house, None
        return landlord_house, f_app.user.get(landlord_boss_id)


def get_id_in_url(url):
    if "property-to-rent/" in url:
        segment = url.split("property-to-rent/")[1]
        if "?" in segment:
            return segment.split("?")[0]
        return segment
    elif "ticketId=" in url:
        return url.split("ticketId=")[1]
    else:
        return None


def get_match(ticket):
    match = []
    if "partial_match" in ticket.get("tags", []):
        match.append("部分满足")
    if "perfect_match" in ticket.get("tags", []):
        match.append("完全满足")
    return '/'.join(match)


def get_detail_address(ticket):
    return ' '.join([ticket.get("country", {}).get("code", ''),
                     ticket.get("city", {}).get("name", ''),
                     ticket.get("maponics_neighborhood", {}).get("name", ''),
                     ticket.get("address", ''),
                     ticket.get("zipcode_index", '')])


def get_all_rent_intention():
    params = {"type": "rent_intention"}
    return f_app.ticket.output(f_app.ticket.search(params, per_page=20))


def get_all_enum_value(enum_singlt_type):
    print "enum type " + enum_singlt_type + " loading."
    enum_list_subdic = {}
    for enumitem in f_app.i18n.process_i18n(f_app.enum.get_all(enum_singlt_type)):
        enum_list_subdic.update({enumitem["id"]: enumitem["value"]})
    enum_type_list.update({enum_singlt_type: enum_list_subdic})


def get_referer_url(referer_id):
    if referer_id is None:
        return ''
    return "http://yangfd.com/admin?_i18n=zh_Hans_CN#/dashboard/rent/"+unicode(referer_id)

# enum_type = ["rent_type", "landlord_type"]

print "find all referer..."
referer_result = f_app.log.output(f_app.log.search({"route": "/api/1/rent_intention_ticket/add"}, per_page=-1))
enum_type_list = {}

wb = Workbook()
ws = wb.active
header = ["状态", "标题", "客户", "联系方式", "邮箱", "微信", "提交时间", "起始日期",
          "终止日期", "出租需求", "预算上限", "预算下限", "period", "出租位置", "备注",
          "样房东有无匹配搭配", "目标房源", "房东类型", "房东姓名", "房东电话", "房东邮箱", "打电话了？", "有接到？", "房子租到了么？", "通过样房东",
          "如果不是通过样房东，那么是通过哪里什么样的房源？有没有交中介费", "对平台体验的想法及反馈",
          "在找房子中用户最疼的点有哪些？", "备注"]
ws.append(header)

print "loading all rent intension ticket..."
for number, ticket in enumerate(get_all_rent_intention()):
    print 'ticket.' + str(number) + ' loading.'
    ticket = f_app.i18n.process_i18n(ticket)
    print 'ticket.' + str(number) + ' i18n process complete.'
    referer_id = get_referer_id(ticket)
    landlord_result, landlord_boss = get_landlord_boss_with_ticket_id(referer_id)
    ws.append(["已提交" if (get_data_directly_as_str(ticket, "status") == "new") else "已出租",
               get_data_directly_as_str(ticket, "title"),
               get_data_directly_as_str(ticket, "nickname"),
               get_data_directly_as_str(ticket, "phone"),
               get_email(ticket),
               get_wechat(ticket),
               get_data_directly_as_str(ticket, "time"),
               get_data_directly_as_str(ticket, "rent_available_time"),
               get_data_directly_as_str(ticket, "rent_deadline_time"),
               get_data_enum(ticket, "rent_type"),
               get_data_directly_as_str(ticket, "rent_budget_max", "value"),
               get_data_directly_as_str(ticket, "rent_budget_min", "value"),
               time_period_label(ticket),
               get_detail_address(ticket),
               get_data_directly_as_str(ticket, "description"),
               get_match(ticket),
               get_referer_url(referer_id),
               get_data_enum(landlord_result, "landlord_type"),  # 房东类型
               get_data_directly_as_str(landlord_boss, "nickname"),
               get_data_directly_as_str(landlord_boss, "phone"),
               get_data_directly_as_str(landlord_boss, "email"),
               ])

format_fit(ws)
add_link(ws, 'Q')
be_colorful(ws, 6)
wb.save("user_rent_intention.xlsx")
