# -*- coding: utf-8 -*-
from __future__ import unicode_literals, absolute_import
import logging
import calendar
from datetime import datetime
from datetime import timedelta
from datetime import date
from hashlib import sha1
from lxml import etree
from six.moves import cStringIO as StringIO
from six.moves import urllib
try:
    import qrcode
except ImportError:
    pass
import bottle
from bson.objectid import ObjectId
from app import f_app
from libfelix.f_interface import f_api, f_get, f_post, static_file, template, request, response, redirect, html_redirect, error, abort, template_gettext as _
import currant_util
import currant_data_helper
from libfelix.f_interface import f_experiment
from openpyxl import Workbook
import six
from openpyxl.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.writer.excel import save_virtual_workbook
from pytz import timezone
# import pytz
from bson.code import Code
import re
f_experiment()

logger = logging.getLogger(__name__)
f_app.dependency_register("qrcode", race="python")


@f_get('/', params=dict(
    _i18n=str
))
@currant_util.check_ip_and_redirect_domain
@f_app.user.login.check()
def default(user, params):
    property_list = []
    homepage_ad_list = []

    if user:
        homepage_ad_list = f_app.ad.get_all_by_channel("homepage_signedin")
        homepage_ad_list = f_app.i18n.process_i18n(homepage_ad_list)
    else:
        homepage_ad_list = f_app.ad.get_all_by_channel("homepage")
        homepage_ad_list = f_app.i18n.process_i18n(homepage_ad_list)
        property_list = currant_data_helper.get_featured_property_list()
        for property in property_list:
            if "news_category" in property:
                property["related_news"] = currant_data_helper.get_property_related_news_list(property)
        property_list = f_app.i18n.process_i18n(property_list)

    news_list = currant_data_helper.get_featured_new_list()
    news_list = f_app.i18n.process_i18n(news_list)

    intention_list = f_app.i18n.process_i18n(f_app.enum.get_all('intention'))

    hot_city_list = []
    hot_city_geonames_params = dict({
        "name": {"$in": ["London", "Liverpool", "Sheffield", "Manchester", "Birmingham"]},
        "feature_code": {"$in": ["PPLC", "PPLA", "PPLA2"]},
        "country": "GB"
    })
    hot_city_list = f_app.geonames.gazetteer.get(f_app.geonames.gazetteer.search(hot_city_geonames_params, per_page=-1))

    rent_type_list = f_app.i18n.process_i18n(f_app.enum.get_all('rent_type'))
    property_type_list = f_app.i18n.process_i18n(f_app.enum.get_all('property_type'))
    property_type_list_without_new_property = filter(lambda item: item.get('slug') != 'new_property', property_type_list)
    property_country_list = currant_util.get_country_list()

    property_city_list = []
    geonames_params = dict({
        "feature_code": {"$in": ["PPLC", "PPLA", "PPLA2"]},
        "country": "GB"
    })
    property_city_list = f_app.geonames.gazetteer.get(f_app.geonames.gazetteer.search(geonames_params, per_page=-1))

    title = _('洋房东')
    description = _('洋房东致力于为英国华人房东和租客提供专业和靠谱的租房找房和海外置业省时省力贴心安全快捷便利华人互联网房产平台')
    keywords = _('洋房东,租房,买房,出租,租房中介,找房子,短租,长租,租金,楼盘,公寓,别墅,学区房,英国置业,留学生租房,海外租房,英国出租,英国租房,伦敦租房,伦敦买房,海外置业,海外投资,英国房价,Youngfunding,for rent,to let,room to rent,property to rent,investment,overseas investment,property,apartment,house,UK property')
    lang = getattr(f_app.i18n, "get_gettext")("web").lang
    if lang == "en_GB":
        homepage_ad_list = f_app.ad.get_all_by_channel("homepage_uk")
        homepage_ad_list = f_app.i18n.process_i18n(homepage_ad_list)
        return currant_util.common_template(
            "index_en",
            title=title,
            description=description,
            keywords=keywords,
            property_list=property_list,
            homepage_ad_list=homepage_ad_list,
            news_list=news_list,
            intention_list=intention_list,
            property_country_list=property_country_list,
            property_city_list=property_city_list,
            rent_type_list=rent_type_list,
            property_type_list=property_type_list,
            property_type_list_without_new_property=property_type_list_without_new_property,
            icon_map=currant_util.icon_map,
            hot_city_list=hot_city_list
        )
    else:
        return currant_util.common_template(
            "index",
            title=title,
            description=description,
            keywords=keywords,
            property_list=property_list,
            homepage_ad_list=homepage_ad_list,
            news_list=news_list,
            intention_list=intention_list,
            property_country_list=property_country_list,
            property_city_list=property_city_list,
            rent_type_list=rent_type_list,
            property_type_list=property_type_list,
            property_type_list_without_new_property=property_type_list_without_new_property,
            icon_map=currant_util.icon_map,
            hot_city_list=hot_city_list
        )


@f_get('/signup')
@currant_util.check_ip_and_redirect_domain
def signup():
    title = _('注册')
    return currant_util.common_template("signup", title=title)


@f_get('/verify-phone')
@currant_util.check_ip_and_redirect_domain
@f_app.user.login.check(force=True)
def verify_phone(user):
    title = _('手机验证')
    return currant_util.common_template("verify_phone", title=title)


@f_get('/vip_sign_up')
@currant_util.check_ip_and_redirect_domain
def vip_sign_up():
    title = _('注册')
    return currant_util.common_template("sign_up_vip", title=title)


@f_get('/signin')
@currant_util.check_ip_and_redirect_domain
def signin():
    title = _('登录')
    return currant_util.common_template("signin", title=title)


@f_get('/affiliate-signup')
@currant_util.check_ip_and_redirect_domain
def affiliate_signup():
    title = _('Affiliate 注册')
    return currant_util.common_template("affiliate_signup", title=title)


@f_get('/intention')
@currant_util.check_ip_and_redirect_domain
@f_app.user.login.check(force=True)
# @currant_util.check_phone_verified_and_redirect_domain
def intention(user):
    title = _('选择服务需求')
    intention_list = f_app.i18n.process_i18n(f_app.enum.get_all('intention'))
    rent_type_list = f_app.i18n.process_i18n(f_app.enum.get_all('rent_type'))
    property_type_list = f_app.i18n.process_i18n(f_app.enum.get_all('property_type'))
    property_type_list_without_new_property = filter(lambda item: item.get('slug') != 'new_property', property_type_list)
    user_type_list = f_app.i18n.process_i18n(f_app.enum.get_all('user_type'))
    property_country_list = currant_util.get_country_list()
    country = "GB"
    geonames_params = dict({
        "feature_code": {"$in": ["PPLC", "PPLA", "PPLA2"]},
        "country": country
    })
    property_city_list = f_app.geonames.gazetteer.get(f_app.geonames.gazetteer.search(geonames_params, per_page=-1))
    return currant_util.common_template("intention", intention_list=intention_list, title=title, icon_map=currant_util.icon_map, rent_type_list=rent_type_list, property_type_list=property_type_list, property_type_list_without_new_property=property_type_list_without_new_property, property_country_list=property_country_list, property_city_list=property_city_list, user_type_list=user_type_list)


@f_get('/reset_password', '/reset-password')
@currant_util.check_ip_and_redirect_domain
def reset_password():
    title = _('重置密码')
    return currant_util.common_template("reset_password", title=title)


@f_get('/reset_password_phone', '/reset-password-phone')
@currant_util.check_ip_and_redirect_domain
def reset_password_phone():
    title = _('使用短信验证码重置密码')
    return currant_util.common_template("reset_password_phone", title=title)


@f_get('/reset_password_email_1', '/reset-password-email-1')
@currant_util.check_ip_and_redirect_domain
def reset_password_email_1():
    title = _('使用绑定邮箱重置密码')
    return currant_util.common_template("reset_password_email_1", title=title)


@f_get('/reset_password_email_2', '/reset-password-email-2', params=dict(
    user_id=str,
    code=str
))
@currant_util.check_ip_and_redirect_domain
def reset_password_email_2(params):
    title = _('使用绑定邮箱重置密码')
    return currant_util.common_template("reset_password_email_2", title=title, user_id=params['user_id'], code=params['code'])


@f_get('/property-for-sale/create')
@currant_util.check_ip_and_redirect_domain
@currant_util.check_crowdfunding_ready
def property_for_sale_create():
    title = _('二手房出售')
    return currant_util.common_template("property_for_sale_create", title=title)


@f_get('/property-for-sale/publish')
@currant_util.check_ip_and_redirect_domain
@currant_util.check_crowdfunding_ready
def property_for_sale_publish():
    title = _('出售预览')
    return currant_util.common_template("property_for_sale_publish", title=title)


@f_get('/host-contact-request/<rent_ticket_id:re:[0-9a-fA-F]{24}>')
@currant_util.check_ip_and_redirect_domain
@f_app.user.login.check(check_role=True)
def rent_ticket_get(rent_ticket_id, user):
    rent_ticket = f_app.i18n.process_i18n(f_app.ticket.output([rent_ticket_id], fuzzy_user_info=True)[0])
    if rent_ticket["status"] not in ["draft", "to rent"]:
        assert user and set(user["role"]) & set(["admin", "jr_admin", "operation", "jr_operation"]), abort(40300, "No access to specify status or target_rent_ticket_id")

    # report = None
    # if rent_ticket.get('zipcode_index') and rent_ticket.get('country').get('code') == 'GB':
    #     report = f_app.i18n.process_i18n(currant_data_helper.get_report(rent_ticket.get('zipcode_index')))

    title = _('我要咨询')
    contact_info_already_fetched = len(f_app.order.search({
        "items.id": f_app.common.view_rent_ticket_contact_info_id,
        "ticket_id": rent_ticket_id,
        "user.id": user["id"],
    })) > 0 if user else False

    private_contact_methods = rent_ticket.get('creator_user', {}).get('private_contact_methods', [])
    private_contact_methods = [] if private_contact_methods is None else private_contact_methods
    return currant_util.common_template("host_contact_request-phone", rent=rent_ticket, title=title, contact_info_already_fetched=contact_info_already_fetched, private_contact_methods=private_contact_methods)


@f_get('/wechat-poster/<rent_ticket_id:re:[0-9a-fA-F]{24}>')
@currant_util.check_ip_and_redirect_domain
def wechat_poster(rent_ticket_id):
    title = _('快来围观')
    rent_ticket = f_app.i18n.process_i18n(f_app.ticket.output([rent_ticket_id], fuzzy_user_info=True)[0])

    report = None
    if rent_ticket["property"].get('report_id'):
        report = f_app.i18n.process_i18n(currant_data_helper.get_report(rent_ticket["property"].get('report_id')))

    # if rent_ticket["status"] not in ["draft", "to rent"]:
    # assert user and set(user["role"]) & set(["admin", "jr_admin", "operation", "jr_operation"]), abort(40300, "No access to specify status or target_rent_ticket_id")

    if rent_ticket.get('title'):
        title = _(rent_ticket.get('title', '')) + ', ' + title
    description = rent_ticket.get('description', '')

    keywords = currant_util.get_country_name_by_code(rent_ticket.get('country', {}).get('code', '')) + ',' + rent_ticket.get('city', {}).get('name', '') + ',' + ','.join(currant_util.BASE_KEYWORDS_ARRAY)
    weixin = f_app.wechat.get_jsapi_signature()

    user = f_app.i18n.process_i18n(currant_data_helper.get_user_with_custom_fields())

    if rent_ticket['status'] not in ['to rent', 'rent'] and rent_ticket.get('creator_user'):
        if not user:
            abort(401)
        elif user['id'] not in (rent_ticket.get('user', {}).get('id'), rent_ticket.get('creator_user', {}).get('id')) and not (set(user['role']) & set(['admin', 'jr_admin', 'support'])):
            abort(403)

    return currant_util.common_template("wechat_poster", rent=rent_ticket, title=title, description=description, keywords=keywords, weixin=weixin, report=report)


@f_get('/wechat-poster/<rent_ticket_id:re:[0-9a-fA-F]{24}>/image')
def wechat_poster_image(rent_ticket_id):
    from libfelix.f_html2png import html2png
    response.set_header(b"Content-Type", b"image/png")
    return html2png("://".join(request.urlparts[:2]) + "/wechat-poster/" + rent_ticket_id, width=480, height=800, url=True)


@f_get('/record-video-tips')
@currant_util.check_ip_and_redirect_domain
def record_video_tips():
    title = _('洋房东平台 - 录制/上传您的看房视频')
    return currant_util.common_template("record_video_tips", title=title)


@f_get('/admin')
@currant_util.check_ip_and_redirect_domain
def admin():
    return template("admin")


@error(401)
@f_app.user.login.check()
def error401_redirect(error, user):
    return html_redirect("/signin?error_code=40100&from=" + urllib.parse.quote(request.url))


# @f_get('/401')
# @error(401)
# def error_401(error=None):
#     title = _('没有访问该页面的权限')
#     return currant_util.common_template("401", title=title)


@f_get('/403')
@error(403)
def error_403(error=None):
    title = _('无法访问该页面')
    return currant_util.common_template("403", title=title)


@f_get('/404')
@error(404)
def error_404(error=None):
    title = _('找不到页面')
    return currant_util.common_template("404", title=title)


@f_get('/500')
@error(500)
def error_500(error=None):
    title = _('服务器开小差了')
    return currant_util.common_template("500", title=title)


@f_get("/static/<filepath:path>")
def static_route(filepath):
    return static_file(filepath, root="views/static")


@f_get("/qrcode/generate", params=dict(
    content=(str, True),
))
def qrcode_generate(params):
    img = qrcode.make(params["content"])
    output = StringIO()
    img.save(output)

    response.set_header(b"Content-Type", b"image/png")

    return output.getvalue()


@f_get("/image/fetch", params=dict(
    link=(str, True),
    news_id=str,
    property_id=str,
    content_id=str,
))
def images_proxy(params):
    def is_in_property(link, property):
        for k, v in property.get("reality_images", {}).iteritems():
            if isinstance(v, list):
                if link in v:
                    return True
        for k, v in property.get("surroundings_images", {}).iteritems():
            if isinstance(v, list):
                if link in v:
                    return True
        for k, v in property.get("effect_pictures", {}).iteritems():
            if isinstance(v, list):
                if link in v:
                    return True
        for k, v in property.get("indoor_sample_room_picture", {}).iteritems():
            if isinstance(v, list):
                if link in v:
                    return True
        for k, v in property.get("planning_map", {}).iteritems():
            if isinstance(v, list):
                if link in v:
                    return True
        for k, v in property.get("floor_plan", {}).iteritems():
            if isinstance(v, list):
                if link in v:
                    return True
        for main_house_type in property.get("main_house_types", []):
            for k, v in main_house_type.get("floor_plan", {}).iteritems():
                if link == v:
                    return True
        return False

    allowed = False
    ssl_bypass = False

    if "bbt-currant.s3.amazonaws.com" in params["link"] or "zoopla.co.uk" in params["link"] or "zoocdn.com" in params["link"]:
        params["link"] = params["link"]
        allowed = True
        ssl_bypass = True
        url_parsed = urllib.parse.urlparse(params["link"])
        url_parsed = url_parsed._replace(scheme="https")
        params["link"] = urllib.parse.urlunparse(url_parsed)

        # TODO: make this saner
        if "yangfd.com" not in request.urlparts[1] and "youngfunding.co.uk" not in request.urlparts[1] and "currant-test" not in request.urlparts[1]:
            params["link"] = params["link"].replace("bbt-currant.s3.amazonaws.com", "s3.yangfd.cn").replace("https://", "http://")

    elif "property_id" in params:
        property = f_app.property.get(params["property_id"])
        allowed = is_in_property(params["link"], property)
        if "target_property_id" in property:
            target_property = f_app.property.get(property["target_property_id"])
            allowed = is_in_property(params["link"], target_property)

    elif "news_id" in params:
        news = f_app.blog.post_get(params["news_id"])
        if params["link"] in news.get("images", []):
            allowed = True

    elif "content_id" in params:
        if params["link"] == f_app.ad.get(params["content_id"]).get("image"):
            allowed = True

    if not allowed:
        abort(40089, logger.warning("Invalid image source: not from existing property or news", exc_info=False))

    if f_app.common.use_ssl and not ssl_bypass:
        result = f_app.request(params["link"])
        if result.status_code == 200:
            response.set_header(b"Content-Type", b"image/png")
            return result.content
        else:
            abort(40000, logger.warning("Failed to fetch image source: timeout or non-exists."))
    else:
        bottle.redirect(params["link"])


@f_get('/reverse_proxy', params=dict(
    link=(str, True),
    content_type=str,
))
def reverse_proxy(params):
    result = f_app.request(params["link"])
    if result.status_code == 200:
        content = result.content
        ext = params["link"].split('.')[-1]
        if ext == "js":
            response.set_header(b"Content-Type", b"application/javascript")
        return content

    else:
        logger.debug("error in proxy %s", result)


@f_get("/logout", params=dict(
    return_url=str,
))
@currant_util.check_ip_and_redirect_domain
def logout(params):
    return_url = params.pop("return_url", "/")
    f_app.user.login.logout()
    baseurl = "://".join(request.urlparts[:2])
    redirect(baseurl + return_url)


@f_get('/upload_image')
def upload_image():
    return template("upload_image")


@f_get('/sitemap_location.xml')
def sitemap():
    xmlns = "http://www.sitemaps.org/schemas/sitemap/0.9"
    domain = request.urlparts[1]
    xhtml = "http://www.w3.org/1999/xhtml"
    root = etree.Element("{%s}urlset" % xmlns, encoding="UTF-8", nsmap={'xhtml': xhtml, None: xmlns})
    logger.debug(etree.tostring(root))
    etree.SubElement(etree.SubElement(root, "url"), "loc").text = "http://%s" % domain
    etree.SubElement(etree.SubElement(root, "url"), "loc").text = "http://%s/signup" % domain
    etree.SubElement(etree.SubElement(root, "url"), "loc").text = "http://%s/vip_sign_up" % domain
    etree.SubElement(etree.SubElement(root, "url"), "loc").text = "http://%s/signin" % domain
    etree.SubElement(etree.SubElement(root, "url"), "loc").text = "http://%s/about" % domain

    etree.SubElement(etree.SubElement(root, "url"), "loc").text = "http://%s/property_list" % domain
    # todo country需要改为code
    etree.SubElement(etree.SubElement(root, "url"), "loc").text = "http://%s/property_list?country=541bf6616b809946e81c2bd3" % domain
    etree.SubElement(etree.SubElement(root, "url"), "loc").text = "http://%s/property_list?country=541c09286b8099496db84f56" % domain
    etree.SubElement(etree.SubElement(root, "url"), "loc").text = "http://%s/property_list?country=541d32eb6b80992a1f209045" % domain
    etree.SubElement(etree.SubElement(root, "url"), "loc").text = "http://%s/property_list?country=541d334d6b80992a1f209046" % domain
    for property in f_app.property.search({"status": {"$in": ["selling", "sold out"]}}, per_page=0):
        etree.SubElement(etree.SubElement(root, "url"), "loc").text = "http://%s/property/%s" % (domain, property)
    for ticket in f_app.ticket.search({"status": {"$in": ["to rent"]}, "type": "rent"}, per_page=0):
        etree.SubElement(etree.SubElement(root, "url"), "loc").text = "http://%s/property-to-rent/%s" % (domain, ticket)

    etree.SubElement(etree.SubElement(root, "url"), "loc").text = "http://%s/news_list" % domain
    etree.SubElement(etree.SubElement(root, "url"), "loc").text = "http://%s/news_list?category=5417ecd46b80992d07638187" % domain
    etree.SubElement(etree.SubElement(root, "url"), "loc").text = "http://%s/news_list?category=5417ecf86b80992d07638188" % domain
    etree.SubElement(etree.SubElement(root, "url"), "loc").text = "http://%s/news_list?category=54180eeb6b80994dcea5600d" % domain
    for news in f_app.blog.post.search({"status": {"$ne": "deleted"}}, per_page=0):
        etree.SubElement(etree.SubElement(root, "url"), "loc").text = "http://%s/news/%s" % (domain, news)

    for child in root:
        if f_app.common.i18n_sitemap_enable_locales:
            if "?" in child[0].text:
                etree.SubElement(child, "{http://www.w3.org/1999/xhtml}link", rel="alternate", hreflang="en", href="%s&_i18n=en_GB" % child[0].text)
                etree.SubElement(child, "{http://www.w3.org/1999/xhtml}link", rel="alternate", hreflang="en", href="%s&_i18n=zh_Hant_HK" % child[0].text)
            else:
                etree.SubElement(child, "{http://www.w3.org/1999/xhtml}link", rel="alternate", hreflang="en", href="%s?_i18n=en_GB" % child[0].text)
                etree.SubElement(child, "{http://www.w3.org/1999/xhtml}link", rel="alternate", hreflang="en", href="%s?_i18n=zh_Hant_HK" % child[0].text)
        if "?" in child[0].text:
            child[0].text = "%s&_i18n=zh_Hans_CN" % child[0].text
        else:
            child[0].text = "%s?_i18n=zh_Hans_CN" % child[0].text

    response.set_header(b"Content-Type", b"application/xml")
    return etree.tostring(root, xml_declaration=True, encoding="UTF-8")


@f_get("/robots.txt")
def robots_txt():
    if "currant-test" in request.urlparts[1]:
        return("User-agent: *\n"
               "Disallow: /\n")

    return("User-agent: *\n"
           "Disallow: /static/\n"
           "Disallow: /admin/\n")


@f_get("/s3_raw/<filename>")
def s3_raw_reverse_proxy(filename):
    if filename.endswith(".jpg"):
        filename = filename[:-4]
    result = f_app.request("http://bbt-currant.s3.amazonaws.com/" + filename)
    return result.content


@f_post("/sendgrid_get_event_endpoint")
def sendgrid_get_event():
    email_event = request.json
    for single_event in email_event:
        # logger.debug(json.dumps(single_event))
        f_app.email.status.email_status_append(single_event)


@f_post("/sendcloud_get_event_endpoint")
def sendcloud_get_event():
    email_event = request.json
    for single_event in email_event:
        f_app.email.status.email_status_append(single_event)


@f_get("/wechat_endpoint", params=dict(
    signature=str,
    timestamp=str,
    nonce=str,
    echostr=str,
))
def wechat_endpoint_verifier(params):
    if params["signature"] == sha1(params["nonce"] + params["timestamp"] + f_app.common.wechat_token).hexdigest():
        return params["echostr"]
    else:
        abort(400)


@f_post("/wechat_endpoint")
def wechat_endpoint():
    orig_xml = etree.fromstring(request.body.getvalue())
    message = {}
    for element in orig_xml.iter():
        message[element.tag] = element.text

    logger.debug("Parsed wechat message:", message)

    if f_app.common.use_ssl:
        schema = "https://"
    else:
        schema = "http://"

    return_str = ""

    def common_root(func):
        def __common_root_replace_func(*args, **kwargs):
            root = etree.Element("xml")
            etree.SubElement(root, "ToUserName").text = message["FromUserName"]
            etree.SubElement(root, "FromUserName").text = message["ToUserName"]
            etree.SubElement(root, "CreateTime").text = str(calendar.timegm(datetime.utcnow().timetuple()))
            func(*args, root=root, **kwargs)
            return etree.tostring(root, encoding="UTF-8")

        return __common_root_replace_func

    @common_root
    def text_reply(text, root):
        etree.SubElement(root, "MsgType").text = etree.CDATA("text")
        etree.SubElement(root, "Content").text = etree.CDATA(text)

    @common_root
    def transfer_customer_service(root):
        etree.SubElement(root, "MsgType").text = etree.CDATA("transfer_customer_service")

    @common_root
    def build_property_list_by_country(country, root):
        if len(country) == 24:
            country = f_app.enum.get(country).get("slug")

        properties = f_app.i18n.process_i18n(f_app.property.output(f_app.property.search({
            "country.code": country,
            "status": {"$in": ["selling", "sold out"]},
            "user_generated": {"$ne": True},
        }, per_page=9, time_field="mtime")))

        etree.SubElement(root, "MsgType").text = "news"
        etree.SubElement(root, "ArticleCount").text = str(len(properties) + 1)

        articles = etree.SubElement(root, "Articles")
        for n, property in enumerate(properties):
            item = etree.SubElement(articles, "item")

            title = ""
            if "city" in property and "value" in property["city"]:
                title += property["city"]["value"] + " "
            if "name" in property:
                title += property["name"]
            if "main_house_types" in property:
                lowest_price = None
                for house_type in property["main_house_types"]:
                    if "total_price" not in house_type:
                        continue
                    if lowest_price is None or float(house_type["total_price"]["value"]) < lowest_price:
                        lowest_price = float(house_type["total_price"]["value"])
                if lowest_price is not None:
                    title += "(起投%.2f万 预期年收益%s)" % (lowest_price / 10000, property.get("annual_return_estimated", ""))
            elif "total_price" in property:
                title += "(起投%.2f万 预期年收益%s)" % (float(property["total_price"]["value"]) / 10000, property.get("annual_return_estimated", ""))

            etree.SubElement(item, "Title").text = etree.CDATA(title)

            if "description" in property:
                etree.SubElement(item, "Description").text = etree.CDATA(property["description"])

            if "reality_images" in property and len(property["reality_images"]):
                picurl = property["reality_images"][0]
                if "bbt-currant.s3.amazonaws.com" in picurl:
                    picurl += "_thumbnail.jpg"
                # from urllib import quote
                etree.SubElement(item, "PicUrl").text = picurl.replace("bbt-currant.s3.amazonaws.com/", "yangfd.com/s3_raw/")
                # etree.SubElement(item, "PicUrl").text = schema + request.urlparts[1] + "/reverse_proxy?link=" + quote(picurl)

            etree.SubElement(item, "Url").text = schema + request.urlparts[1] + "/property/" + property["id"]

        if len(properties):
            more = etree.SubElement(articles, "item")

            etree.SubElement(more, "Title").text = etree.CDATA("更多%s房产..." % (currant_util.get_country_name_by_code(property["country"]["code"]), ))
            etree.SubElement(more, "Url").text = schema + request.urlparts[1] + "/property_list?country=" + country

    if "MsgType" in message:
        if message["MsgType"] == "event":
            if message["Event"] == "CLICK":
                if message["EventKey"].startswith("property_by_country/"):
                    return_str = build_property_list_by_country(message["EventKey"][len("property_by_country/"):])

            elif message["Event"] == "subscribe":
                return_str = text_reply(
                    "感谢您关注洋房东微信号\n"
                    "我们将为你带来：\n"
                    "最实时的英国房产资讯\n"
                    "最有趣的房产新闻\n"
                    "最专业的房产解读\n"
                    "记得每天点开洋房东的消息看一看哦！"
                )

        elif message["MsgType"] == "text":
            return_str = transfer_customer_service()

            # Disabled
            if False and message["Content"] == "英国":
                # TODO: don't hardcode
                return_str = build_property_list_by_country("541c09286b8099496db84f56")

    response.set_header(b"Content-Type", b"application/xml")
    logger.debug("Responding to wechat:", return_str)
    return return_str


@f_get('/app-download')
@currant_util.check_ip_and_redirect_domain
def app_download():
    if any(pattern in request.get_header('User-Agent') for pattern in (b'iPhone', b'iPod', b'iPad')) and b"MicroMessenger" not in request.get_header('User-Agent'):
        redirect('https://itunes.apple.com/cn/app/yang-fang-dong-ying-guo-zu/id980469674')
    weixin = f_app.wechat.get_jsapi_signature()
    title = _('洋房东APP下载页')
    description = _('洋房东官方手机app客户端应用下载PC电脑端网站英国英国大不列颠英格兰苏格兰伦敦租客搜索房源租房找房住宿房东发布房源省时省力贴心安全快捷便利')
    keywords = _('洋房东,租房,出租,租房中介,找房子,短租,长租,租金,公寓,别墅,学区房,留学生租房,海外租房,英国出租,英国租房,伦敦租房,官方应用,官方app,Youngfunding,for rent,to let,room to rent,property to rent,property,apartment,house,UK property,official app, official application')

    return currant_util.common_template("app_download", title=title, weixin=weixin, description=description, keywords=keywords)


@f_get("/beta-app-download")
@currant_util.check_ip_and_redirect_domain
def beta_app_download():
    redirect('https://itunes.apple.com/cn/app/yang-fang-dong-ying-guo-zu/id980469674')


@f_get("/track", "/track.png", "/track/<ticket_id>/<property_id>/<image_type>.png", params=dict(
    ticket_id=ObjectId,
    property_id=ObjectId,
    image_type=(str, "1px"),
))
def track(params, ticket_id=None, property_id=None, image_type=None):
    """
    ``image_type`` should be ``1px`` or ``logo``.
    """

    if ticket_id is not None and ticket_id != "none":
        params["ticket_id"] = ObjectId(ticket_id)

    if property_id is not None and property_id != "none":
        params["property_id"] = ObjectId(property_id)

    if image_type is not None:
        params["image_type"] = image_type

    params["log_type"] = "share_visit"

    assert {"property_id", "ticket_id"} & set(params), abort(40000, "No track target specified")

    f_app.log.add(params)

    if params["image_type"] == "1px":
        return static_file("images/1px.png", root="views/static")
    elif params["image_type"] == "logo":
        return static_file("images/logo/logo.png", root="views/static")
    else:
        raise NotImplementedError


@f_get('/rental-available')
@currant_util.check_ip_and_redirect_domain
def rental_available():
    # issue #6990
    title = _('洋房东租房服务')
    return currant_util.common_template("rental_available", title=title)


@f_get('/how-it-works/landlord')
@currant_util.check_ip_and_redirect_domain
def how_it_works_landlord():
    title = _('洋房东房东服务')
    return currant_util.common_template("how_it_works_landlord", title=title)


@f_get('/how-it-works/tenant')
@currant_util.check_ip_and_redirect_domain
def how_it_works_tenant():
    title = _('洋房东租客服务')
    return currant_util.common_template("how_it_works_tenant", title=title)


@f_get('/how-it-works/student-house')
@currant_util.check_ip_and_redirect_domain
def how_it_works_student_house():
    title = _('洋房东学生公寓服务')
    return currant_util.common_template("how_it_works_student_house", title=title)


@f_get('/rent-intention/<rent_intention_ticket_id:re:[0-9a-fA-F]{24}>/edit')
@currant_util.check_ip_and_redirect_domain
def rent_intention_edit(rent_intention_ticket_id):

    title = _('求租意向单编辑')
    rent_intention_ticket = f_app.i18n.process_i18n(f_app.ticket.output([rent_intention_ticket_id], fuzzy_user_info=True)[0])
    user = f_app.i18n.process_i18n(currant_data_helper.get_user_with_custom_fields())
    if not user or user.get('id') != rent_intention_ticket.get('creator_user', {}).get('id'):
        redirect('/')

    return currant_util.common_template("rent_intention_edit", title=title, rent_intention_ticket=rent_intention_ticket)


@f_api('/aggregation-general', params=dict(
    date_from=(datetime, None),
    date_to=(datetime, None)
))
@f_app.user.login.check(force=True, role=['admin', 'jr_admin', 'sales', 'operation'])
def aggregation_general(user, params):
    value = {}
    with f_app.mongo() as m:
        if 'date_from' in params and 'date_to' in params:
            value.update({
                "aggregation_user_total": m.users.find({
                    "register_time": {
                        "$gte": params['date_from'],
                        "$lt": params['date_to']
                    }
                }).count()
            })
            value.update({
                "aggregation_register_user_total": m.users.find({
                    "register_time": {
                        "$gte": params['date_from'],
                        "$lt": params['date_to']
                    },
                    "status": {"$ne": "deleted"},
                }).count()
            })
            aggregate_params = [
                {"$match": {
                    "register_time": {
                        "$gte": params['date_from'],
                        "$lt": params['date_to']
                    }
                }},
                {"$unwind": "$user_type"},
                {"$group": {"_id": "$user_type", "count": {"$sum": 1}}}
            ]
        else:
            value.update({
                "aggregation_user_total": m.users.find().count()
            })
            value.update({
                "aggregation_register_user_total": m.users.find({
                    "register_time": {"$exists": True},
                    "status": {"$ne": "deleted"},
                }).count()
            })
            aggregate_params = [
                {"$unwind": "$user_type"},
                {"$group": {"_id": "$user_type", "count": {"$sum": 1}}}
            ]
        cursor = m.users.aggregate(aggregate_params)
        user_type = []
        for document in cursor:
            user_type.append({"type": f_app.enum.get(document['_id']['_id'])['value']['zh_Hans_CN'],
                              "total": document['count']})
        cursor.close()
        value.update({"aggregation_user_type": user_type})
    return value


@f_api('/aggregation-rent-ticket', params=dict(
    date_from=(datetime, None),
    date_to=(datetime, None)
))
@f_app.user.login.check(force=True, role=['admin', 'jr_admin', 'sales', 'operation'])
def aggregation_rent_ticket(user, params):

    def get_aggregation_params(params_list):
        result = []
        if 'date_from' in params and 'date_to' in params:
            result = [{
                "$match": {
                    "time": {
                        "$gte": params['date_from'],
                        "$lt": params['date_to']
                    }
                }
            }]
        result.extend(params_list)
        return result

    value = {}
    with f_app.mongo() as m:
        cursor = m.tickets.aggregate(get_aggregation_params(
            [
                {'$match': {'type': "rent"}},
                {'$group': {'_id': "$type", 'count': {'$sum': 1}}}
            ]
        ))
        if cursor.alive:
            document = cursor.next()
        else:
            document = {}
        value.update({"aggregation_rent_ticket_total": document.get('count', 0)})
        cursor.close()
        cursor = m.log.aggregate(get_aggregation_params(
            [
                {'$match': {'route': '/api/1/rent_ticket/add'}},
                {'$group': {'_id': None, 'count': {'$sum': 1}}}
            ]
        ))
        if cursor.alive:
            document = cursor.next()
        else:
            document = {}
        value.update({"aggregation_rent_ticket_create_total": document.get('count', 0)})
        cursor.close()
        cursor = m.log.aggregate(get_aggregation_params(
            [
                {'$match': {'route': '/api/1/rent_ticket/add', 'useragent': {'$regex': '.*currant.*'}}},
                {'$group': {'_id': None, 'count': {'$sum': 1}}}
            ]
        ))
        if cursor.alive:
            document = cursor.next()
        else:
            document = {}
        value.update({"aggregation_rent_ticket_create_total_from_mobile": document.get('count', 0)})
        if value['aggregation_rent_ticket_create_total'] == 0:
            aggregation_rent_ticket_create_total_from_mobile_ratio = 0
        else:
            aggregation_rent_ticket_create_total_from_mobile_ratio = value['aggregation_rent_ticket_create_total_from_mobile'] * 1.0 / value['aggregation_rent_ticket_create_total']
        value.update({"aggregation_rent_ticket_create_total_from_mobile_ratio": aggregation_rent_ticket_create_total_from_mobile_ratio})
        cursor.close()
        cursor = m.tickets.aggregate(get_aggregation_params(
            [
                {'$match': {'type': "rent"}},
                {'$group': {'_id': "$status", 'count': {'$sum': 1}}}
            ]
        ))
        status_dic = {
            'rent': '已出租',
            'to rent': '发布中',
            'draft': '草稿',
            'deleted': '已删除'
        }
        aggregation_rent_ticket_status = []
        for document in cursor:
            aggregation_rent_ticket_status.append({"status": status_dic[document['_id']],
                                                   "total": document['count']})
        cursor.close()
        value.update({"aggregation_rent_ticket_status": aggregation_rent_ticket_status})

        cursor = m.tickets.aggregate(get_aggregation_params(
            [
                {'$match': {'type': "rent"}},
                {'$group': {'_id': "$rent_type", 'count': {'$sum': 1}}}
            ]
        ))
        aggregation_rent_ticket_type = []
        for document in cursor:
            if(document['_id']):
                aggregation_rent_ticket_type.append({"type": f_app.enum.get(document['_id']['_id'])['value']['zh_Hans_CN'],
                                                     "total": document['count']})
        value.update({"aggregation_rent_ticket_type": aggregation_rent_ticket_type})
        cursor.close()
        cursor = m.tickets.aggregate(get_aggregation_params(
            [
                {'$match': {'type': "rent", 'status': "to rent"}},
                {'$group': {'_id': "$rent_type", 'count': {'$sum': 1}}}
            ]
        ))
        aggregation_rent_ticket_type_available = []
        for document in cursor:
            if(document['_id']):
                aggregation_rent_ticket_type_available.append({"type": f_app.enum.get(document['_id']['_id'])['value']['zh_Hans_CN'],
                                                               "total": document['count']})
        value.update({"aggregation_rent_ticket_type_available": aggregation_rent_ticket_type_available})
        cursor.close()
        cursor = m.tickets.aggregate(get_aggregation_params(
            [
                {'$match': {'type': "rent", 'status': "rent"}},
                {'$group': {'_id': "$rent_type", 'count': {'$sum': 1}}}
            ]
        ))
        aggregation_rent_ticket_type_rent = []
        for document in cursor:
            if(document['_id']):
                aggregation_rent_ticket_type_rent.append({"type": f_app.enum.get(document['_id']['_id'])['value']['zh_Hans_CN'],
                                                          "total": document['count']})
        cursor.close()
        value.update({"aggregation_rent_ticket_type_rent": aggregation_rent_ticket_type_rent})
        cursor = m.tickets.aggregate(get_aggregation_params(
            [
                {'$match': {'type': "rent", 'status': "to rent"}},
                {'$group': {'_id': "$landlord_type", 'count': {'$sum': 1}}}
            ]
        ))
        aggregation_landlord_type_has_available_rent_ticket = []
        for document in cursor:
            if(document['_id']):
                aggregation_landlord_type_has_available_rent_ticket.append({"type": f_app.enum.get(document['_id']['_id'])['value']['zh_Hans_CN'],
                                                                            "total": document['count']})
        cursor.close()
        value.update({"aggregation_landlord_type_has_available_rent_ticket": aggregation_landlord_type_has_available_rent_ticket})
        cursor = m.tickets.aggregate(get_aggregation_params(
            [
                {'$match': {'type': "rent", 'status': "to rent", 'rent_type._id': ObjectId('55645cf5666e3d0f57d6e284')}},
                {'$group': {'_id': "$landlord_type", 'count': {'$sum': 1}}}
            ]
        ))
        aggregation_landlord_type_has_available_whole = []

        for document in cursor:
            if(document['_id']):
                aggregation_landlord_type_has_available_whole.append({"type": f_app.enum.get(document['_id']['_id'])['value']['zh_Hans_CN'],
                                                                      "total": document['count']})
        cursor.close()
        value.update({"aggregation_landlord_type_has_available_whole": aggregation_landlord_type_has_available_whole})

        # aggregation_rent_ticket_shortest_rent_period TODO

        cursor = m.tickets.aggregate(get_aggregation_params(
            [
                {'$match': {
                    'type': "rent",
                    'status': "to rent"
                }},
                {'$group': {'_id': "$minimum_rent_period", 'count': {'$sum': 1}}}
            ]
        ))

        period_count = {
            'short': 0,
            'short_middle': 0,
            'middle_long': 0,
            'long': 0,
            'extra_long': 0
        }

        def covert_to_month(period):
            if(period['unit'] == 'week'):
                period['value'] = float(period['value']) / 4
            if(period['unit'] == 'day'):
                period['value'] = float(period['value']) / 31
            if(period['unit'] == 'year'):
                period['value'] = float(period['value']) * 12
            else:
                period['value'] = float(period['value'])
            return period

        for document in cursor:
            if(document['_id']):
                period = covert_to_month(document['_id'])
                if(period['value'] < 1.0):
                    period_count['short'] += document['count']
                if(period['value'] >= 1.0 and period['value'] < 3.0):
                    period_count['short_middle'] += document['count']
                if(period['value'] >= 3.0 and period['value'] < 6.0):
                    period_count['middle_long'] += document['count']
                if(period['value'] >= 6.0 and period['value'] < 12.0):
                    period_count['long'] += document['count']
                if(period['value'] >= 12.0):
                    period_count['extra_long'] += document['count']
        sample_text = {
            'short': 'less than 1 month',
            'short_middle': '1 month ~ 3 month',
            'middle_long': '3 month ~ 6 month',
            'long': '6 month ~ 12 month',
            'extra_long': 'longer than 12 month'
        }
        aggregation_rent_ticket_shortest_rent_period = []
        for rang in period_count:
            aggregation_rent_ticket_shortest_rent_period.append({"period": sample_text[rang],
                                                                 "total": period_count[rang]})
        value.update({"aggregation_rent_ticket_shortest_rent_period": aggregation_rent_ticket_shortest_rent_period})
        cursor.close()

    return value


@f_api('/aggregation-rent-intention-ticket', params=dict(
    date_from=(datetime, None),
    date_to=(datetime, None)
))
@f_app.user.login.check(force=True, role=['admin', 'jr_admin', 'sales', 'operation'])
def aggregation_rent_intention_ticket(user, params):

    def get_aggregation_params(params_list):
        result = []
        if 'date_from' in params and 'date_to' in params:
            result = [{
                "$match": {
                    "time": {
                        "$gte": params['date_from'],
                        "$lt": params['date_to']
                    }
                }
            }]
        result.extend(params_list)
        return result

    def get_find_params(params_dic):
        if 'date_from' in params and 'date_to' in params:
            params_dic.update({
                "time": {
                    "$gte": params['date_from'],
                    "$lt": params['date_to']
                }
            })
        return params_dic

    value = {}
    with f_app.mongo() as m:
        value.update({
            "aggregation_rent_intention_total": m.tickets.find(
                get_find_params({
                    'type': "rent_intention"
                })
            ).count()
        })
        cursor = m.tickets.aggregate(
            get_aggregation_params([
                {'$match': {'type': "rent_intention"}},
                {'$group': {'_id': '$city', 'count': {'$sum': 1}}},
                {'$sort': {'count': -1}}
            ])
        )
        aggregation_rent_intention_total_city = []
        for document in cursor:
            city = {}
            city_id = document['_id']
            if city_id is None:
                city = {}
            else:
                city_id = city_id.get('_id', None)
                if city_id is None:
                    city = {}
                else:
                    city = f_app.geonames.gazetteer.get(city_id)
            aggregation_rent_intention_total_city.append({
                "city": city.get('name', ''),
                "total": document['count']
            })
        cursor.close()
        value.update({"aggregation_rent_intention_total_city": aggregation_rent_intention_total_city})
        cursor = m.tickets.aggregate(
            get_aggregation_params(
                [
                    {'$match': {'type': "rent_intention", 'city._id': ObjectId('555966cd666e3d0f578ad2cf')}},
                    {'$group': {'_id': "$rent_type", 'count': {'$sum': 1}}}
                ]
            )
        )
        aggregation_rent_intention_type_london = []
        for document in cursor:
            if(document['_id']):
                aggregation_rent_intention_type_london.append({
                    "type": f_app.enum.get(document['_id']['_id'])['value']['zh_Hans_CN'],
                    "total": document['count']
                })
        cursor.close()
        value.update({"aggregation_rent_intention_type_london": aggregation_rent_intention_type_london})

        target_budget_currency = 'GBP'
        target_budget = {'min': 200.0}
        budget_filter = []

        for currency in f_app.common.currency:
            condition = {}
            conditions = {}
            if currency == target_budget_currency:
                if 'min' in target_budget:
                    condition['rent_budget_min.unit'] = currency
                    condition['rent_budget_min.value_float'] = {}
                    condition['rent_budget_min.value_float']['$gte'] = target_budget['min']
                if 'max' in target_budget:
                    condition['rent_budget_max.unit'] = currency
                    condition['rent_budget_max.value_float'] = {}
                    condition['rent_budget_max.value_float']['$lte'] = target_budget['max']
            else:
                if 'min' in target_budget:
                    condition['rent_budget_min.unit'] = currency
                    condition['rent_budget_min.value_float'] = {}
                    condition['rent_budget_min.value_float']['$gte'] = float(f_app.i18n.convert_currency({"unit": target_budget_currency, "value_float": target_budget['min']}, currency))
                if 'max' in target_budget:
                    condition['rent_budget_max.unit'] = currency
                    condition['rent_budget_max.value_float'] = {}
                    condition['rent_budget_max.value_float']['$lte'] = float(f_app.i18n.convert_currency({"unit": target_budget_currency, "value_float": target_budget['max']}, currency))
            conditions['$and'] = []
            conditions['$and'].append(condition)
            budget_filter.append(conditions)

        value.update({
            "aggregation_rent_intention_total_above_200": m.tickets.find(
                get_find_params({
                    'type': "rent_intention",
                    'city._id': ObjectId('555966cd666e3d0f578ad2cf'),
                    '$or': budget_filter
                })
            ).count()
        })

        value.update({
            "aggregation_rent_intentionl_has_neighborhood_total": m.tickets.find(
                get_find_params({
                    'type': "rent_intention",
                    'city._id': ObjectId('555966cd666e3d0f578ad2cf'),
                    'maponics_neighborhood': {'$exists': 'true'}
                })
            ).count()
        })
        cursor = m.tickets.aggregate(
            get_aggregation_params(
                [
                    {'$match': {'type': "rent_intention", 'city._id': ObjectId('555966cd666e3d0f578ad2cf'), 'maponics_neighborhood': {'$exists': 'true'}}},
                    {'$group': {'_id': '$maponics_neighborhood._id', 'count': {'$sum': 1}}},
                    {'$sort': {'count': -1}}
                ]
            )
        )
        aggregation_rent_intentionl_has_neighborhood = []
        for document in cursor:
            target_regions = f_app.maponics.neighborhood.get(document['_id'])
            aggregation_rent_intentionl_has_neighborhood.append({
                "neighborhood": target_regions.get("name", "") + "," + target_regions.get("parent_name", ""),
                "total": document['count']
            })
        value.update({"aggregation_rent_intentionl_has_neighborhood": aggregation_rent_intentionl_has_neighborhood})

    return value


@f_api('/aggregation-favorite', params=dict(
    date_from=(datetime, None),
    date_to=(datetime, None)
))
@f_app.user.login.check(force=True, role=['admin', 'jr_admin', 'sales', 'operation'])
def aggregation_favorite(user, params):

    def get_aggregation_params(params_list):
        result = []
        if 'date_from' in params and 'date_to' in params:
            result = [{
                "$match": {
                    "time": {
                        "$gte": params['date_from'],
                        "$lt": params['date_to']
                    }
                }
            }]
        result.extend(params_list)
        return result

    value = {}
    with f_app.mongo() as m:
        cursor = m.favorites.aggregate(
            get_aggregation_params([
                {'$group': {'_id': "$type", 'count': {'$sum': 1}}}
            ])
        )
        fav_type_dic = {
            'rent_ticket': '出租房源',
            'property': '海外房产',
            'item': '众筹'
        }
        aggregation_ticket_favorite_times_by_type = []
        for document in cursor:
            if(document['_id']):
                aggregation_ticket_favorite_times_by_type.append({
                    "type": fav_type_dic[document['_id']],
                    "total": document['count']
                })
        cursor.close()
        value.update({"aggregation_ticket_favorite_times_by_type": aggregation_ticket_favorite_times_by_type})
        cursor = m.favorites.aggregate(
            get_aggregation_params([
                {'$match': {'type': "rent_ticket"}},
                {'$group': {'_id': "$user_id", 'count': {'$sum': 1}}},
                {'$sort': {'count': -1}},
                {'$limit': 10}
            ])
        )
        aggregation_rent_ticket_favorite_times_by_user = []
        for document in cursor:
            user_name = f_app.user.output([document['_id']], custom_fields=f_app.common.user_custom_fields)
            if user_name is None or 'nickname' not in user_name[0]:
                user_name = six.text_type(document['_id'])
            else:
                user_name = user_name[0]['nickname']
            aggregation_rent_ticket_favorite_times_by_user.append({
                "user_id": document['_id'],
                "user": user_name,
                "total": document['count']
            })
        cursor.close()
        value.update({"aggregation_rent_ticket_favorite_times_by_user": aggregation_rent_ticket_favorite_times_by_user})

        cursor = m.favorites.aggregate(
            get_aggregation_params([
                {'$match': {'type': "property"}},
                {'$group': {'_id': "$user_id", 'count': {'$sum': 1}}},
                {'$sort': {'count': -1}},
                {'$limit': 10}
            ])
        )
        aggregation_property_favorite_times_by_user = []
        for document in cursor:
            user_name = f_app.user.output([document['_id']], custom_fields=f_app.common.user_custom_fields)
            if user_name is None or 'nickname' not in user_name[0]:
                user_name = six.text_type(document['_id'])
            else:
                user_name = user_name[0]['nickname']
            aggregation_property_favorite_times_by_user.append({
                "user_id": document['_id'],
                "user": user_name,
                "total": document['count']
            })
        value.update({"aggregation_property_favorite_times_by_user": aggregation_property_favorite_times_by_user})
    return value


@f_api('/aggregation-view-contact', params=dict(
    date_from=(datetime, None),
    date_to=(datetime, None)
))
@f_app.user.login.check(force=True, role=['admin', 'jr_admin', 'sales', 'operation'])
def aggregation_view_contact(user, params):

    def get_aggregation_params(params_list):
        result = []
        if 'date_from' in params and 'date_to' in params:
            result = [{
                "$match": {
                    "time": {
                        "$gte": params['date_from'],
                        "$lt": params['date_to']
                    }
                }
            }]
        result.extend(params_list)
        return result

    value = {}
    with f_app.mongo() as m:
        cursor = m.orders.aggregate(
            get_aggregation_params([
                {'$unwind': "$items"},
                {'$group': {'_id': "$user.nickname", 'count': {'$sum': 1}}},
                {'$group': {'_id': None, 'totalUsersCount': {'$sum': 1}, 'totalRequestCount': {'$sum': "$count"}}}
            ])
        )
        if cursor.alive:
            document = cursor.next()
        else:
            document = {}
        value.update({"aggregation_view_contact_user_total": document.get('totalUsersCount', 0)})
        value.update({"aggregation_view_contact_times": document.get('totalRequestCount', 0)})
        aggregation_view_contact_detail = []
        for i in range(5):
            cursor = m.orders.aggregate(
                get_aggregation_params([
                    {'$unwind': "$items"},
                    {'$group': {'_id': "$user.nickname", 'count': {'$sum': 1}}},
                    {'$match': {'count': i}},
                    {'$group': {'_id': 'null', 'totalUsersCount': {'$sum': 1}, 'totalRequestCount': {'$sum': "$count"}}}
                ])
            )
            if cursor.alive:
                document = cursor.next()
                aggregation_view_contact_detail.append({
                    "times": i,
                    "user_total": document['totalUsersCount'],
                    "view_times": document['totalRequestCount'],
                    "ratio": document['totalUsersCount'] * 1.0 / value['aggregation_view_contact_user_total']
                })
        value.update({"aggregation_view_contact_detail": aggregation_view_contact_detail})
        cursor.close()
        cursor = m.orders.aggregate(
            get_aggregation_params([
                {'$unwind': "$items"},
                {
                    '$group':
                    {
                        '_id': "$user.nickname",
                        'count': {'$sum': 1},
                        'user_id': {'$first': "$user.id"}
                    }
                },
                {'$sort': {'count': -1}}
            ])
        )
        aggregation_view_contact_by_user = []
        for document in cursor:
            aggregation_view_contact_by_user.append({
                "user_id": document['user_id'],
                "user": document['_id'],
                "total": document['count']
            })
        value.update({"aggregation_view_contact_by_user": aggregation_view_contact_by_user})
        cursor.close()
        cursor = m.orders.aggregate(
            get_aggregation_params([
                {'$unwind': "$items"},
                {'$group': {'_id': "$ticket_id", 'count': {'$sum': 1}}},
                {'$group': {'_id': None, 'totalUsersCount': {'$sum': 1}, 'totalRequestCount': {'$sum': "$count"}}}
            ])
        )
        if cursor.alive:
            document = cursor.next()
        else:
            document = {}
        value.update({
            "aggregation_view_contact_ticket_total": document.get('totalUsersCount', 0),
            "aggregation_view_contact_total": document.get('totalRequestCount', 0)
        })
        cursor.close()
        cursor = m.orders.aggregate(
            get_aggregation_params([
                {'$unwind': "$items"},
                {'$group': {'_id': "$ticket_id", 'count': {'$sum': 1}}},
                {'$sort': {'count': -1}},
                {'$limit': 10}
            ])
        )
        aggregation_view_contact_times_sort = []
        for document in cursor:
            try:
                ticket = f_app.ticket.get(document['_id'])
            except:
                ticket = None
            if ticket is None:
                continue
            aggregation_view_contact_times_sort.append({
                "name": ticket.get("title", ''),
                "url_id": document['_id'],
                "total": document['count']
            })
        value.update({"aggregation_view_contact_times_sort": aggregation_view_contact_times_sort})
    return value


@f_api('/aggregation-property-view', params=dict(
    date_from=(datetime, None),
    date_to=(datetime, None)
))
@f_app.user.login.check(force=True, role=['admin', 'jr_admin', 'sales', 'operation'])
def aggregation_property_view(user, params):

    def get_aggregation_params(params_list):
        result = []
        if 'date_from' in params and 'date_to' in params:
            result = [{
                "$match": {
                    "time": {
                        "$gte": params['date_from'],
                        "$lt": params['date_to']
                    }
                }
            }]
        result.extend(params_list)
        return result

    def get_find_params(params_dic):
        if 'date_from' in params and 'date_to' in params:
            params_dic.update({
                "time": {
                    "$gte": params['date_from'],
                    "$lt": params['date_to']
                }
            })
        return params_dic

    value = {}
    with f_app.mongo() as m:
        value.update({
            "aggregation_property_view_total": m.log.find(
                get_find_params({
                    'property_id': {
                        '$exists': True
                    }
                })
            ).count()
        })
        value.update({
            "aggregation_property_view_register_user": m.log.find(
                get_find_params({
                    'property_id': {
                        '$exists': True
                    },
                    'id': {'$ne': None}
                })
            ).count()
        })

        func_map = Code('''
            function() {
                emit(this.id, 1);
            }
        ''')
        func_reduce = Code('''
            function(key, value) {
                return Array.sum(value)
            }
        ''')
        result = m.log.map_reduce(
            func_map,
            func_reduce,
            "aggregation_property_view_by_user",
            query=get_find_params({
                'property_id': {
                    '$exists': True
                },
                'id': {
                    '$ne': None
                }
            })
        )
        aggregation_property_view_times_by_user_sort = []
        for single in result.find().sort('value', -1):
            if 'nickname' in f_app.user.get(single['_id']):
                user_name = f_app.user.get(single['_id'])['nickname']
            else:
                user_name = single['_id']
            aggregation_property_view_times_by_user_sort.append({
                "user_id": single['_id'],
                "user": user_name,
                "total": single['value']
            })
        value.update({"aggregation_property_view_times_by_user_sort": aggregation_property_view_times_by_user_sort})
        func_map = Code('''
            function() {
                emit(this.property_id, 1);
            }
        ''')
        result = m.log.map_reduce(
            func_map,
            func_reduce,
            "aggregation_property_view_by_user",
            query=get_find_params({
                'property_id': {
                    '$exists': True
                }
            })
        )
        aggregation_property_view_times_by_property_sort = []
        for single in result.find().sort('value', -1):
            property_id = single["_id"]
            property_domain = f_app.property.output([property_id], permission_check=False, ignore_nonexist=True)
            name = ''
            if property_domain is None:
                continue
            if property_domain[0]:
                name = f_app.i18n.process_i18n(property_domain[0].get("name"))
            if property_id:
                if name is None:
                    name = ''
                aggregation_property_view_times_by_property_sort.append({
                    "title": name,
                    "url_id": property_id,
                    "total": single['value']
                })
        value.update({"aggregation_property_view_times_by_property_sort": aggregation_property_view_times_by_property_sort})
    return value


@f_api('/aggregation-rent-request', params=dict(
    date_from=(datetime, None),
    date_to=(datetime, None)
))
@f_app.user.login.check(force=True, role=['admin', 'jr_admin', 'sales', 'operation'])
def aggregation_rent_request(user, params):

    def get_aggregation_params(params_list):
        result = []
        if 'date_from' in params and 'date_to' in params:
            result = [{
                "$match": {
                    "time": {
                        "$gte": params['date_from'],
                        "$lt": params['date_to']
                    }
                }
            }]
        result.extend(params_list)
        return result

    def get_find_params(params_dic):
        if 'date_from' in params and 'date_to' in params:
            params_dic.update({
                "time": {
                    "$gte": params['date_from'],
                    "$lt": params['date_to']
                }
            })
        return params_dic

    value = {}
    with f_app.mongo() as m:

        value.update({
            'aggregation_rent_request_total_count': m.tickets.find(
                get_find_params({
                    "type": "rent_intention",
                    "interested_rent_tickets": {"$exists": True},
                    "status": {
                        "$in": [
                            "requested", "assigned", "in_progress", "rejected", "confirmed_video", "booked", "holding_deposit_paid", "checked_in"
                        ]
                    }
                })
            ).count()
        })

        func_map = Code('''
            function() {
                this.interested_rent_tickets.forEach(
                    function(e) {
                        emit(e, 1)
                    }
                )
            }
        ''')
        func_reduce = Code('''
            function(key, value) {
                return Array.sum(value)
            }
        ''')
        result = m.tickets.map_reduce(
            func_map,
            func_reduce,
            "aggregation_rent_request",
            query=get_find_params({
                "type": "rent_intention",
                "interested_rent_tickets": {"$exists": True},
                "status": {
                    "$in": [
                        "requested", "assigned", "in_progress", "rejected", "confirmed_video", "booked", "holding_deposit_paid", "checked_in"
                    ]
                }
            })
        )
        aggregation_rent_request_count_sort = []
        for single in result.find().sort('value', -1):
            rent_ticket_id = single['_id']
            count = single['value']
            try:
                target_ticket = f_app.ticket.get(rent_ticket_id)
            except:
                pass
            else:
                aggregation_rent_request_count_sort.append({
                    'ticket_id': rent_ticket_id,
                    'title': target_ticket.get('title', ''),
                    'count': count
                })
        value.update({'aggregation_rent_request_count_sort': aggregation_rent_request_count_sort})

        cursor = m.users.aggregate(
            [
                {"$unwind": "$user_type"},
                {"$group": {"_id": "$user_type", "count": {"$sum": 1}}}
            ]
        )
        user_type = []
        for document in cursor:
            user_type.append({"type": f_app.enum.get(document['_id']['_id'])['value']['zh_Hans_CN'],
                              "total": document['count']})
        cursor.close()
        for single in user_type:
            if single['type'] == '租客':
                percent_base = single['total']
        result = m.tickets.find(
            get_find_params({
                "type": "rent_intention",
                "interested_rent_tickets": {"$exists": True},
                "status": {
                    "$in": [
                        "requested", "assigned", "in_progress", "rejected", "confirmed_video", "booked", "holding_deposit_paid", "checked_in"
                    ]
                }
            })
        )
        user_set = set()
        get_request_period = []
        for single in result:
            user_set.add(single['user_id'])
            for ticket_id in single['interested_rent_tickets']:
                time_end = single.get('time', None)
                try:
                    time_start = f_app.ticket.get(ticket_id).get('time', None)
                except:
                    time_start = None
                if time_end is not None and time_start is not None:
                    get_request_period.append(time_end - time_start)
        value.update({"aggregation_rent_request_user_do_request_total": len(user_set)})
        if percent_base == 0:
            aggregation_rent_request_user_ratio = 0
        else:
            aggregation_rent_request_user_ratio = len(user_set) * 1.0 / percent_base
        value.update({'aggregation_rent_request_user_ratio': aggregation_rent_request_user_ratio})
        if len(get_request_period) == 0:
            average = "无"
        else:
            average = sum(get_request_period, timedelta()) / len(get_request_period)
        value.update({'aggregation_rent_request_average_period': average})

        result = m.tickets.find(
            get_find_params({
                "type": "rent_intention",
                "interested_rent_tickets": {"$exists": True},
                "status": {
                    "$in": [
                        "requested", "assigned", "in_progress", "rejected", "confirmed_video", "booked", "holding_deposit_paid", "checked_in"
                    ]
                }
            })
        )

        request_count_before_rent_list = []
        relocate_different_city_count = 0
        user_have_request_set = set()
        user_different_city_set = set()
        period_count = {
            "longer than 12 months": {
                'single': 0,
                'whole': 0
            },
            "6 ~ 12 months": {
                'single': 0,
                'whole': 0
            },
            "3 ~ 6 months": {
                'single': 0,
                'whole': 0
            },
            "1 ~ 3 months": {
                'single': 0,
                'whole': 0
            },
            "less than 1 month": {
                'single': 0,
                'whole': 0
            }
        }
        for index, ticket in enumerate(result):
            if 'rent_available_time' in ticket and 'rent_deadline_time' in ticket:
                try:
                    period = ticket['rent_deadline_time'] - ticket['rent_available_time']
                    rent_type = f_app.enum.get(f_app.ticket.get(ticket['interested_rent_tickets'][0])['rent_type']['id'])['slug'].split(':')[-1]
                except:
                    pass
                else:
                    if period.days >= 365:
                        period_count["longer than 12 months"][rent_type] += 1
                    elif 365 > period.days >= 180:
                        period_count["6 ~ 12 months"][rent_type] += 1
                    elif 180 > period.days >= 90:
                        period_count["3 ~ 6 months"][rent_type] += 1
                    elif 90 > period.days >= 30:
                        period_count["1 ~ 3 months"][rent_type] += 1
                    elif period.days <= 30:
                        period_count["less than 1 month"][rent_type] += 1
            if 'user_id' in ticket:
                user_have_request_set.add(ticket['user_id'])
            if 'custom_fields' in ticket:
                relocate_city = None
                want_rent_city = None
                for single in ticket['custom_fields']:
                    if single.get('key', '') == 'relocate':
                        relocate_city = single.get('value', '').split(',')[-1].lstrip(' ')
                try:
                    interested_ticket_id = ticket['interested_rent_tickets'][0]
                    interested_ticket = f_app.ticket.get(interested_ticket_id)
                    want_rent_city = f_app.i18n.process_i18n({'city': f_app.property.get(interested_ticket['property_id']).get('city', {})})['city'].get('name', None)
                except:
                    continue
                if relocate_city is not None and want_rent_city is not None:
                    if relocate_city != want_rent_city:
                        if 'user_id' in ticket:
                            user_different_city_set.add(ticket['user_id'])
                        relocate_different_city_count += 1
            if ticket['status'] == 'rent':
                for single in aggregation_rent_request_count_sort:
                    if single['ticket_id'] in ticket['interested_rent_tickets']:
                        request_count_before_rent_list.append(single['count'])
        value.update({"aggregation_rent_request_average_count_before_rent": sum(request_count_before_rent_list) * 1.0 / len(request_count_before_rent_list) if len(request_count_before_rent_list) else 0})
        value.update({'aggregation_rent_request_user_rent_different_city_ratio': float(1.0 * len(user_different_city_set) / len(user_have_request_set) if len(user_have_request_set) else 0)})
        value.update({
            'aggregation_rent_request_period_count':
            [
                {
                    "period": single,
                    "detail": [
                        {
                            "type": type,
                            "count": period_count[single][type]
                        }
                        for type in period_count[single]
                    ]
                }
                for single in period_count
            ]
        })
    return value


@f_api('/aggregation-email-detail', params=dict(
    date_from=(datetime, None),
    date_to=(datetime, None)
))
@f_app.user.login.check(force=True, role=['admin', 'jr_admin', 'sales', 'operation'])
def aggregation_email_detail(user, params):

    def get_aggregation_params(params_list):
        result = []
        if 'date_from' in params and 'date_to' in params:
            result = [{
                "$match": {
                    "time": {
                        "$gte": params['date_from'],
                        "$lt": params['date_to']
                    }
                }
            }]
        result.extend(params_list)
        return result

    def get_find_params(params_dic, time="time"):
        if 'date_from' in params and 'date_to' in params:
            params_dic.update({
                time: {
                    "$gte": params['date_from'],
                    "$lt": params['date_to']
                }
            })
        return params_dic

    value = {}
    with f_app.mongo() as m:
        func_map = Code('''
            function() {
            var list = []
                if (this.tag && this.email_id) {
                    if (Array.isArray(this.target)) {
                        for (var index = 0; index < this.target.length; index ++) {
                            list = []
                            list.push({target: this.target[index],
                                       email_id: this.email_id});
                            emit(this.tag, {a:list});
                        }
                    }
                    else {
                        list.push({target: this.target,
                                   email_id: this.email_id});
                        emit(this.tag, {a:list});
                    }
                }
            }
        ''')
        func_reduce = Code('''
            function(key, values) {
                var list = []
                values.forEach(function(e) {
                    if (e.a) {
                        list = list.concat(e.a)
                    }
                    else {
                        list = list.concat(e)
                    }
                });
                return {a:list}
            }
        ''')
        # specify_date = datetime(2015, 11, 18)
        '''if 'time' in params:
            result = f_app.task.get_database(m).map_reduce(func_map, func_reduce, "aggregation_tag", query={"type": "email_send", "start": {"$gte": params['time']}})
        else:
            result = f_app.task.get_database(m).map_reduce(func_map, func_reduce, "aggregation_tag", query={"type": "email_send"})'''
        result = f_app.task.get_database(m).map_reduce(
            func_map,
            func_reduce,
            "aggregation_tag",
            query=get_find_params({"type": "email_send"}, "start")
        )
        value.update({"aggregation_email_tag_total": result.find().count()})
        # total_email_drop = 0
        # total_email_contain_new_only = 0
        aggregation_email_tag_detail = []
        for tag in result.find():
            # we only aggregation hava processed mark or dropped mark email into total_email
            func_status_map = Code('''
                function() {
                    var event = this.email_status_set;
                    var event_detail = this.email_status_detail;
                    if (event_detail && event) {
                        if (event.indexOf("processed") != -1 || event.indexOf("dropped") != -1) {
                            emit("total_email", 1);
                        }
                        if (event.indexOf("dropped") != -1) {
                            emit("total_email_drop", 1);
                            emit("total_email_drop_id", {email_id:[this.email_id]});
                        }
                        if (event.length == 1 && event.indexOf("new") != -1) {
                            emit("total_email_contain_new_only", 1);
                            emit("total_email_contain_new_only_id", {email_id:[this.email_id]});
                        }
                        event.forEach(function(e) {
                            emit(e, 1);
                            if (event_detail) {
                                event_detail.forEach(function(c) {
                                    if (c.event == e) {
                                        emit(e+" (repeat)", 1);
                                    }
                                });
                            }
                        });
                    }
                }
            ''')
            func_status_reduce = Code('''
                function(key, value) {
                    if (key == 'total_email_drop_id' || key == 'total_email_contain_new_only_id') {
                        value_list = [];
                        value.forEach(function(e) {
                            value_list = value_list.concat(e.email_id);
                        });
                        return {email_id:value_list}
                    }
                    else {
                        return Array.sum(value)
                    }
                }
            ''')
            query_param = {}
            or_param = []
            for single_param in tag["value"]["a"]:
                or_param.append(single_param)
            query_param.update({"$or": or_param})
            tag_result = f_app.email.status.get_database(m).map_reduce(
                func_status_map,
                func_status_reduce,
                "aggregation_tag_event",
                query=query_param
            )
            final_result = {}
            for thing in tag_result.find():
                final_result.update({thing["_id"]: thing["value"]})
            open_unique = final_result.get("open", 0)
            open_times = final_result.get("open (repeat)", 0)
            click_unique = final_result.get("click", 0)
            click_times = final_result.get("click (repeat)", 0)
            delivered_times = final_result.get("delivered", 0)
            total_email = final_result.get("total_email", 0)
            # total_email_drop += final_result.get("total_email_drop", 0)
            # total_email_contain_new_only += final_result.get("total_email_contain_new_only", 0)
            # total_email_drop_id = final_result.get("total_email_drop_id", {}).get("email_id", [])
            # total_email_contain_new_only_id = final_result.get("total_email_contain_new_only_id", {}).get("email_id", [])
            single_value = {
                "tag": tag['_id'],
                "total": total_email,
                "delivered": delivered_times,
                "delivered_ratio": delivered_times / total_email if total_email else 0,
                "open": open_unique,
                "open_ratio": open_unique / total_email if total_email else 0,
                "open_repeat": open_times,
                "click": click_unique,
                "click_ratio": click_unique / total_email if total_email else 0,
                "click_repeat": click_times
            }
            '''if 'time' in params:
                single_value.update({
                    "total_email_drop_id": total_email_drop_id,
                    "total_email_contain_new_only_id": total_email_contain_new_only_id
                })'''
            aggregation_email_tag_detail.append(single_value)
        value.update({"aggregation_email_tag_detail": aggregation_email_tag_detail})
        '''if 'time' in params:
            value.update({"aggregation_email_contain_new_only": total_email_contain_new_only})
            value.update({"aggregation_email_drop": total_email_drop})'''

    return value


@f_api('/update-user-analyze')
@f_app.user.login.check(force=True, role=['admin', 'jr_admin', 'sales', 'operation'])
def user_analyze_update(user):
    for user_id in f_app.user.get_active():
        f_app.user.analyze.data_update(user_id)


@f_get('/export-excel/rent-ticket-analyze.xlsx', params=dict(
    date_from=(datetime, None),
    date_to=(datetime, None)
))
@f_app.user.login.check(force=True, role=['admin', 'jr_admin', 'sales', 'operation'])
def rent_ticket_analyze(user, params):

    def get_gender(user, ticket):
        trans = {
            'male': '男',
            'female': '女'
        }
        if 'gender' in user:
            return trans.get(user['gender'], '')
        elif 'gender' in ticket:
            return trans.get(ticket['gender'], '')
        else:
            with f_app.mongo() as m:
                ticket_loc = m.tickets.find_one({
                    "user_id": ObjectId(user['id']),
                    "gender": {"$exists": True}
                })
                if ticket_loc is not None:
                    return trans.get(ticket_loc['gender'], '')
            return ''

    def get_user_age(user, ticket):
        birth_date = None
        if 'date_of_birth' in user:
            birth_date = user['date_of_birth']
        elif 'date_of_birth' in ticket:
            birth_date = ticket['date_of_birth']
        else:
            with f_app.mongo() as m:
                ticket_loc = m.tickets.find_one({
                    "user_id": ObjectId(user['id']),
                    "date_of_birth": {"$exists": True}
                })
                if ticket_loc is not None:
                    birth_date = ticket_loc['date_of_birth']
        if birth_date is None:
            return ''
        else:
            birth_date = birth_date.date()
            today = date.today()
            age = today.year - birth_date.year
            today = today.replace(year=birth_date.year)
            if today < birth_date and age:
                age = age - 1
            return age

    def get_occupation(user, ticket):
        occupation_id = None
        if 'occupation' in user:
            occupation_id = user['occupation']['id']
        elif 'occupation' in ticket:
            occupation_id = ticket['occupation']['id']
        else:
            with f_app.mongo() as m:
                ticket_loc = m.tickets.find_one({
                    "user_id": ObjectId(user['id']),
                    "occupation": {"$exists": True}
                })
                if ticket_loc is not None:
                    occupation_id = ticket_loc['occupation']['_id']
        if occupation_id is None:
            return ''
        else:
            return f_app.enum.get(occupation_id)['value']['zh_Hans_CN']

    def get_detail_address(ticket):
        ticket = f_app.i18n.process_i18n(ticket)
        if f_app.util.batch_iterable(ticket.get("maponics_neighborhood", {})):
            maponics_neighborhood = ticket.get("maponics_neighborhood", {})[0]
        else:
            maponics_neighborhood = ticket.get("maponics_neighborhood", {})
        return {
            'whole': ' '.join([
                ticket.get("country", {}).get("code", ''),
                ticket.get("city", {}).get("name", ''),
                maponics_neighborhood.get("name", ''),
                ticket.get("address", ''),
                ticket.get("zipcode_index", '')
            ]),
            'country': ticket.get("country", {}).get("code", ''),
            'city': ticket.get("city", {}).get("name", ''),
            'neighborhood': maponics_neighborhood.get("name", '')
        }

    def get_user_landlord_type(ticket):
        if ticket is None:
            return ''
        landlord_type = ticket.get('landlord_type', None)
        if landlord_type is None:
            return ''
        return f_app.enum.get(landlord_type['id'])['value']['zh_Hans_CN']

    def get_to_rent_local(ticket):
        if 'property_id' not in ticket:
            return {}
        try:
            property_value = f_app.property.get(ticket['property_id'])
        except:
            return {}
        return get_detail_address(property_value)

    def get_request_ticket_total(ticket):
        with f_app.mongo() as m:
            total = m.tickets.find({
                "type": "rent_intention",
                "interested_rent_tickets": ObjectId(ticket['id']),
                "status": {
                    "$in": [
                        "requested", "assigned", "in_progress", "rejected", "confirmed_video", "booked", "holding_deposit_paid", "checked_in"
                    ]
                }
            }).count()
        return total

    def get_own_house_viewed_time(ticket):
        with f_app.mongo() as m:
            total = m.log.find({
                "type": "route",
                "route": "/property-to-rent/" + six.text_type(ticket['id'])
            }).count()
        return total

    def get_own_house_favorite_time(ticket):
        with f_app.mongo() as m:
            total = m.favorites.find({
                "status": {"$ne": "deleted"},
                "type": "rent_ticket",
                "ticket_id": ObjectId(ticket['id'])
            }).count()
        return total

    def time_period_label(ticket):
        if ticket is None:
            return ''
        time = ""
        period_start = ticket.get("rent_available_time", None)
        period_end = ticket.get("rent_deadline_time", None)
        if period_end is None or period_start is None:
            time = "不明"
        else:
            period = period_end - period_start
            if period.days >= 365:
                time = "longer than 12 months"
            elif 365 > period.days >= 180:
                time = "6 ~ 12 months"
            elif 180 > period.days >= 90:
                time = "3 ~ 6 months"
            elif 90 > period.days >= 30:
                time = "1 ~ 3 months"
            elif period.days <= 30:
                time = "less than 1 month"
        return time

    def get_own_house_share_time(ticket):
        with f_app.mongo() as m:
            total = 0
            total = m.log.find({
                "type": "route",
                "route": '/wechat-poster/' + ticket['id']
            }).count()
        return total

    def get_rent_type(ticket):
        if 'rent_type' not in ticket:
            return ''
        return f_app.enum.get(ticket['rent_type']['id'])['value']['zh_Hans_CN']

    def prepare_data(value):
        if f_app.util.batch_iterable(value):
            value_list = []
            for single_value in value:
                value_list.append(prepare_data(single_value))
            return value_list
        elif isinstance(value, datetime):
            loc_t = timezone('Europe/London')
            loc_dt = loc_t.localize(value)
            return six.text_type(loc_dt.strftime('%Y-%m-%d %H:%M:%S'))
        elif value is not None:
            return six.text_type(value)
        else:
            return ''

    def get_correct_col_index(num):
        if num > 26 * 26:
            return "ZZ"
        if num >= 26:
            return get_correct_col_index(num / 26 - 1) + get_correct_col_index(num % 26)
        else:
            return chr(num + 65)

    def format_fit(sheet):
        simsun_font = Font(name="SimSun")
        alignment_fit = Alignment(shrink_to_fit=True)
        for row in sheet.rows:
            for cell in row:
                cell.font = simsun_font
                cell.alignment = alignment_fit
        for num, col in enumerate(sheet.columns):
            lenmax = 0
            for cell in col:
                lencur = 0
                if cell.value is None:
                    cell.value = ''
                if isinstance(cell.value, int) or isinstance(cell.value, datetime):
                    lencur = len(six.text_type(cell.value).encode("GBK"))
                elif cell.value is not None:
                    lencur = len(cell.value.encode("GBK", "replace"))
                if lencur > lenmax:
                    lenmax = lencur
            sheet.column_dimensions[get_correct_col_index(num)].width = lenmax * 0.86
            print("col", get_correct_col_index(num), "fit.")

    def get_user_type(user):
        user_type_list = []
        if 'user_type' in user:
            if f_app.util.batch_iterable(user['user_type']):
                for user_type in user['user_type']:
                    user_type_list.append(f_app.enum.get(user_type['id'])['value']['zh_Hans_CN'])
                return '/'.join(user_type_list)
            else:
                user_type = user['user_type']
                return f_app.enum.get(user_type['id'])['value']['zh_Hans_CN']
        else:
            return ''

    def check_download(user):
        credit = f_app.user.credit.get("view_rent_ticket_contact_info", user['id']).get("credits", [])
        for single in credit:
            if single.get("tag", None) == "download_ios_app":
                return '已下载'
        return '未下载'

    header = [
        '房东用户名', '注册时间', '性别', '年龄', '职业', '电话', '邮箱', '国家', '城市', '用户类型',
        'app下载',
        '房东类型',
        '房源收到的咨询单数量',
        '房源被查看的次数',
        '分享房产',
        '查看与分享次数合计',
        '房源被收藏次数',
        '发布时间',
        '国家', '城市', '街区', '单间还是整套', '短租长租', '租金', '已发布时间(天)', '提前多久开始出租(天)',
    ]

    wb = Workbook()
    ws = wb.active

    ws.append(header)

    date_from = datetime(2015, 12, 7)
    tickets = f_app.ticket.get(f_app.ticket.search({
        "type": "rent",
        "status": {"$nin": ["deleted", "draft"]},
        "time": {
            "$gte": date_from
        }
    }, per_page=-1, notime=True))
    for index, rent_ticket in enumerate(tickets):
        print(index)
        # if index > 15:
        #     break
        if 'user_id' not in rent_ticket or 'time' not in rent_ticket:
            continue
        if rent_ticket['time'] < date_from:
            continue
        # if rent_ticket['time'] < params.date_from:
        #     continue
        user = f_app.user.get(rent_ticket['user_id'])

        wait_rent_priod = ''
        wait_to_rent_priod = ''
        location = get_to_rent_local(rent_ticket)
        if 'rent_available_time' in rent_ticket and 'time' in rent_ticket:
            wait_to_rent_priod = rent_ticket['rent_available_time'] - rent_ticket['time']
            wait_to_rent_priod = wait_to_rent_priod.days
        if 'time' in rent_ticket:
            wait_rent_priod = datetime.utcnow() - rent_ticket['time']
            wait_rent_priod = wait_rent_priod.days
        view_times = get_own_house_viewed_time(rent_ticket)
        share_times = get_own_house_share_time(rent_ticket)
        ws.append(prepare_data([
            user.get("nickname", ''),  # 用户名
            user.get("register_time", ''),  # 注册时间注册时间
            get_gender(user, rent_ticket),  # gender
            get_user_age(user, rent_ticket),  # age
            get_occupation(user, rent_ticket),  # 职业
            user.get('phone', ''),  # phone
            user.get('email', ''),  # email
            user.get("country", {}).get('code', ''),  # 国家
            '',  # city
            get_user_type(user),
            # f_app.enum.get(user.get('user_type', {}).get('id', ''))['value']['zh_Hans_CN'],  # 用户类型
            check_download(user),
            # user.get("analyze_guest_downloaded", ''),  # app下载
            get_user_landlord_type(rent_ticket),  # 房东类型
            get_request_ticket_total(rent_ticket),  # 房东收到的咨询单数量
            view_times,  # 房源被查看的次数
            share_times,  # 分享房产
            view_times + share_times,  # 查看与分享次数合计
            get_own_house_favorite_time(rent_ticket),  # 房源被收藏次数
            rent_ticket['time'],  # 发布时间
            location.get('country', ''),  # 地区 国家
            location.get('city', ''),  # 地区 城市
            location.get('neighborhood', ''),  # 地区 街区
            get_rent_type(rent_ticket),  # 单间还是整套
            time_period_label(rent_ticket),  # 短租长租
            rent_ticket.get("price", {}).get('value', ''),  # 租金
            wait_rent_priod,  # 已发布时间
            wait_to_rent_priod,  # 提前多久开始出租(天)
        ]))
    format_fit(ws)
    out = StringIO(save_virtual_workbook(wb))
    response.set_header(b"Content-Type", b"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    return out


@f_get('/export-excel/user-analyze.xlsx')
@f_app.user.login.check(force=True, role=['admin', 'jr_admin', 'sales', 'operation'])
def user_analyze(user):
    def prepare_data(value):
        if f_app.util.batch_iterable(value):
            value_list = []
            for single_value in value:
                value_list.append(prepare_data(single_value))
            return value_list
        elif isinstance(value, datetime):
            loc_t = timezone('Europe/London')
            loc_dt = loc_t.localize(value)
            return six.text_type(loc_dt.strftime('%Y-%m-%d %H:%M:%S %Z%z'))
        elif value is not None:
            return six.text_type(value)
        else:
            return ''

    def get_correct_col_index(num):
        if num > 26 * 26:
            return "ZZ"
        if num >= 26:
            return get_correct_col_index(num / 26 - 1) + get_correct_col_index(num % 26)
        else:
            return chr(num + 65)

    def format_fit(sheet):
        simsun_font = Font(name="SimSun")
        alignment_fit = Alignment(shrink_to_fit=True)
        for row in sheet.rows:
            for cell in row:
                cell.font = simsun_font
                cell.alignment = alignment_fit
        for num, col in enumerate(sheet.columns):
            lenmax = 0
            for cell in col:
                lencur = 0
                if cell.value is None:
                    cell.value = ''
                if isinstance(cell.value, int) or isinstance(cell.value, datetime):
                    lencur = len(six.text_type(cell.value).encode("GBK"))
                elif cell.value is not None:
                    lencur = len(cell.value.encode("GBK", "replace"))
                if lencur > lenmax:
                    lenmax = lencur
            sheet.column_dimensions[get_correct_col_index(num)].width = lenmax * 0.86
            print("col", get_correct_col_index(num), "fit.")

    def get_diff_color(total):
        fill = []
        base_color = 0x999999
        color = 0x0
        if 4 <= total <= 6:
            s = 0x111111
        elif total == 3:
            s = 0x222222
        elif total == 2:
            s = 0x333333
        else:
            return
        for index in range(total):
            color = base_color + s * index
            color_t = '00' + "%x" % color
            fill.append(PatternFill(fill_type='solid', start_color=color_t, end_color=color_t))
        return fill

    def be_colorful(sheet, max_segment):
        header_fill = PatternFill(fill_type='solid', start_color='00dddddd', end_color='00dddddd')
        for cell in sheet.rows[0]:
            cell.fill = header_fill
        for col in sheet.columns:
            col_set = set()
            cell_fill = []
            col_list = []
            for num, cell in enumerate(col):
                if num and len(cell.value):
                    col_set.add(cell.value)
                if len(col_set) > max_segment:
                    break
            if max_segment >= len(col_set) > 1:
                cell_fill = get_diff_color(len(col_set))
                col_list = list(col_set)
                for number, cell in enumerate(col):
                    if not number:
                        continue
                    for index, value in enumerate(col_list):
                        if cell.value == value:
                            cell.fill = cell_fill[index]

    def get_request_total(user):
        if user is None:
            return ''
        with f_app.mongo() as m:
            return m.tickets.find({
                "user_id": ObjectId(user['id']),
                "type": "rent_intention",
                "interested_rent_tickets": {"$exists": True},
                "status": {
                    "$in": [
                        "requested", "assigned", "in_progress", "rejected", "confirmed_video", "booked", "holding_deposit_paid", "checked_in"
                    ]
                }
            }).count()

    def get_recent_ticket_to_rent(user):
        if user is None:
            return ''
        result = f_app.ticket.get(f_app.ticket.search({
            "user_id": ObjectId(user['id']),
            "type": "rent",
            "status": {
                "$nin": ['deleted', 'draft']
            }
        }, per_page=-1))
        if result is None:
            return {}
        if len(result) == 0:
            return {}
        time_max = result[0]['time']
        for single in result:
            time_max = max(time_max, single['time'])
        for single in result:
            if single['time'] == time_max:
                return single

    def get_recent_ticket_request(user):
        if user is None:
            return ''
        result = f_app.ticket.get(f_app.ticket.search({
            "user_id": ObjectId(user['id']),
            "type": "rent_intention",
            "interested_rent_tickets": {"$exists": True},
            "status": {
                "$in": [
                    "requested", "assigned", "in_progress", "rejected", "confirmed_video", "booked", "holding_deposit_paid", "checked_in"
                ]
            }
        }, per_page=-1))
        if result is None:
            return {}
        if len(result) == 0:
            return {}
        time_max = result[0]['time']
        for single in result:
            time_max = max(time_max, single['time'])
        for single in result:
            if single['time'] == time_max:
                return single

    def get_user_age(ticket):
        if 'date_of_birth' in ticket:
            birth_date = ticket['date_of_birth'].date()
            today = date.today()
            age = today.year - birth_date.year
            today = today.replace(year=birth_date.year)
            if today < birth_date and age:
                age = age - 1
            return six.text_type(age)
        return ''

    def get_relocate_city(ticket):
        relocate = ''
        if 'custom_fields' in ticket and ticket['custom_fields'] is not None:
            for field in ticket['custom_fields']:
                if field.get('key', None) == 'relocate':
                    relocate = field.get('value', '')
                    break
        return relocate

    def get_user_landlord_type(user):
        ticket_list = f_app.ticket.get(f_app.ticket.search({
            "user_id": ObjectId(user['id']),
            "type": "rent",
            "status": {"$ne": "deleted"}
        }, per_page=-1))
        landlord_type_list = set()
        if ticket_list is None:
            return ''
        for single_ticket in ticket_list:
            landlord_type = single_ticket.get('landlord_type', None)
            if landlord_type is None:
                continue
            landlord_type_value = f_app.enum.get(landlord_type['id'])['value']['zh_Hans_CN']
            landlord_type_list.add(landlord_type_value)
        return '/'.join(landlord_type_list)

    def get_own_house_favorite_time(user):
        ticket_list = f_app.ticket.search({
            "user_id": ObjectId(user['id']),
            "type": "rent",
            "status": {"$ne": "deleted"}
        }, per_page=-1)
        with f_app.mongo() as m:
            total = 0
            if ticket_list is None:
                return total
            for single_ticket_id in ticket_list:
                total += m.favorites.find({
                    "status": {"$ne": "deleted"},
                    "type": "rent_ticket",
                    "ticket_id": ObjectId(single_ticket_id)
                }).count()
        return total

    def get_own_house_share_time(user):
        ticket_list = f_app.ticket.search({
            "user_id": ObjectId(user['id']),
            "type": "rent",
            "status": {"$ne": "deleted"}
        }, per_page=-1)
        with f_app.mongo() as m:
            total = 0
            if ticket_list is None:
                return total
            for single_ticket_id in ticket_list:
                total += m.log.find({
                    "type": "route",
                    "route": '/wechat-poster/' + single_ticket_id
                }).count()
        return total

    def get_own_house_viewed_time(user):
        ticket_list = f_app.ticket.search({
            "user_id": ObjectId(user['id']),
            "type": "rent",
            "status": {"$ne": "deleted"}
        }, per_page=-1)
        with f_app.mongo() as m:
            total = 0
            if ticket_list is None:
                return total
            for single_ticket_id in ticket_list:
                total += m.log.find({
                    "type": "route",
                    "route": '/property-to-rent/' + single_ticket_id
                }).count()
        return total

    def get_request_ticket_total(user):
        ticket_list = f_app.ticket.search({
            "user_id": ObjectId(user['id']),
            "type": "rent",
            "status": {"$ne": "deleted"}
        }, per_page=-1)
        with f_app.mongo() as m:
            total = 0
            if ticket_list is None:
                return total
            for single_ticket_id in ticket_list:
                total += m.tickets.find({
                    "type": "rent_intention",
                    "interested_rent_tickets": ObjectId(single_ticket_id),
                    "status": {
                        "$in": [
                            "requested", "assigned", "in_progress", "rejected", "confirmed_video", "booked", "holding_deposit_paid", "checked_in"
                        ]
                    }
                }).count()
        return total

    def get_intention_ticket_total(user):
        with f_app.mongo() as m:
            return m.tickets.find({
                "user_id": ObjectId(user['id']),
                "type": "intention",
                "status": {
                    "$ne": "deleted"
                }
            }).count()

    def get_to_rent_ticket_total(user):
        with f_app.mongo() as m:
            return m.tickets.find({
                "user_id": ObjectId(user['id']),
                "type": "rent",
                "status": "to rent"
            }).count()

    def time_period_label(ticket):
        if ticket is None:
            return ''
        time = ""
        period_start = ticket.get("rent_available_time", None)
        period_end = ticket.get("rent_deadline_time", None)
        if period_end is None or period_start is None:
            time = "不明"
        else:
            period = period_end - period_start
            if period.days >= 365:
                time = "longer than 12 months"
            elif 365 > period.days >= 180:
                time = "6 ~ 12 months"
            elif 180 > period.days >= 90:
                time = "3 ~ 6 months"
            elif 90 > period.days >= 30:
                time = "1 ~ 3 months"
            elif period.days <= 30:
                time = "less than 1 month"
        return time

    def get_detail_address(ticket):
        ticket = f_app.i18n.process_i18n(ticket)
        if f_app.util.batch_iterable(ticket.get("maponics_neighborhood", {})):
            maponics_neighborhood = ticket.get("maponics_neighborhood", {})[0]
        else:
            maponics_neighborhood = ticket.get("maponics_neighborhood", {})
        return {
            'whole': ' '.join([
                ticket.get("country", {}).get("code", ''),
                ticket.get("city", {}).get("name", ''),
                maponics_neighborhood.get("name", ''),
                ticket.get("address", ''),
                ticket.get("zipcode_index", '')
            ]),
            'country': ticket.get("country", {}).get("code", ''),
            'city': ticket.get("city", {}).get("name", ''),
            'neighborhood': maponics_neighborhood.get("name", '')
        }

    def get_to_rent_local(ticket):
        if 'property_id' not in ticket:
            return ''
        try:
            property_value = f_app.property.get(ticket['property_id'])
        except:
            return ''
        return get_detail_address(property_value)

    def get_own_house_refresh_total(ticket):
        single_ticket_id = ticket.get('id', None)
        total = 0
        if single_ticket_id is None:
            return ''
        with f_app.mongo() as m:
            total = m.log.find({
                "type": "route",
                "route": "/api/1/rent_ticket/" + six.text_type(single_ticket_id) + "/refresh"
            }).count()
        return total
    category = [
        '基本属性', '', '', '', '', '', '', '', '', '',
        '网站使用情况', '', '', '',
        '房东数据', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '',
        '租客咨询数据', '', '', '', '', '', '', '', '', '', '', '',
        '求租单统计', '', '', '', '', '',
        '投资者分析', '', '', '', '', '', '', ''
    ]
    merge = []
    header = [
        '用户名', '注册时间', '性别', '年龄', '职业', '电话', '邮箱', '国家', '城市', '用户类型',
        '单独访问次数', '活跃天数', 'app下载', '跳出页面',
        '发布过的房源数量', '房东类型', '正在发布中的房源数量', '房东收到的咨询单数量', '房源被查看的次数', '分享房产', '房源被收藏次数', '房源被刷新的次数', '有没有草稿', '发布时间', '地区', '国家', '城市', '街区', '单间还是整套', '短租长租', '租金', '已发布时间(天)', '提前多久开始出租(天)',
        '查看房产次数', '收藏房产次数', '停留时间最多的页面或rental房产', '咨询单数量', '最近一次提交时间', '最近一次入住开始时间', '最近一次入住结束时间', '提前多久开始找房', '入住人数', '吸烟', '带小孩', '带宠物',
        '求租时间', '预算', '地区', '匹配级别', '查看房东联系方式的次数', '分享房产',
        '浏览数量', '投资单提交数量', '投资意向时间', '投资预算', '期房还是现房', '几居室', '停留时间最多的页面或sales房产', '跳出的页面'
    ]

    wb = Workbook()
    ws = wb.active

    ws.append(category)
    ws.append(header)
    merge_ops_start = 0
    merge_ops_end = 0
    for index, value in enumerate(category):
        if value == '':
            merge_ops_end = index
        else:
            if index:
                merge.append(get_correct_col_index(merge_ops_start) + '1:' + get_correct_col_index(merge_ops_end) + '1')
            merge_ops_start = index
            merge_ops_end = index
    merge.append(get_correct_col_index(merge_ops_start) + '1:' + get_correct_col_index(merge_ops_end) + '1')
    for merge_ops in merge:
        ws.merge_cells(merge_ops)
        cell = ws.cell(merge_ops.split(':')[0])
        cell.alignment = Alignment(horizontal='center')

    for index, user in enumerate(f_app.user.get(f_app.user.get_active())):
        print(index)
        if user['register_time'] < datetime(2015, 12, 7):
            continue

        ticket_to_rent = get_recent_ticket_to_rent(user)
        ticket_request = get_recent_ticket_request(user)
        finding_priod = ''
        occupation = ''
        wait_rent_priod = ''
        wait_to_rent_priod = ''
        location = get_to_rent_local(ticket_to_rent)
        if location == '':
            location = {}
        if 'rent_available_time' in ticket_to_rent and 'time' in ticket_to_rent:
            wait_to_rent_priod = ticket_to_rent['rent_available_time'] - ticket_to_rent['time']
            wait_to_rent_priod = wait_to_rent_priod.days
        if 'analyze_rent_time' in user and isinstance(user.get('analyze_rent_time', ''), datetime):
            wait_rent_priod = datetime.utcnow() - user.get("analyze_rent_time", '')
            wait_rent_priod = wait_rent_priod.days
        if 'occupation' in ticket_request:
            occupation = f_app.enum.get(ticket_request['occupation']['id'])['value']['zh_Hans_CN']
        if 'rent_available_time' in ticket_request and 'time' in ticket_request:
            finding_priod = ticket_request.get('rent_available_time', '') - ticket_request.get('time', '')
            finding_priod = finding_priod.days
        ws.append(prepare_data([
            user.get("nickname", ''),  # 用户名
            user.get("register_time", ''),  # 注册时间注册时间
            '男' if ticket_request.get('gender', '') == 'male' else '女' if ticket_request.get('gender', '') == 'female' else '',  # gender
            get_user_age(ticket_request),  # age
            occupation,  # 职业
            user.get('phone', ''),  # phone
            user.get('email', ''),  # email
            user.get("country", {}).get('code', ''),  # 国家
            get_relocate_city(ticket_request).split(',')[-1],  # city
            user.get("analyze_guest_user_type", ''),  # 用户类型

            '',  # 单独访问次数
            user.get("analyze_guest_active_days", ''),  # 活跃天数
            user.get("analyze_guest_downloaded", ''),  # app下载
            '',  # exit page url

            user.get("analyze_rent_estate_total", ''),  # 发布过的房源数量
            # user.get("analyze_rent_landlord_type", ''),  # 房东类型
            get_user_landlord_type(user),  # 房东类型
            get_to_rent_ticket_total(user),  # 正在发布中的房源数量
            get_request_ticket_total(user),  # 房东收到的咨询单数量
            get_own_house_viewed_time(user),  # 房源被查看的次数
            get_own_house_share_time(user),  # 分享房产
            get_own_house_favorite_time(user),  # 房源被收藏次数
            get_own_house_refresh_total(ticket_to_rent),  # 房源被刷新的次数
            user.get("analyze_rent_has_draft", ''),  # 有没有草稿
            user.get("analyze_rent_commit_time", ''),  # 发布时间
            # user.get("analyze_rent_local", ''),  # 地区
            location.get('whole', ''),  # 地区
            location.get('country', ''),  # 地区 国家
            location.get('city', ''),  # 地区 城市
            location.get('neighborhood', ''),  # 地区 街区
            user.get("analyze_rent_single_or_whole", ''),  # 单间还是整套
            # user.get("analyze_rent_period_range", ''),  # 短租长租
            time_period_label(ticket_to_rent),  # 短租长租
            user.get("analyze_rent_price", ''),  # 租金
            wait_rent_priod,  # 已发布时间
            wait_to_rent_priod,  # 提前多久开始出租(天)

            user.get("analyze_rent_intention_views_times", ''),  # 查看房产次数
            user.get("analyze_rent_intention_favorite_times", ''),  # 收藏房产次数
            '',  # 停留时间最多的页面或rental房产
            get_request_total(user),  # 咨询单数量
            six.text_type(ticket_request.get('time', '')),  # 最近一次提交时间
            six.text_type(ticket_request.get('rent_available_time', '')),  # 最近一次入住开始时间
            six.text_type(ticket_request.get('rent_deadline_time', '')),  # 最近一次入住结束时间
            finding_priod,  # 提前多久开始找房
            ticket_request.get('tenant_count', ''),  # 入住人数
            "有" if ticket_request.get('smoke', '') is True else "否" if ticket_request.get('smoke', '') is False else '',  # 吸烟
            "有" if ticket_request.get('baby', '') is True else "否" if ticket_request.get('baby', '') is False else '',  # 带小孩
            "有" if ticket_request.get('pet', '') is True else "否" if ticket_request.get('pet', '') is False else '',  # 带宠物

            user.get("analyze_rent_intention_time", ''),  # 求租时间
            user.get("analyze_rent_intention_budget", ''),  # 预算
            user.get("analyze_rent_intention_local", ''),  # 地区
            user.get("analyze_rent_intention_match_level", ''),  # 匹配级别
            user.get("analyze_rent_intention_view_contact_times", ''),  # 查看房东联系方式的次数
            '',  # 分享房产

            user.get("analyze_intention_views_times", ''),  # 浏览数量
            get_intention_ticket_total(user),  # 投资单提交数量
            user.get("analyze_intention_time", ''),  # 投资意向时间
            user.get("analyze_intention_budget", ''),  # 投资预算
            '',  # 期房还是现房
            '',  # 几居室
            '',  # 停留时间最多的页面或sales房产
            '',  # 跳出的页面
        ]))
    format_fit(ws)
    # be_colorful(ws, 6)
    out = StringIO(save_virtual_workbook(wb))
    response.set_header(b"Content-Type", b"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    return out


@f_get('/export-excel/user-rent-request.xlsx', params=dict(
    date_from=(datetime, None),
    date_to=(datetime, None)
))
@f_app.user.login.check(force=True, role=['admin', 'jr_admin', 'sales', 'operation'])
def user_rent_request(user, params):

    def get_all_rent_request():
        search_params = {
            "type": "rent_intention",
            "interested_rent_tickets": {"$exists": True},
            "status": {
                "$nin": ['new', 'deleted'],
                "$exists": True
            }
        }
        if 'date_from' in params and 'date_to' in params:
            condition = {
                "time": {
                    "$gte": params['date_from'],
                    "$lt": params['date_to']
                }
            }
            search_params.update(condition)
        return f_app.ticket.get(f_app.ticket.search(search_params, per_page=-1, notime=True))

    def get_request_status_translate(ticket):
        status = ticket.get('status', None)
        status_dic = {
            'requested': '咨询申请已提交',
            'assigned': '已处理',
            'in_progress': '沟通中',
            'rejected': '已拒绝',
            'confirmed_video': '已确认视频看房',
            'booked': '预订确认，等待支付',
            'holding_deposit_paid': '定金已支付，确认入住',
            'canceled': '预订取消',
            'checked_in': '租客已入住'
        }
        if status not in status_dic:
            return '未知'
        return status_dic[status]

    def time_request_available(ticket):
        period_start = ticket.get("rent_available_time", None)
        request_time = ticket.get('time', None)
        period = period_start - request_time
        return period.days

    def time_period_label(ticket):
        time = ""
        period_start = ticket.get("rent_available_time", None)
        period_end = ticket.get("rent_deadline_time", None)
        if period_end is None or period_start is None:
            time = "不明"
        else:
            period = period_end - period_start
            if period.days >= 365:
                time = "longer than 12 months"
            elif 365 > period.days >= 180:
                time = "6 ~ 12 months"
            elif 180 > period.days >= 90:
                time = "3 ~ 6 months"
            elif 90 > period.days >= 30:
                time = "1 ~ 3 months"
            elif period.days <= 30:
                time = "less than 1 month"
        return time

    def get_detail_address(ticket):
        if 'property_id' in ticket:
            try:
                single_property = f_app.i18n.process_i18n(f_app.property.get(ticket['property_id']))
            except:
                return ''
            else:
                return ' '.join([
                    single_property.get("country", {}).get("code", ''),
                    single_property.get("city", {}).get("name", ''),
                    single_property.get("maponics_neighborhood", {}).get("name", ''),
                    single_property.get("address", ''),
                    single_property.get("zipcode_index", '')
                ])
        return ''

    def get_short_id(ticket):
        if 'property_id' in ticket:
            try:
                single_property = f_app.i18n.process_i18n(f_app.property.get(ticket['property_id']))
            except:
                return ''
            else:
                return single_property.get('short_id', '')
        return ''

    def get_referer_id(ticket):
        ticket_id = ticket['id']
        result = f_app.log.get(f_app.log.search({"ticket_type": "rent_intention", "type": "ticket_add", "ticket_id": ticket_id}, per_page=-1))
        if result is not None and len(result):
            return get_id_in_url(result[0].get('referer', None))
        time = ticket.get("time", None)
        diff_time = timedelta(milliseconds=500)
        flag = 0
        if time is None:
            return None
        for num, single in enumerate(referer_result):
            rst_time = single.get("time", None)
            if rst_time is None:
                continue
            if rst_time - diff_time < time < rst_time + diff_time:
                flag = 1
                record = single
        if flag:
            return get_id_in_url(record.get("referer", None))
        return None

    def get_ticket_add_ip(ticket):
        ticket_id = ticket['id']
        result = f_app.log.get(f_app.log.search({"ticket_type": "rent_intention", "type": "ticket_add", "ticket_id": ticket_id}, per_page=-1))
        if result is not None and len(result):
            ip = result[0].get('ip', None)
            if ip is not None and len(ip):
                return six.text_type(ip[0])
            else:
                return 'unknow'
        time = ticket.get("time", None)
        diff_time = timedelta(milliseconds=500)
        flag = 0
        if time is None:
            return None
        for num, single in enumerate(referer_result):
            rst_time = single.get("time", None)
            if rst_time is None:
                continue
            if rst_time - diff_time < time < rst_time + diff_time:
                flag = 1
                record = single
        if flag:
            ip = record.get('ip', None)
            if ip is not None and len(ip):
                return six.text_type(ip[0])
            else:
                return 'unknow'
        return 'unknow'

    def get_id_in_url(url):
        if url is None:
            return None
        if "property-to-rent/" in url:
            segment = url.split("property-to-rent/")[1]
            if "?" in segment:
                return segment.split("?")[0]
            return segment
        elif "ticketId=" in url:
            return url.split("ticketId=")[1]
        else:
            return None

    def get_user_age(ticket):
        if 'date_of_birth' in ticket:
            birth_date = ticket['date_of_birth'].date()
            today = date.today()
            age = today.year - birth_date.year
            today = today.replace(year=birth_date.year)
            if today < birth_date and age:
                age = age - 1
            return six.text_type(age)
        return ''

    def get_user_occupation(ticket):
        if 'occupation' in ticket:
            return f_app.enum.get(ticket['occupation']['id'])['value']['zh_Hans_CN']
        return ''

    def format_fit(sheet):
        simsun_font = Font(name="SimSun")
        alignment_fit = Alignment(shrink_to_fit=True)
        for row in sheet.rows:
            for cell in row:
                cell.font = simsun_font
                cell.alignment = alignment_fit
        for num, col in enumerate(sheet.columns):
            lenmax = 0
            for cell in col:
                lencur = 0
                if cell.value is None:
                    cell.value = ''
                if isinstance(cell.value, int) or isinstance(cell.value, datetime):
                    lencur = len(str(cell.value).encode("GBK"))
                elif cell.value is not None:
                    lencur = len(cell.value.encode("GBK", "replace"))
                if lencur > lenmax:
                    lenmax = lencur
            sheet.column_dimensions[get_correct_col_index(num)].width = lenmax * 0.86

    def get_correct_col_index(num):
        if num > 26 * 26:
            return "ZZ"
        if num >= 26:
            return get_correct_col_index(num / 26 - 1) + chr(num - 26 + 65)
        else:
            return chr(num + 65)

    def add_link(sheet, target, link=None):
        if target is None:
            return
        if f_app.util.batch_iterable(target):
            pass
        else:
            for index in range(2, len(sheet.rows) + 1):
                cell = sheet[target + six.text_type(index)]
                if len(cell.value):
                    if link is None:
                        cell.hyperlink = cell.value
                    else:
                        cell.hyperlink = six.text_type(link)

    def get_relocate(ticket):
        relocate = {}
        if 'custom_fields' in ticket and ticket['custom_fields'] is not None:
            for field in ticket['custom_fields']:
                if field.get('key', None) == 'relocate':
                    relocate['exists'] = True
                    relocate['value'] = field.get('value', '')
                    break
        return relocate

    def get_referrer_source(ticket):
        if ticket is None:
            return ''
        if 'referrer' in ticket:
            try:
                referrer_id = ObjectId(ticket['referrer'])
            except:
                return ticket['referrer']
            else:
                return f_app.enum.get(referrer_id)['value']['zh_Hans_CN']

    wb = Workbook()
    ws = wb.active
    '''header = ["咨询单状态", "咨询单描述", "咨询人昵称", "咨询人性别", "咨询人年龄", "咨询人职业", "咨询人联系方式", "咨询人邮箱", "微信", "提交时间", "起始日期",
              "终止日期", "period", "入住人数", "吸烟", "小孩", "宠物", "备注",
              "房源标题", "房源地址", "房源短id", "房东姓名", "房东电话", "房东邮箱", "房源链接"]
    ws.append(header)'''

    Header = [
        ["提交时间", "意向房源", "租期", "", "", "", "客户", "", "", "", "", "",
         "入住信息", "", "", "", "TBC", "", "Location", "", "", "relocate", "租客对房东的问题",
         "咨询处理状态", "备注", "咨询房源提交量", "房源地址", "short ID", "url", "来源"
         ],
        ["", "", "入住", "结束", "租期描述", "入住与提交时间差", "名字", "性别",
         "现状", "年龄", "电话", "邮件", "人数", "吸烟", "带小孩", "带宠物",
         "签证类型", "签证到期时间", "IP", "国家", "城市"
         ]
    ]
    merge = [
        'A1:A2',
        'B1:B2',
        'C1:F1',
        'G1:L1',
        'M1:P1',
        'Q1:R1',
        'S1:U1',
        'V1:V2',
        'W1:W2',
        'X1:X2',
        'Y1:Y2',
        'Z1:Z2',
        'AA1:AA2',
        'AB1:AB2',
        'AC1:AC2',
        'AD1:AD2'
    ]
    for header in Header:
        ws.append(header)
    for merge_ops in merge:
        ws.merge_cells(merge_ops)

    referer_result = f_app.log.output(f_app.log.search({"route": "/api/1/rent_intention_ticket/add"}, per_page=-1))
    target_ticket = []
    for ticket_request in get_all_rent_request():
        target_ticket.extend(ticket_request['interested_rent_tickets'])

    for ticket_request in get_all_rent_request():
        relocate = get_relocate(ticket_request)
        url = get_referer_id(ticket_request)
        if url is not None:
            url = "http://yangfd.com/admin?_i18n=zh_Hans_CN#/dashboard/rent/" + six.text_type(url)
        else:
            url = ''

        for ticket_id in ticket_request['interested_rent_tickets']:
            ticket = f_app.ticket.get(ticket_id, ignore_nonexist=True)
            if ticket is None:
                continue

            '''boss_id = ticket.get('user_id', None)
            if boss_id is None:
                landlord_boss = {}
            else:
                landlord_boss = f_app.user.get(boss_id)'''

            '''guest_id = ticket_request.get('user_id', None)
            if guest_id is None:
                guest_user = {}
            else:
                guest_user = f_app.user.get(guest_id)'''
            '''ws.append([
                get_request_status_translate(ticket_request),
                ticket_request.get('description', ''),
                ticket_request.get('nickname', ''),
                "男" if ticket_request.get('gender', None) == 'male' else "女",
                get_user_age(ticket_request),
                get_user_occupation(ticket_request),
                ticket_request.get('phone', ''),
                ticket_request.get('email', ''),
                guest_user.get('wechat', ''),
                unicode(timezone('Europe/London').localize(ticket_request['time'])),
                unicode(timezone('Europe/London').localize(ticket_request['rent_available_time'])),
                unicode(timezone('Europe/London').localize(ticket_request['rent_deadline_time'])),
                time_period_label(ticket_request),
                ticket_request.get('tenant_count', None),
                "是" if ticket_request.get('smoke', False) is True else "否",
                "是" if ticket_request.get('baby', False) is True else "否",
                "是" if ticket_request.get('pet', False) is True else "否",
                '',
                ticket.get('title', ''),
                get_detail_address(ticket),
                get_short_id(ticket),
                landlord_boss.get("nickname", ""),
                landlord_boss.get("phone", ""),
                landlord_boss.get("email", ""),
                url if url else ''
            ])'''
            result_final = [
                six.text_type(timezone('Europe/London').localize(ticket_request['time']).strftime("%Y-%m-%d %H:%M:%S")),
                ILLEGAL_CHARACTERS_RE.sub(r'', ticket.get('title', '')),
                six.text_type(timezone('Europe/London').localize(ticket_request['rent_available_time']).strftime("%Y-%m-%d %H:%M:%S")),
                six.text_type(timezone('Europe/London').localize(ticket_request['rent_deadline_time']).strftime("%Y-%m-%d %H:%M:%S")),
                time_period_label(ticket_request),
                time_request_available(ticket_request),
                ticket_request.get('nickname', ''),
                "男" if ticket_request.get('gender', None) == 'male' else "女",
                get_user_occupation(ticket_request),
                get_user_age(ticket_request),
                ticket_request.get('phone', ''),
                ticket_request.get('email', ''),
                ticket_request.get('tenant_count', None),
                "是" if ticket_request.get('smoke', False) is True else "否",
                "是" if ticket_request.get('baby', False) is True else "否",
                "是" if ticket_request.get('pet', False) is True else "否",
                '',
                '',
                get_ticket_add_ip(ticket_request),
                '',
                relocate.get('value', ''),
                '是' if relocate.get('exists', False) else '否',
                ticket_request.get('description', ''),
                get_request_status_translate(ticket_request),
                '',
                target_ticket.count(ticket_id),
                get_detail_address(ticket),
                get_short_id(ticket),
                url if url else '',
                get_referrer_source(ticket_request)
            ]
            try:
                ws.append(result_final)
            except:
                print("ERROR EXCEL DATA", result_final)

    format_fit(ws)
    add_link(ws, 'AC')
    out = StringIO(save_virtual_workbook(wb))
    response.set_header(b"Content-Type", b"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    return out


@f_get('/export-excel/user-rent-intention.xlsx', params=dict(
    days=(int, -1)
))
@f_app.user.login.check(force=True, role=['admin', 'jr_admin', 'sales', 'operation'])
def user_rent_intention(user, params):

    def get_data_directly_as_str(user, part, deep=None):
        if user is None:
            return ''
        user_part = user.get(part, None)
        if (user_part is not None) and (deep is not None):
            return user.get(part).get(deep, None)
        if f_app.util.batch_iterable(user_part):
            user_part = '/'.join(user_part)
        if user_part is None:
            return ''
        if isinstance(user_part, datetime):
            loc_t = timezone('Europe/London')
            loc_dt = loc_t.localize(user_part)
            return six.text_type(loc_dt.strftime('%Y-%m-%d %H:%M:%S %Z%z'))
        return six.text_type(user_part)

    def get_data_enum(user, enum_name):
        if user is None:
            return
        if enum_name not in enum_type_list:
            get_all_enum_value(enum_name)
        single = user.get(enum_name, None)
        value_list = []
        if f_app.util.batch_iterable(single):
            for true_single in single:
                if true_single is None:
                    continue
                enum_id = true_single.get("id", None)
                value = enum_type_list[enum_name].get(enum_id, None)
                value_list.append(value)
        elif single is not None:
            if single.get("id", None):
                enum_id = str(single.get("id", None))
                value = enum_type_list[enum_name].get(enum_id, None)
                if value is not None:
                    value_list.append(value)
        if not f_app.util.batch_iterable(value_list):
            value_list = [value_list]
        value_set = set(value_list)
        value_list = list(value_set)
        return '/'.join(value_list)

    def get_correct_col_index(num):
        if num > 26 * 26:
            return "ZZ"
        if num >= 26:
            return get_correct_col_index(num / 26 - 1) + chr(num - 26 + 65)
        else:
            return chr(num + 65)

    def add_link(sheet, target, link=None):
        if target is None:
            return
        if f_app.util.batch_iterable(target):
            pass
        else:
            for index in range(2, len(sheet.rows) + 1):
                cell = sheet[target + six.text_type(index)]
                if len(cell.value):
                    if link is None:
                        cell.hyperlink = cell.value
                    else:
                        cell.hyperlink = six.text_type(link)

    def get_diff_color(total):
        fill = []
        base_color = 0x999999
        color = 0x0
        if 4 <= total <= 6:
            s = 0x111111
        elif total == 3:
            s = 0x222222
        elif total == 2:
            s = 0x333333
        else:
            return
        for index in range(total):
            color = base_color + s * index
            color_t = '00' + "%x" % color
            fill.append(PatternFill(fill_type='solid', start_color=color_t, end_color=color_t))
        return fill

    def be_colorful(sheet, max_segment):
        header_fill = PatternFill(fill_type='solid', start_color='00dddddd', end_color='00dddddd')
        for cell in sheet.rows[0]:
            cell.fill = header_fill
        for col in sheet.columns:
            col_set = set()
            cell_fill = []
            col_list = []
            for num, cell in enumerate(col):
                if num and len(cell.value):
                    col_set.add(cell.value)
                if len(col_set) > max_segment:
                    break
            if max_segment >= len(col_set) > 1:
                cell_fill = get_diff_color(len(col_set))
                col_list = list(col_set)
                for number, cell in enumerate(col):
                    if not number:
                        continue
                    for index, value in enumerate(col_list):
                        if cell.value == value:
                            cell.fill = cell_fill[index]

    def format_fit(sheet):
        simsun_font = Font(name="SimSun")
        alignment_fit = Alignment(shrink_to_fit=True)
        for row in sheet.rows:
            for cell in row:
                cell.font = simsun_font
                cell.alignment = alignment_fit
        for num, col in enumerate(sheet.columns):
            lenmax = 0
            for cell in col:
                lencur = 0
                if cell.value is None:
                    cell.value = ''
                if isinstance(cell.value, int) or isinstance(cell.value, datetime):
                    lencur = len(str(cell.value).encode("GBK"))
                elif cell.value is not None:
                    lencur = len(cell.value.encode("GBK", "replace"))
                if lencur > lenmax:
                    lenmax = lencur
            sheet.column_dimensions[get_correct_col_index(num)].width = lenmax * 0.86

    def get_referer_id(ticket):
        time = ticket.get("time", None)
        diff_time = timedelta(milliseconds=500)
        flag = 0
        if time is None:
            return ''
        for num, single in enumerate(referer_result):
            rst_time = single.get("time", None)
            if rst_time is None:
                continue
            if rst_time - diff_time < time < rst_time + diff_time:
                flag = 1
                record = single
        if flag:
            return get_id_in_url(record.get("referer", ''))
        return ''

    def get_email(ticket):
        user = ticket.get("user", None)
        if user is None:
            return ''
        return user.get("email", '')

    def get_wechat(ticket):
        user = ticket.get("user", None)
        if user is None:
            return ''
        return user.get("wechat", '')

    def time_period_label(ticket):
        time = ""
        period_start = ticket.get("rent_available_time", None)
        period_end = ticket.get("rent_deadline_time", None)
        if period_end is None or period_start is None:
            time = "不明"
        else:
            period = period_end - period_start
            if period.days >= 365:
                time = "longer than 12 months"
            elif 365 > period.days >= 180:
                time = "6 ~ 12 months"
            elif 180 > period.days >= 90:
                time = "3 ~ 6 months"
            elif 90 > period.days >= 30:
                time = "1 ~ 3 months"
            elif period.days <= 30:
                time = "less than 1 month"
        return time

    def get_landlord_boss_with_ticket_id(landlord_ticket_id):
        if landlord_ticket_id is None:
            return None, None
        try:
            landlord_house = f_app.ticket.get(landlord_ticket_id)
        except:
            return None, None
        else:
            landlord_boss_id = landlord_house.get("user_id", None)
            if landlord_boss_id is None:
                return landlord_house, None
            return landlord_house, f_app.user.get(landlord_boss_id)

    def get_id_in_url(url):
        if "property-to-rent/" in url:
            segment = url.split("property-to-rent/")[1]
            if "?" in segment:
                return segment.split("?")[0]
            return segment
        elif "ticketId=" in url:
            return url.split("ticketId=")[1]
        else:
            return None

    def get_match(ticket):
        match = []
        if "partial_match" in ticket.get("tags", []):
            match.append("部分满足")
        if "perfect_match" in ticket.get("tags", []):
            match.append("完全满足")
        return '/'.join(match)

    def get_detail_address(ticket):
        return ' '.join([ticket.get("country", {}).get("code", ''),
                         ticket.get("city", {}).get("name", ''),
                         ticket.get("maponics_neighborhood", {}).get("name", ''),
                         ticket.get("address", ''),
                         ticket.get("zipcode_index", '')])

    def get_all_rent_intention(days):
        params = {
            "type": "rent_intention",
            "status": "new",
            "interested_rent_tickets": {"$exists": False}
        }
        if days > 0:
            time_now = datetime.utcnow()
            time_diff = timedelta(days=days)
            condition = {"time": {"$gt": time_now - time_diff}}
            params.update(condition)
        return f_app.ticket.output(f_app.ticket.search(params, per_page=-1, notime=True))

    def get_all_enum_value(enum_singlt_type):
        enum_list_subdic = {}
        for enumitem in f_app.i18n.process_i18n(f_app.enum.get_all(enum_singlt_type)):
            enum_list_subdic.update({enumitem["id"]: enumitem["value"]})
        enum_type_list.update({enum_singlt_type: enum_list_subdic})

    def get_referer_url(referer_id):
        if referer_id is None:
            return ''
        return "http://yangfd.com/admin?_i18n=zh_Hans_CN#/dashboard/rent/" + six.text_type(referer_id)

    # enum_type = ["rent_type", "landlord_type"]

    referer_result = f_app.log.output(f_app.log.search({"route": "/api/1/rent_intention_ticket/add"}, per_page=-1))
    enum_type_list = {}

    wb = Workbook()
    ws = wb.active
    header = ["状态", "标题", "客户", "联系方式", "邮箱", "微信", "提交时间", "起始日期",
              "终止日期", "出租需求", "预算上限", "预算下限", "period", "出租位置", "备注",
              "样房东有无匹配搭配", "目标房源", "房东类型", "房东姓名", "房东电话", "房东邮箱", "打电话了？", "有接到？", "房子租到了么？", "通过样房东",
              "如果不是通过样房东，那么是通过哪里什么样的房源？有没有交中介费", "对平台体验的想法及反馈",
              "在找房子中用户最疼的点有哪些？", "备注"]
    ws.append(header)

    for number, ticket in enumerate(get_all_rent_intention(params.get("days", -1))):
        ticket = f_app.i18n.process_i18n(ticket)
        referer_id = get_referer_id(ticket)
        landlord_result, landlord_boss = get_landlord_boss_with_ticket_id(referer_id)
        ws.append(["已提交" if (get_data_directly_as_str(ticket, "status") == "new") else "已出租",
                   get_data_directly_as_str(ticket, "title"),
                   get_data_directly_as_str(ticket, "nickname"),
                   get_data_directly_as_str(ticket, "phone"),
                   get_email(ticket),
                   get_wechat(ticket),
                   get_data_directly_as_str(ticket, "time"),
                   get_data_directly_as_str(ticket, "rent_available_time"),
                   get_data_directly_as_str(ticket, "rent_deadline_time"),
                   get_data_enum(ticket, "rent_type"),
                   get_data_directly_as_str(ticket, "rent_budget_max", "value"),
                   get_data_directly_as_str(ticket, "rent_budget_min", "value"),
                   time_period_label(ticket),
                   get_detail_address(ticket),
                   get_data_directly_as_str(ticket, "description"),
                   get_match(ticket),
                   get_referer_url(referer_id),
                   get_data_enum(landlord_result, "landlord_type"),  # 房东类型
                   get_data_directly_as_str(landlord_boss, "nickname"),
                   get_data_directly_as_str(landlord_boss, "phone"),
                   get_data_directly_as_str(landlord_boss, "email"),
                   ])

    format_fit(ws)
    add_link(ws, 'Q')
    be_colorful(ws, 6)
    out = StringIO(save_virtual_workbook(wb))
    # wb.save(out)
    response.set_header(b"Content-Type", b"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    return out


@f_api('/aggregation_featured_facility', params=dict(
    date_from=(datetime, None),
    date_to=(datetime, None),
    facility=(str, None),
    total=(int, 0)
))
@f_app.user.login.check(force=True, role=['admin', 'jr_admin', 'sales', 'operation'])
def get_featured_facility_sort(user, params):
    result_end = []
    result = []
    if 'date_from' not in params:
        params['date_from'] = datetime(2015, 12, 7)
    if 'date_to' not in params:
        params['date_to'] = datetime.utcnow()
    tickets = f_app.ticket.get(f_app.ticket.search({
        "type": "rent",
        "property_id": {"$exists": True},
        "time": {
            "$gte": params['date_from'],
            "$lte": params['date_to']
        }
    }, per_page=-1, notime=True))
    property_id_list = []
    for index, ticket in enumerate(tickets):
        if 'property_id' in ticket:
            property_id_list.append(ticket['property_id'])
    featured_facility_list = []
    for index, property in enumerate(f_app.property.output(property_id_list)):
        if property is None:
            continue
        if 'featured_facility' in property:
            for single_index, single_facility in enumerate(property['featured_facility']):
                if 'traffic_time' not in property['featured_facility'][single_index]:
                    continue
                property['featured_facility'][single_index].pop('traffic_time')
            featured_facility_list.extend(property['featured_facility'])

    for single_facility in f_app.i18n.process_i18n(featured_facility_list):
        single_result = {}
        if 'facility' not in params:
            single_result.update({
                'type': single_facility['type']['slug'],
                'id': single_facility.get('doogal_station', {}).get('id', '') if 'doogal_station' in single_facility else single_facility.get('hesa_university', {}).get('id', ''),
                'name': single_facility.get('doogal_station', {}).get('name', '') if 'doogal_station' in single_facility else single_facility.get('hesa_university', {}).get('name', '')
            })
            result.append(single_result)
        elif params['facility'] == single_facility['type']['slug']:
            single_result.update({
                'type': single_facility['type']['slug'],
                'id': single_facility.get('doogal_station', {}).get('id', '') if 'doogal_station' in single_facility else single_facility.get('hesa_university', {}).get('id', ''),
                'name': single_facility.get('doogal_station', {}).get('name', '') if 'doogal_station' in single_facility else single_facility.get('hesa_university', {}).get('name', '')
            })
            result.append(single_result)

    result_mark = []
    result_final = []
    count_list = []
    for single in result:
        if single['id'] in result_mark:
            continue
        result_mark.append(single['id'])
        single.update({'count': result.count(single)})
        count_list.append(single['count'])
        result_final.append(single)
    count_list.sort(reverse=True)
    if 'total' not in params:
        params['total'] = 0
    for index in count_list:
        for point, single in enumerate(result_final):
            if single['count'] == index:
                result_end.append(single)
                result_final.remove(single)
                break
        if params['total'] and len(result_end) > params['total']:
            break
    return result_end


@f_api('/affiliate-get-new-user-behavior', params=dict(
    date_from=(datetime, None),
    date_to=(datetime, None),
    user_id=(str, None),
    days=(int, None)
))
@f_app.user.login.check(force=True, role=['admin', 'jr_admin', 'sales', 'operation', 'affiliate'])
def affiliate_get_new_user_behavior(user, params):

    today = date.today()
    date_begin = datetime.strptime(today.replace(day=1).isoformat(), "%Y-%m-%d")
    date_end = datetime.strptime(today.replace(day=1, month=today.month + 1).isoformat(), "%Y-%m-%d")
    if 'date_from' in params:
        date_begin = params['date_from']
    if 'date_to' in params:
        date_end = params['date_to']
    result = []

    if 'affiliate' in user.get('role', []) and 'admin' not in user.get('role', []):
        if 'user_id' in params:
            request_user = f_app.user.get(params['user_id'])
            user_select_rang = [ObjectId(params['user_id'])]
            if unicode(user['id']) not in [unicode(request_user.get('referral', '')), unicode(params['user_id'])]:
                return
        else:
            user_select_rang = f_app.user.search({
                "status": {"$ne": "deleted"},
                'referral': ObjectId(user['id'])
            })
    else:
        user_select_rang = f_app.user.get_active()
        if 'user_id' in params:
            user_select_rang = [ObjectId(params['user_id'])]

    for single_user in f_app.user.get(user_select_rang):
        if 'affiliate' in single_user.get('role', []) or 'referral_code' in single_user:

            search_condition = {
                "status": {"$ne": "deleted"},
                'referral': ObjectId(single_user['id'])
            }
            if 'date_from' in params:
                search_condition.update({
                    "register_time": {
                        "$gte": date_begin
                    }
                })
            if 'date_to' in params:
                search_condition.update({
                    "register_time": {
                        "$lt": date_end
                    }
                })
            if 'date_from' in params and 'date_to' in params:
                search_condition.update({
                    "register_time": {
                        "$gte": date_begin,
                        "$lt": date_end
                    }
                })
            user_list = f_app.user.search(search_condition)

            affiliate_user_request_count = 0
            affiliate_user_success_rent = 0
            user_count = 0
            if user_list is not None:

                user_count = len(user_list)
                for affliate_user in user_list:
                    affiliate_user_list = f_app.ticket.search({
                        "user_id": ObjectId(affliate_user),
                        "type": "rent_intention",
                        "interested_rent_tickets": {"$exists": True},
                        "status": {
                            "$in": [
                                "requested", "assigned", "in_progress", "rejected", "confirmed_video", "booked", "holding_deposit_paid", "checked_in"
                            ]
                        }
                    }, per_page=-1)
                    if affiliate_user_list is not None:
                        affiliate_user_request_count += len(affiliate_user_list)

                for affliate_user in user_list:
                    search_ticket_condition = {
                        "user_id": ObjectId(affliate_user),
                        "type": "rent_intention",
                        "status": "checked_in"
                    }
                    if 'date_from' in params:
                        search_ticket_condition.update({
                            "time": {
                                "$gte": date_begin
                            }
                        })
                    if 'date_to' in params:
                        search_ticket_condition.update({
                            "time": {
                                "$lt": date_end
                            }
                        })
                    if 'date_from' in params and 'date_to' in params:
                        search_ticket_condition.update({
                            "time": {
                                "$gte": date_begin,
                                "$lt": date_end
                            }
                        })
                    affiliate_user_list = f_app.ticket.search(search_ticket_condition, per_page=-1, notime=True)
                    if affiliate_user_list is not None:
                        affiliate_user_success_rent += len(affiliate_user_list)
            result.append({
                "affiliate_user_id": single_user['id'],
                "affiliate_new_user_total": user_count,
                "affiliate_user_request_count": affiliate_user_request_count,
                "affiliate_user_success_rent": affiliate_user_success_rent
            })

    return result


@f_api('/affiliate-get-all-user-behavior', params=dict(
    date_from=(datetime, None),
    date_to=(datetime, None),
    user_id=(str, None)
))
@f_app.user.login.check(force=True, role=['admin', 'jr_admin', 'sales', 'operation', 'affiliate'])
def affiliate_get_all_user_behavior(user, params):
    today = date.today()
    date_begin = datetime.strptime(today.replace(day=1).isoformat(), "%Y-%m-%d")
    date_end = datetime.strptime(today.replace(day=1, month=today.month + 1).isoformat(), "%Y-%m-%d")
    if 'date_from' in params:
        date_begin = params['date_from']
    if 'date_to' in params:
        date_end = params['date_to']
    result = []

    if 'affiliate' in user.get('role', []) and 'admin' not in user.get('role', []):
        if 'user_id' in params:
            request_user = f_app.user.get(params['user_id'])
            user_select_rang = [ObjectId(params['user_id'])]
            if unicode(user['id']) not in [unicode(request_user.get('referral', '')), unicode(params['user_id'])]:
                return
        else:
            user_select_rang = f_app.user.search({
                "status": {"$ne": "deleted"},
                'referral': ObjectId(user['id'])
            })
    else:
        user_select_rang = f_app.user.get_active()
        if 'user_id' in params:
            user_select_rang = [ObjectId(params['user_id'])]

    for single_user in f_app.user.get(user_select_rang):
        if 'affiliate' in single_user.get('role', []) or 'referral_code' in single_user:

            search_condition = {
                "status": {"$ne": "deleted"},
                'referral': ObjectId(single_user['id'])
            }
            user_list = f_app.user.search(search_condition)

            affiliate_all_user_request_count = 0
            affiliate_all_user_success_rent = 0
            if user_list is not None:

                for affliate_user in user_list:
                    search_ticket_condition = {
                        "user_id": ObjectId(affliate_user),
                        "type": "rent_intention",
                        "status": "checked_in"
                    }
                    search_request_ticket_condition = {
                        "user_id": ObjectId(affliate_user),
                        "type": "rent_intention",
                        "interested_rent_tickets": {"$exists": True},
                        "status": {
                            "$in": [
                                "requested", "assigned", "in_progress", "rejected", "confirmed_video", "booked", "holding_deposit_paid", "checked_in"
                            ]
                        }
                    }
                    if 'date_from' in params:
                        search_ticket_condition.update({
                            "time": {
                                "$gte": date_begin
                            }
                        })
                        search_request_ticket_condition.update({
                            "time": {
                                "$gte": date_begin
                            }
                        })
                    if 'date_to' in params:
                        search_ticket_condition.update({
                            "time": {
                                "$lt": date_end
                            }
                        })
                        search_request_ticket_condition.update({
                            "time": {
                                "$lt": date_end
                            }
                        })
                    if 'date_from' in params and 'date_to' in params:
                        search_ticket_condition.update({
                            "rent_time": {
                                "$gte": date_begin,
                                "$lt": date_end
                            }
                        })
                        search_request_ticket_condition.update({
                            "time": {
                                "$gte": date_begin,
                                "$lt": date_end
                            }
                        })
                    affiliate_user_list = f_app.ticket.search(search_ticket_condition, per_page=-1, notime=True)
                    if affiliate_user_list is not None:
                        affiliate_all_user_success_rent += len(affiliate_user_list)
                    affiliate_user_request_list = f_app.ticket.search(search_request_ticket_condition, per_page=-1, notime=True)
                    if affiliate_user_request_list is not None:
                        affiliate_all_user_request_count += len(affiliate_user_request_list)
            result.append({
                "affiliate_user_id": single_user['id'],
                "affiliate_all_user_request_count": affiliate_all_user_request_count,
                "affiliate_all_user_success_rent": affiliate_all_user_success_rent
            })
    return result


@f_api('/affiliate-get-invited-user-count-detail', params=dict(
    date_from=(datetime, None),
    date_to=(datetime, None),
    user_id=(str, None)
))
@f_app.user.login.check(force=True, role=['admin', 'jr_admin', 'sales', 'operation', 'affiliate'])
def affiliate_get_invited_user_count_detail(user, params):
    today = date.today()
    date_begin = datetime.strptime(today.replace(day=1).isoformat(), "%Y-%m-%d")
    date_end = datetime.strptime(today.replace(day=1, month=today.month + 1).isoformat(), "%Y-%m-%d")
    if 'date_from' in params:
        date_begin = params['date_from']
    if 'date_to' in params:
        date_end = params['date_to']

    if 'user_id' in params:
        user_id = params['user_id']
    single_user = f_app.user.get(ObjectId(user_id))

    search_condition = {
        "status": {"$ne": "deleted"},
        'referral': {"$exists": True},
    }

    if 'affiliate' in user.get('role', []) and 'admin' not in user.get('role', []):
        if unicode(user['id']) not in [unicode(single_user.get('referral', '')), unicode(user_id)]:
            return
        else:
            search_condition = {
                "referral": ObjectId(user['id'])
            }

    if 'user_id' in params:
        search_condition.update({
            'referral': ObjectId(params['user_id'])
        })
    if 'date_from' in params:
        search_condition.update({
            "register_time": {
                "$gte": date_begin
            }
        })
    if 'date_to' in params:
        search_condition.update({
            "register_time": {
                "$lt": date_end
            }
        })
    if 'date_from' in params and 'date_to' in params:
        search_condition.update({
            "register_time": {
                "$gte": date_begin,
                "$lt": date_end
            }
        })
    user_list = f_app.user.search(search_condition)

    affiliate_user_success_invited_user_count = 0
    if user_list is not None:
        user_list = f_app.user.get(user_list)
        affiliate_user_set = set()
        for single_user in user_list:
            if single_user['referral'] is not None:
                affiliate_user_set.add(single_user['referral'])
        affiliate_user_success_invited_user_count = len(affiliate_user_set)
        result = []
        for single_affiliate_user in affiliate_user_set:
            affiliate_invited_user_count = 0
            for single_user in user_list:
                if unicode(single_user['referral']) == unicode(single_affiliate_user):
                    affiliate_invited_user_count += 1
            result.append({
                'affiliate_user_id': unicode(single_affiliate_user),
                'invited_user_count': affiliate_invited_user_count,
            })
    return {
        "affiliate_invitor_total": affiliate_user_success_invited_user_count,
        "affiliate_user_invited_detail_count": result
    }


@f_api('/affiliate-get-sub-user-count', params=dict(
    date_from=(datetime, None),
    date_to=(datetime, None),
    user_id=(str, None)
))
@f_app.user.login.check(force=True, role=['admin', 'jr_admin', 'sales', 'operation', 'affiliate'])
def affiliate_get_sub_user_count(user, params):
    result_level_info = {}
    if 'user_id' in params:
        user_id = params['user_id']
    else:
        return
    single_user = f_app.user.get(ObjectId(user_id))
    if 'affiliate' in user.get('role', []) and 'admin' not in user.get('role', []):
        if unicode(user['id']) not in [unicode(single_user.get('referral', '')), unicode(user_id)]:
            return
    if 'affiliate' in single_user.get('role', []) or 'referral_code' in single_user:
        this_level_user_list = [single_user['id']]
        if single_user.get('referral', None) is not None:
            this_level_user_list = f_app.user.search({
                "referral": ObjectId(single_user['referral']),
                "status": {"$ne": "deleted"}
            })
        level = 1
        this_level_count = 0
        if this_level_user_list is not None:
            this_level_count = len(this_level_user_list)
        result_level_info.update({
            "level" + unicode(level): this_level_count
        })
        while True:
            level += 1
            this_level_count = 0
            down_level_user_list = []
            for this_level_single_user in f_app.user.get(this_level_user_list):
                if this_level_single_user.get('referral_code', None) is not None:
                    down_level_user_list_extend = f_app.user.search({
                        "referral": ObjectId(this_level_single_user['id']),
                        "status": {"$ne": "deleted"}
                    })
                    if down_level_user_list_extend is not None:
                        down_level_user_list.extend(down_level_user_list_extend)
            this_level_count = len(down_level_user_list)
            if this_level_count == 0:
                break
            result_level_info.update({
                "level" + unicode(level): this_level_count
            })
            this_level_user_list = list(down_level_user_list)
    return {
        "affiliate_invited_all_level_total": sum([result_level_info[item] for item in result_level_info]) - result_level_info.get('level1', 0),
        "affiliate_invited_each_level_detail": result_level_info
    }


@f_get('/export-excel/facility-around-rent-info.xlsx', params=dict(
    date_from=(datetime, None),
    date_to=(datetime, None)
))
@f_app.user.login.check(force=True, role=['admin', 'jr_admin', 'sales', 'operation'])
def get_featured_facilities_around_rent(user, params):

    if 'date_from' in params:
        date_from = params['date_from']
    else:
        date_from = datetime(2016, 2, 3)

    if 'date_to' in params:
        date_to = params['date_to']
    else:
        date_to = datetime.utcnow()

    def get_own_house_viewed_time(ticket):
        with f_app.mongo() as m:
            total = m.log.find({
                "type": "route",
                "route": {"$in": ['/wechat-poster/' + unicode(ticket['id']), "/property-to-rent/" + unicode(ticket['id'])]},
                "time": {
                    "$gte": date_from,
                    "$lte": date_to
                }
            }).count()
        return int(total)

    def get_request_ticket_total(ticket):
        with f_app.mongo() as m:
            total = m.tickets.find({
                "type": "rent_intention",
                "interested_rent_tickets": ObjectId(ticket['id']),
                "time": {
                    "$gte": date_from,
                    "$lte": date_to
                },
                "status": {
                    "$in": [
                        "requested", "assigned", "in_progress", "rejected", "confirmed_video", "booked", "holding_deposit_paid", "checked_in"
                    ]
                }
            }).count()
        return int(total)

    tickets_id = []
    with f_app.mongo() as m:
        cursor = m.tickets.aggregate(
            [
                {
                    '$match': {
                        "type": "rent",
                        "status": {"$ne": "draft"}
                    }
                },
                {
                    '$group': {
                        "_id": '$_id'
                    }
                }
            ]
        )
        for single in cursor:
            tickets_id.append(ObjectId(single['_id']))

    doogal_station_list = {
        'view_times': {},
        'request_times': {},
        'ticket_total': {}
    }
    hesa_university_list = {
        'view_times': {},
        'request_times': {},
        'ticket_total': {}
    }
    maponics_neighborhood_list = {
        'view_times': {},
        'request_times': {},
        'ticket_total': {}
    }
    city_list = {
        'view_times': {},
        'request_times': {},
        'ticket_total': {}
    }

    print len(tickets_id)
    for index, rent_ticket_id in enumerate(tickets_id):
        # print index
        if rent_ticket_id is not None:
            try:
                with f_app.mongo() as m:
                    rent_ticket = m.tickets.find_one({'_id': rent_ticket_id})
                    rent_ticket['id'] = ObjectId(rent_ticket.pop('_id'))
            except:
                continue
        else:
            continue
        if rent_ticket is None:
            continue
        if 'property_id' in rent_ticket:
            try:
                with f_app.mongo() as m:
                    this_property = m.propertys.find_one({'_id': rent_ticket['property_id']})
                    this_property['id'] = ObjectId(this_property.pop('_id'))
            except:
                print "property get fail"
                continue
            if this_property is None:
                continue
        else:
            continue

        doogal_station_id = None
        hesa_university_id = None
        maponics_neighborhood_id = None
        city_id = None

        if 'featured_facility' in this_property:
            doogal_station_id = []
            hesa_university_id = []
            for single_facility in this_property['featured_facility']:
                if "doogal_station" in single_facility:
                    doogal_station_id.append(single_facility['doogal_station'])
                elif "hesa_university" in single_facility:
                    hesa_university_id.append(single_facility['hesa_university'])

        if 'maponics_neighborhood' in this_property:
            if '_id' in this_property['maponics_neighborhood']:
                maponics_neighborhood_id = this_property['maponics_neighborhood']['_id']

        if 'city' in this_property:
            if '_id' in this_property['city']:
                city_id = this_property['city']['_id']

        view_times = get_own_house_viewed_time(rent_ticket)
        request_times = get_request_ticket_total(rent_ticket)

        if doogal_station_id is not None:
            for single_doogal_station in doogal_station_id:
                doogal_station_list['ticket_total'].update({single_doogal_station: doogal_station_list['ticket_total'].get(single_doogal_station, 0) + 1})
        if hesa_university_id is not None:
            for single_university in hesa_university_id:
                hesa_university_list['ticket_total'].update({single_university: hesa_university_list['ticket_total'].get(single_university, 0) + 1})
        maponics_neighborhood_list['ticket_total'].update({maponics_neighborhood_id: maponics_neighborhood_list['ticket_total'].get(maponics_neighborhood_id, 0) + 1})
        city_list['ticket_total'].update({city_id: city_list['ticket_total'].get(city_id, 0) + 1})

        if view_times:
            if doogal_station_id is not None:
                for single_doogal_station in doogal_station_id:
                    doogal_station_list['view_times'].update({single_doogal_station: doogal_station_list['view_times'].get(single_doogal_station, 0) + 1 * view_times})
            if hesa_university_id is not None:
                for single_university in hesa_university_id:
                    hesa_university_list['view_times'].update({single_university: hesa_university_list['view_times'].get(single_university, 0) + 1 * view_times})
            if maponics_neighborhood_id is not None:
                maponics_neighborhood_list['view_times'].update({maponics_neighborhood_id: maponics_neighborhood_list['view_times'].get(maponics_neighborhood_id, 0) + 1 * view_times})
            if city_id is not None:
                city_list['view_times'].update({city_id: city_list['view_times'].get(city_id, 0) + 1 * view_times})

        if request_times:
            if doogal_station_id is not None:
                for single_doogal_station in doogal_station_id:
                    doogal_station_list['request_times'].update({single_doogal_station: doogal_station_list['request_times'].get(single_doogal_station, 0) + 1 * request_times})
            if hesa_university_id is not None:
                for single_university in hesa_university_id:
                    hesa_university_list['request_times'].update({single_university: hesa_university_list['request_times'].get(single_university, 0) + 1 * request_times})
            if maponics_neighborhood_id is not None:
                maponics_neighborhood_list['request_times'].update({maponics_neighborhood_id: maponics_neighborhood_list['request_times'].get(maponics_neighborhood_id, 0) + 1 * request_times})
            if city_id is not None:
                city_list['request_times'].update({city_id: city_list['request_times'].get(city_id, 0) + 1 * request_times})

        # if index >= 15:
        #     break

    header = ["名", "查看", "咨询", "房源"]
    doogal_station_result = []
    hesa_university_result = []
    maponics_neighborhood_result = []
    city_result = []

    sort_temp = [doogal_station_list['view_times'][single] for single in doogal_station_list['view_times']]
    sort_temp.sort(reverse=True)
    for value in sort_temp:
        for single in doogal_station_list['view_times']:
            if doogal_station_list['view_times'][single] == value:
                try:
                    station = f_app.doogal.station.get(single)
                except:
                    station = {}
                doogal_station_result.append([unicode(station.get('name', '')), unicode(value), unicode(doogal_station_list['request_times'].get(single, 0)), unicode(doogal_station_list['ticket_total'].get(single, 0))])
                doogal_station_list['view_times'].pop(single)
                break

    sort_temp = [hesa_university_list['view_times'][single] for single in hesa_university_list['view_times']]
    sort_temp.sort(reverse=True)
    for value in sort_temp:
        for single in hesa_university_list['view_times']:
            if hesa_university_list['view_times'][single] == value:
                try:
                    university = f_app.hesa.university.get(single)
                except:
                    university = {}
                hesa_university_result.append([unicode(university.get('name', '')), unicode(value), unicode(hesa_university_list['request_times'].get(single, 0)), unicode(hesa_university_list['ticket_total'].get(single, 0))])
                hesa_university_list['view_times'].pop(single)
                break

    sort_temp = [maponics_neighborhood_list['view_times'][single] for single in maponics_neighborhood_list['view_times']]
    sort_temp.sort(reverse=True)
    for value in sort_temp:
        for single in maponics_neighborhood_list['view_times']:
            if maponics_neighborhood_list['view_times'][single] == value:
                try:
                    neighborhood = f_app.maponics.neighborhood.get(single)
                except:
                    neighborhood = {}
                maponics_neighborhood_result.append([unicode(neighborhood.get('name', '')), unicode(value), unicode(maponics_neighborhood_list['request_times'].get(single, 0)), unicode(maponics_neighborhood_list['ticket_total'].get(single, 0))])
                maponics_neighborhood_list['view_times'].pop(single)
                break

    sort_temp = [city_list['view_times'][single] for single in city_list['view_times']]
    sort_temp.sort(reverse=True)
    for value in sort_temp:
        for single in city_list['view_times']:
            if city_list['view_times'][single] == value:
                try:
                    city = f_app.geonames.gazetteer.get(single)
                except:
                    city = {}
                city_result.append([unicode(city.get('name', '')), unicode(value), unicode(city_list['request_times'].get(single, 0)), unicode(city_list['ticket_total'].get(single, 0))])
                city_list['view_times'].pop(single)
                break

    def get_correct_col_index(num):
        if num > 26 * 26:
            return "ZZ"
        if num >= 26:
            return get_correct_col_index(num / 26 - 1) + get_correct_col_index(num % 26)
        else:
            return chr(num + 65)

    def format_fit(sheet):
        simsun_font = Font(name="SimSun")
        alignment_fit = Alignment(shrink_to_fit=True)
        for row in sheet.rows:
            for cell in row:
                cell.font = simsun_font
                cell.alignment = alignment_fit
        for num, col in enumerate(sheet.columns):
            lenmax = 0
            for cell in col:
                lencur = 0
                if cell.value is None:
                    cell.value = ''
                if isinstance(cell.value, int) or isinstance(cell.value, datetime):
                    lencur = len(six.text_type(cell.value).encode("GBK"))
                elif cell.value is not None:
                    lencur = len(cell.value.encode("GBK", "replace"))
                if lencur > lenmax:
                    lenmax = lencur
            sheet.column_dimensions[get_correct_col_index(num)].width = lenmax * 0.86
            print("col", get_correct_col_index(num), "fit.")

    wb = Workbook()

    ws_doogal_station_result = wb.active
    ws_hesa_university_result = wb.create_sheet()
    ws_maponics_neighborhood_result = wb.create_sheet()
    ws_city_result = wb.create_sheet()

    ws_doogal_station_result.title = "doogal_station_result"
    ws_hesa_university_result.title = "hesa_university_result"
    ws_maponics_neighborhood_result.title = "maponics_neighborhood_result"
    ws_city_result.title = "city_result"

    ws_doogal_station_result.append(header)
    ws_hesa_university_result.append(header)
    ws_maponics_neighborhood_result.append(header)
    ws_city_result.append(header)
    for single in doogal_station_result:
        ws_doogal_station_result.append(single)
    for single in hesa_university_result:
        ws_hesa_university_result.append(single)
    for single in maponics_neighborhood_result:
        ws_maponics_neighborhood_result.append(single)
    for single in city_result:
        ws_city_result.append(single)

    format_fit(ws_doogal_station_result)
    format_fit(ws_hesa_university_result)
    format_fit(ws_maponics_neighborhood_result)
    format_fit(ws_city_result)
    out = StringIO(save_virtual_workbook(wb))
    response.set_header(b"Content-Type", b"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    return out


@f_api('/get-users-portrait', params=dict(
    date_from=(datetime, None),
    date_to=(datetime, None)
))
@f_app.user.login.check(force=True, role=['admin', 'jr_admin', 'sales', 'operation'])
def get_users_portrait(user, params):

    def get_user_age(total):
        source = {}
        result = {
            "age_distribution": {
                "32~999": 0,
                "32~27": 0,
                "23~27": 0,
                "18~23": 0,
                "0~18": 0
            }
        }
        condition = {
            "type": "rent_intention",
            "date_of_birth": {"$exists": True},
            "status": {
                "$in": [
                    "requested", "assigned", "in_progress", "rejected", "confirmed_video", "booked", "holding_deposit_paid", "checked_in"
                ]
            }
        }
        if 'date_from' in params and 'date_to' in params:
            condition.update({
                "time": {
                    "$gte": params['date_from'],
                    "$lt": params['date_to']
                }
            })
        ticekts = f_app.ticket.search(condition, per_page=-1, notime=True)
        for ticket in f_app.ticket.get(ticekts):
            if 'date_of_birth' in ticket:
                birth_date = ticket['date_of_birth'].date()
                today = date.today()
                age = today.year - birth_date.year
                today = today.replace(year=birth_date.year)
                if today < birth_date and age:
                    age = age - 1
                source.update({ticket['user_id']: age})
        for single in source:
            if source[single] > 32:
                result["age_distribution"]['32~999'] += 1
            elif 32 >= source[single] > 27:
                result["age_distribution"]['32~27'] += 1
            elif 27 >= source[single] > 23:
                result["age_distribution"]['23~27'] += 1
            elif 23 >= source[single] > 18:
                result["age_distribution"]['18~23'] += 1
            elif 18 >= source[single]:
                result["age_distribution"]['0~18'] += 1

        result['age_distribution'].update({
            "other": total - len(source)
        })

        return result['age_distribution']

    def get_occupation(total):
        source = {}
        condition = {
            "type": "rent_intention",
            "occupation": {"$exists": True},
            "status": {
                "$in": [
                    "requested", "assigned", "in_progress", "rejected", "confirmed_video", "booked", "holding_deposit_paid", "checked_in"
                ]
            }
        }
        if 'date_from' in params and 'date_to' in params:
            condition.update({
                "time": {
                    "$gte": params['date_from'],
                    "$lt": params['date_to']
                }
            })
        ticekts = f_app.ticket.search(condition, per_page=-1, notime=True)
        occupation_id = None
        for ticket in f_app.ticket.get(ticekts):
            if 'occupation' in ticket:
                occupation_id = ticket['occupation']['id']
                source.update({ticket['user_id']: occupation_id})
        occupation_count = {}
        for single in source:
            if source[single] not in occupation_count:
                occupation_count[source[single]] = 0
            occupation_count[source[single]] += 1
        result = {}
        for single in occupation_count:
            result.update({f_app.enum.get(single)['value']['zh_Hans_CN']: occupation_count[single]})
        result.update({"other": total - len(source)})
        return result

    value = {}
    with f_app.mongo() as m:
        if 'date_from' in params and 'date_to' in params:
            value.update({
                "user_portrait_register_user_total": m.users.find({
                    "register_time": {
                        "$gte": params['date_from'],
                        "$lt": params['date_to']
                    },
                    "status": {"$ne": "deleted"},
                }).count()
            })
            aggregate_params = [
                {
                    "$match": {
                        "register_time": {
                            "$gte": params['date_from'],
                            "$lt": params['date_to']
                        }
                    }
                },
                {"$unwind": "$user_type"},
                {"$group": {"_id": "$user_type._id", "count": {"$sum": 1}}}
            ]
            aggregate_params_gender = [
                {
                    "$match": {
                        "register_time": {
                            "$gte": params['date_from'],
                            "$lt": params['date_to']
                        }
                    }
                },
                {"$group": {"_id": "$gender", "count": {"$sum": 1}}}
            ]
            aggregate_params_country = [
                {
                    "$match": {
                        "register_time": {
                            "$gte": params['date_from'],
                            "$lt": params['date_to']
                        }
                    }
                },
                {"$group": {"_id": "$country.code", "count": {"$sum": 1}}}
            ]
        else:
            value.update({
                "user_portrait_register_user_total": m.users.find({
                    "register_time": {"$exists": True},
                    "status": {"$ne": "deleted"},
                }).count()
            })
            aggregate_params = [
                {"$unwind": "$user_type"},
                {"$group": {"_id": "$user_type._id", "count": {"$sum": 1}}}
            ]
            aggregate_params_gender = [
                {"$group": {"_id": "$gender", "count": {"$sum": 1}}}
            ]
            aggregate_params_country = [
                {"$group": {"_id": "$country.code", "count": {"$sum": 1}}}
            ]
        value["user_portrait_active_days"] = {}
        value["user_portrait_active_days"].update({
            "0~1": m.users.find({
                "register_time": {"$exists": True},
                "status": {"$ne": "deleted"},
                "analyze_guest_active_days": {
                    "$lte": 1
                }
            }).count(),
            "2~7": m.users.find({
                "register_time": {"$exists": True},
                "status": {"$ne": "deleted"},
                "analyze_guest_active_days": {
                    "$gte": 2,
                    "$lte": 7
                }
            }).count(),
            "8~14": m.users.find({
                "register_time": {"$exists": True},
                "status": {"$ne": "deleted"},
                "analyze_guest_active_days": {
                    "$gte": 8,
                    "$lte": 14
                }
            }).count(),
            "14+": m.users.find({
                "register_time": {"$exists": True},
                "status": {"$ne": "deleted"},
                "analyze_guest_active_days": {
                    "$gte": 15
                }
            }).count()
        })

        cursor = m.users.aggregate(aggregate_params)
        user_type = {}
        for document in cursor:
            user_type.update({f_app.enum.get(document['_id'])['value']['zh_Hans_CN']: document['count']})
        user_type.update({
            "other": m.users.find({"user_type": {"$exists": False}}).count()
        })
        cursor.close()
        value.update({"user_portrait_user_type": user_type})

        cursor = m.users.aggregate(aggregate_params_gender)
        gender_type = {}
        for document in cursor:
            gender_type.update({document['_id'] if document['_id'] else "other": document['count']})
        cursor.close()
        value.update({"user_portrait_gender_type": gender_type})

        cursor = m.users.aggregate(aggregate_params_country)
        country_distribution = {}
        for document in cursor:
            country_distribution.update({document['_id'] if document['_id'] else "other": document['count']})
        cursor.close()
        value.update({"user_portrait_country_distribution": country_distribution})

        value.update({"user_portrait_age_distribution": get_user_age(value['user_portrait_register_user_total'])})
        value.update({"user_portrait_occupation_distribution": get_occupation(value['user_portrait_register_user_total'])})

        app_user = m.users.find({"analyze_guest_downloaded": "已下载"}).count()
        value.update({"user_portrait_user_device": {
            "app": app_user,
            "other": value['user_portrait_register_user_total'] - app_user
        }})

    return value


@f_api('/get-users-portrait-tenants-behavior', params=dict(
    date_from=(datetime, None),
    date_to=(datetime, None)
))
@f_app.user.login.check(force=True, role=['admin', 'jr_admin', 'sales', 'operation'])
def get_users_portrait_tenants_behavior(user, params):
    value = {}
    source_finding = {}
    source_rent = {}
    condition = {
        "type": "rent_intention",
        "status": {
            "$in": [
                "requested", "assigned", "in_progress", "rejected", "confirmed_video", "booked", "holding_deposit_paid", "checked_in"
            ]
        }
    }
    if 'date_from' in params and 'date_to' in params:
        condition.update({
            "time": {
                "$gte": params['date_from'],
                "$lt": params['date_to']
            }
        })
    ticekts = f_app.ticket.search(condition, per_page=-1, notime=True)
    for ticket in f_app.ticket.get(ticekts):
        if 'rent_available_time' not in ticket:
            continue
        finding_priod = ticket['rent_available_time'] - ticket['time']
        source_finding.update({unicode(finding_priod.days): source_finding.get(unicode(finding_priod.days), 0) + 1})
        rent_priod = ticket['rent_deadline_time'] - ticket['rent_available_time']
        source_rent.update({unicode(rent_priod.days): source_rent.get(unicode(rent_priod.days), 0) + 1})
    value.update({"want_rent_days_distribution": {
        "0~1 month": sum([source_rent[single] if -1 <= int(single) < 30 else 0 for single in source_rent]),
        "1~3 month": sum([source_rent[single] if 30 <= int(single) < 91 else 0 for single in source_rent]),
        "3~6 month": sum([source_rent[single] if 91 <= int(single) < 183 else 0 for single in source_rent]),
        "6~9 month": sum([source_rent[single] if 183 <= int(single) < 275 else 0 for single in source_rent]),
        "9~12 month": sum([source_rent[single] if 275 <= int(single) < 365 else 0 for single in source_rent]),
        "12+ month": sum([source_rent[single] if 365 <= int(single) else 0 for single in source_rent]),
    }})
    value.update({"finding_rent_days_distribution": {
        "0~1 weeks": sum([source_finding[single] if -1 <= int(single) < 7 else 0 for single in source_finding]),
        "1~2 weeks": sum([source_finding[single] if 7 <= int(single) < 14 else 0 for single in source_finding]),
        "2~4 weeks": sum([source_finding[single] if 14 <= int(single) < 30 else 0 for single in source_finding]),
        "1~2 month": sum([source_finding[single] if 30 <= int(single) < 60 else 0 for single in source_finding]),
        "2+ month": sum([source_finding[single] if 60 <= int(single) else 0 for single in source_finding]),
    }})

    def get_tenant_count(min, max=None):
        search_condition = {
            "type": "rent_intention",
            "status": {
                "$in": [
                    "requested", "assigned", "in_progress", "rejected", "confirmed_video", "booked", "holding_deposit_paid", "checked_in"
                ]
            },
            "tenant_count": 1
        }
        if 'date_from' in params and 'date_to' in params:
            search_condition.update({
                "time": {
                    "$gte": params['date_from'],
                    "$lt": params['date_to']
                }
            })
        if max is None:
            search_condition.update({"tenant_count": {"$gte": min}})
        else:
            search_condition.update({"tenant_count": min})
        ticket = f_app.ticket.search(search_condition, per_page=-1, notime=True)
        return len(ticket)

    value.update({
        "tenant_count": {
            "single": get_tenant_count(1, 1),
            "couple": get_tenant_count(2, 2),
            "family": get_tenant_count(3)
        }
    })

    with f_app.mongo() as m:
        match_condition = {
            "type": "route",
            "rent_ticket_id": {"$exists": True},
            "id": {"$exists": True, "$ne": None}
        }
        if 'date_from' in params and 'date_to' in params:
            match_condition.update({
                "time": {
                    "$gte": params['date_from'],
                    "$lt": params['date_to']
                }
            })
        cursor = m.log.aggregate([
            {"$match": match_condition},
            {"$group": {
                "_id": "$id",
                "count": {"$sum": 1}
            }}
        ])
        result = {}
        for single in cursor:
            if 1 <= single['count'] < 10:
                result.update({"1~10": result.get('1~10', 0) + 1})
            elif 10 <= single['count'] < 20:
                result.update({"10~20": result.get('10~20', 0) + 1})
            elif 20 <= single['count'] < 30:
                result.update({"20~30": result.get('20~30', 0) + 1})
            elif 30 <= single['count'] < 100:
                result.update({"30~100": result.get('30~100', 0) + 1})
            elif 100 <= single['count']:
                result.update({"100+": result.get('100+', 0) + 1})

        value.update({"ticket_access_time": result})

        match_condition = {
            "user_id": {"$exists": True, "$ne": None},
            "property_id": {"$exists": True},
            "status": "new"
        }
        if 'date_from' in params and 'date_to' in params:
            match_condition.update({
                "time": {
                    "$gte": params['date_from'],
                    "$lt": params['date_to']
                }
            })
        cursor = m.favorites.aggregate([
            {'$match': match_condition},
            {'$group': {
                "_id": "$user_id",
                "count": {"$sum": 1}
            }}
        ])
        result = {}
        for single in cursor:
            if 1 <= single['count'] < 3:
                result.update({"1~3": result.get('1~3', 0) + 1})
            elif 3 <= single['count'] < 8:
                result.update({"3~8": result.get('3~8', 0) + 1})
            elif 8 <= single['count']:
                result.update({"8+": result.get('8+', 0) + 1})

        value.update({"ticket_favorite_time": result})

        match_condition = {
            "user_id": {"$exists": True, "$ne": None},
            "type": "rent_intention",
            "status": {"$in": ["requested", "assigned", "in_progress", "rejected", "confirmed_video", "booked", "holding_deposit_paid", "checked_in"]}
        }
        if 'date_from' in params and 'date_to' in params:
            match_condition.update({
                "time": {
                    "$gte": params['date_from'],
                    "$lt": params['date_to']
                }
            })
        cursor = m.tickets.aggregate([
            {'$match': match_condition},
            {'$group': {
                "_id": "$user_id",
                "count": {"$sum": 1}
            }}
        ])
        result = {}
        for single in cursor:
            if 1 <= single['count'] < 3:
                result.update({"1~3": result.get('1~3', 0) + 1})
            elif 3 <= single['count'] < 8:
                result.update({"3~8": result.get('3~8', 0) + 1})
            elif 8 <= single['count']:
                result.update({"8+": result.get('8+', 0) + 1})

        value.update({"ticket_request_time": result})

    return value


@f_api('/get-users-portrait-landlord-behavior', params=dict(
    date_from=(datetime, None),
    date_to=(datetime, None)
))
@f_app.user.login.check(force=True, role=['admin', 'jr_admin', 'sales', 'operation'])
def get_users_portrait_landlord_behavior(user, params):
    final_result = {}
    with f_app.mongo() as m:
        cursor = m.tickets.aggregate([
            {'$match': {
                "user_id": {"$exists": True},
                "type": "rent",
                "status": "to rent"
            }},
            {'$group': {
                "_id": "$user_id",
                "count": {"$sum": 1}
            }}
        ])
        result = {}
        for single in cursor:
            if 1 == single['count']:
                result.update({"1": result.get('1', 0) + 1})
            elif 1 < single['count'] <= 3:
                result.update({"1~3": result.get('1~3', 0) + 1})
            elif 3 < single['count']:
                result.update({"3+": result.get('3+', 0) + 1})
        # print "房源发布数量:"
        # print json.dumps(result, indent=2)
        final_result.update({'rent_ticket_renting_count_distribution': result})

        def get_direct_data_from_newest_ticket(field_name):
            cursor = m.tickets.aggregate([
                {'$match': {
                    "user_id": {"$exists": True},
                    "type": "rent",
                    "status": "to rent",
                    field_name: {'$exists': True}
                }},
                {'$group': {
                    '_id': "$user_id",
                    'count': {'$sum': 1},
                    'landlord_type_id': {'$first': '$' + field_name}
                }},
                {'$match': {
                    'count': {'$lte': 1}
                }},
                {'$group': {
                    '_id': '$landlord_type_id',
                    'count': {'$sum': 1}
                }}
            ])
            source = {}
            for single in cursor:
                source.update({unicode(single['_id']): single['count']})

            cursor = m.tickets.aggregate([
                {'$match': {
                    "user_id": {"$exists": True},
                    "type": "rent",
                    "status": "to rent",
                    field_name: {'$exists': True}
                }},
                {'$group': {
                    '_id': "$user_id",
                    'count': {'$sum': 1},
                    'time': {'$max': '$time'}
                }},
                {'$match': {
                    'count': {'$gt': 1}
                }}
            ])
            search_condition = {'$or': []}
            for single in cursor:
                search_condition['$or'].append({
                    'user_id': ObjectId(single['_id']),
                    'time': single['time'],
                    'type': 'rent',
                    'status': 'to rent'
                })

            tickets = m.tickets.aggregate([
                {'$match': search_condition},
                {'$project': {
                    '_id': False,
                    'landlord_type_id': '$' + field_name
                }}
            ])

            for ticket in tickets:
                source.update({unicode(ticket['landlord_type_id']): source.get(unicode(ticket['landlord_type_id']), 0) + 1})
            return source

        source = get_direct_data_from_newest_ticket('rent_type._id')
        result = {}
        for single in source:
            result.update({f_app.enum.get(ObjectId(single))['value']['zh_Hans_CN']: source[single]})
        # print "出租类型:"
        # print json.dumps(result, indent=2, ensure_ascii=False)
        final_result.update({'rent_ticket_renting_type_distribution': result})

        source = get_direct_data_from_newest_ticket('landlord_type._id')
        result = {}
        for single in source:
            result.update({f_app.enum.get(ObjectId(single))['value']['zh_Hans_CN']: source[single]})
        # print "房东类型:"
        # print json.dumps(result, indent=2, ensure_ascii=False)
        final_result.update({'rent_ticket_renting_landlordtype_distribution': result})

        source = get_direct_data_from_newest_ticket('price.value')
        result_temp = {}
        match_pattern = re.compile('\.[0]*$')
        for single in source:
            result_temp.update({match_pattern.sub('', single): result_temp.get(match_pattern.sub('', single), 0) + source[single]})
        result = {}
        for single in result_temp:
            key = single
            key_float = float(key)
            value = result_temp[key]
            if 0 < key_float <= 100:
                result.update({'0~100': result.get('0~100', 0) + value})
            elif 100 < key_float <= 200:
                result.update({'100~200': result.get('100~200', 0) + value})
            elif 200 < key_float <= 300:
                result.update({'200~300': result.get('200~300', 0) + value})
            elif 300 < key_float <= 400:
                result.update({'300~400': result.get('300~400', 0) + value})
            elif 400 < key_float <= 500:
                result.update({'400~500': result.get('400~500', 0) + value})
            elif 500 < key_float:
                result.update({'500+': result.get('500+', 0) + value})
        # print "租金分布:"
        # print json.dumps(result, indent=2, ensure_ascii=False)
        final_result.update({'rent_ticket_renting_price_distribution': result})

        source = {}
        cursor = m.tickets.aggregate([
            {'$match': {
                "user_id": {"$exists": True},
                "type": "rent",
                "status": "to rent",
                "rent_available_time": {'$exists': True},
                "rent_deadline_time": {'$exists': True}
            }},
            {'$group': {
                '_id': "$user_id",
                'count': {'$sum': 1},
                'rent_available_time': {'$first': '$rent_available_time'},
                'rent_deadline_time': {'$first': '$rent_deadline_time'}
            }},
            {'$match': {
                'count': {'$lte': 1}
            }},
            {'$project': {
                'rent_time_length_days': {'$divide': [{'$subtract': ['$rent_deadline_time', '$rent_available_time']}, 86400000]},
                '_id': False,
            }},
            {'$group': {
                '_id': '$rent_time_length_days',
                'count': {'$sum': 1}
            }}
        ])
        for single in cursor:
            key_float = float(single['_id'])
            value = single['count']
            if 0 < key_float <= 31:
                source.update({'1_month': source.get('1_month', 0) + value})
            elif 31 < key_float <= 91:
                source.update({'1~3_month': source.get('1~3_month', 0) + value})
            elif 91 < key_float <= 183:
                source.update({'3~6_month': source.get('3~6_month', 0) + value})
            elif 183 < key_float <= 365:
                source.update({'6~12_month': source.get('6~12_month', 0) + value})
            elif 365 < key_float:
                source.update({'12+_month': source.get('12+_month', 0) + value})

        cursor = m.tickets.aggregate([
            {'$match': {
                "user_id": {"$exists": True},
                "type": "rent",
                "status": "to rent",
                "rent_available_time": {'$exists': True},
                "rent_deadline_time": {'$exists': True}
            }},
            {'$group': {
                '_id': "$user_id",
                'count': {'$sum': 1},
                'time': {'$max': '$time'}
            }},
            {'$match': {
                'count': {'$gt': 1}
            }},
            {'$project': {
                '_id': True,
                'time': True
            }}
        ])
        search_condition = {'$or': []}
        for single in cursor:
            search_condition['$or'].append({
                'user_id': ObjectId(single['_id']),
                'time': single['time'],
                'type': 'rent',
                'status': 'to rent',
                "rent_available_time": {'$exists': True},
                "rent_deadline_time": {'$exists': True}
            })

        cursor = m.tickets.aggregate([
            {'$match': search_condition},
            {'$project': {
                'rent_time_length_days': {'$divide': [{'$subtract': ['$rent_deadline_time', '$rent_available_time']}, 86400000]},
                '_id': False,
            }},
            {'$group': {
                '_id': '$rent_time_length_days',
                'count': {'$sum': 1}
            }}
        ])
        for single in cursor:
            key_float = float(single['_id'])
            value = single['count']
            if 0 < key_float <= 31:
                source.update({'1_month': source.get('1_month', 0) + value})
            elif 31 < key_float <= 91:
                source.update({'1~3_month': source.get('1~3_month', 0) + value})
            elif 91 < key_float <= 183:
                source.update({'3~6_month': source.get('3~6_month', 0) + value})
            elif 183 < key_float <= 365:
                source.update({'6~12_month': source.get('6~12_month', 0) + value})
            elif 365 < key_float:
                source.update({'12+_month': source.get('12+_month', 0) + value})

        source.update({'no_limit': m.tickets.find({
            "user_id": {"$exists": True},
            "type": "rent",
            "status": "to rent",
            "rent_available_time": {'$exists': True},
            "rent_deadline_time": {'$exists': False}
        }).count()})
        # print "出租多久:"
        # print json.dumps(source, indent=2, ensure_ascii=False)
        final_result.update({'rent_ticket_renting_time_length_plan_distribution': source})

        source = {}
        cursor = m.tickets.aggregate([
            {'$match': {
                "user_id": {"$exists": True},
                "type": "rent",
                "status": "to rent",
                "time": {'$exists': True},
                "rent_available_time": {'$exists': True}
            }},
            {'$group': {
                '_id': "$user_id",
                'count': {'$sum': 1},
                'time': {'$first': '$time'},
                'rent_available_time': {'$first': '$rent_available_time'}
            }},
            {'$match': {
                'count': {'$lte': 1}
            }},
            {'$project': {
                'rent_time_length_days': {'$divide': [{'$subtract': ['$rent_available_time', '$time']}, 86400000]},
                '_id': False,
            }},
            {'$group': {
                '_id': '$rent_time_length_days',
                'count': {'$sum': 1}
            }}
        ])
        for single in cursor:
            key_float = float(single['_id'])
            value = single['count']
            if 0 < key_float <= 7:
                source.update({'1_week': source.get('1_week', 0) + value})
            elif 7 < key_float <= 14:
                source.update({'1~2_week': source.get('1~2_week', 0) + value})
            elif 14 < key_float <= 30:
                source.update({'2~4_week': source.get('2~4_week', 0) + value})
            elif 30 < key_float <= 61:
                source.update({'1~2_month': source.get('1~2_month', 0) + value})
            elif 61 < key_float:
                source.update({'2+_month': source.get('2+_month', 0) + value})

        cursor = m.tickets.aggregate([
            {'$match': {
                "user_id": {"$exists": True},
                "type": "rent",
                "status": "to rent",
                "time": {'$exists': True},
                "rent_available_time": {'$exists': True}
            }},
            {'$group': {
                '_id': "$user_id",
                'count': {'$sum': 1},
                'time': {'$max': '$time'}
            }},
            {'$match': {
                'count': {'$gt': 1}
            }},
            {'$project': {
                '_id': True,
                'time': True
            }}
        ])
        search_condition = {'$or': []}
        for single in cursor:
            search_condition['$or'].append({
                'user_id': ObjectId(single['_id']),
                'time': single['time'],
                'type': 'rent',
                'status': 'to rent',
                "time": {'$exists': True},
                "rent_available_time": {'$exists': True}
            })

        cursor = m.tickets.aggregate([
            {'$match': search_condition},
            {'$project': {
                'rent_time_length_days': {'$divide': [{'$subtract': ['$rent_available_time', '$time']}, 86400000]},
                '_id': False,
            }},
            {'$group': {
                '_id': '$rent_time_length_days',
                'count': {'$sum': 1}
            }}
        ])
        for single in cursor:
            key_float = float(single['_id'])
            value = single['count']
            if 0 < key_float <= 7:
                source.update({'1_week': source.get('1_week', 0) + value})
            elif 7 < key_float <= 14:
                source.update({'1~2_week': source.get('1~2_week', 0) + value})
            elif 14 < key_float <= 30:
                source.update({'2~4_week': source.get('2~4_week', 0) + value})
            elif 30 < key_float <= 61:
                source.update({'1~2_month': source.get('1~2_month', 0) + value})
            elif 61 < key_float:
                source.update({'2+_month': source.get('2+_month', 0) + value})

        # print "提前多久开始租房:"
        # print json.dumps(source, indent=2, ensure_ascii=False)
        final_result.update({'rent_ticket_renting_time_length_ahead_distribution': source})

        cursor = m.log.aggregate([
            {'$match': {
                'type': 'route',
                'route': {'$exists': True},
            }},
            {'$project': {
                '_id': False,
                'route': True,
                'judge': {'$strcasecmp': ['$route', '/property-to-rent/']}
            }},
            {'$match': {
                'judge': 1
            }},
            {'$project': {
                '_id': False,
                'route': True,
                'route_page_pre': {'$substr': ['$route', 0, 18]},
            }},
            {'$match': {
                'route_page_pre': '/property-to-rent/',
                'route': {'$ne': '/property-to-rent/create'}
            }},
            {'$group': {
                '_id': {'$substr': ['$route', 18, 24]},
                'count': {'$sum': 1}
            }},
            {'$group': {
                '_id': '$count',
                'count': {'$sum': 1}
            }}
        ])

        result = {}
        for single in cursor:
            if 0 < single['_id'] <= 30:
                result.update({'0~30': result.get('0~30', 0) + single['count']})
            elif 30 < single['_id'] <= 100:
                result.update({'30~100': result.get('30~100', 0) + single['count']})
            elif 100 < single['_id'] <= 300:
                result.update({'100~300': result.get('100~300', 0) + single['count']})
            elif 300 < single['_id']:
                result.update({'300+': result.get('300+', 0) + single['count']})
        # print "房源详情页面被查看的次数:"
        # print json.dumps(result, indent=2)
        final_result.update({'rent_ticket_renting_page_view_times_distribution': result})

        cursor = m.log.aggregate([
            {'$match': {
                'type': 'route',
                'route': {'$exists': True},
            }},
            {'$project': {
                '_id': False,
                'route': True,
                'judge': {'$strcasecmp': ['$route', '/wechat-poster/']}
            }},
            {'$match': {
                'judge': 1
            }},
            {'$project': {
                '_id': False,
                'route': True,
                'route_page_pre': {'$substr': ['$route', 0, 15]},
            }},
            {'$match': {
                'route_page_pre': '/wechat-poster/',
                # 'route': {'$ne': '/property-to-rent/create'}
            }},
            {'$group': {
                '_id': {'$substr': ['$route', 15, 24]},
                'count': {'$sum': 1}
            }},
            {'$group': {
                '_id': '$count',
                'count': {'$sum': 1}
            }}
        ])

        result = {}
        for single in cursor:
            if 0 < single['_id'] <= 30:
                result.update({'0~30': result.get('0~30', 0) + single['count']})
            elif 30 < single['_id'] <= 100:
                result.update({'30~100': result.get('30~100', 0) + single['count']})
            elif 100 < single['_id'] <= 300:
                result.update({'100~300': result.get('100~300', 0) + single['count']})
            elif 300 < single['_id']:
                result.update({'300+': result.get('300+', 0) + single['count']})
        # print "房源微信Poster页面被查看的次数:"
        # print json.dumps(result, indent=2)
        final_result.update({'rent_ticket_renting_wechat_share_times_distribution': result})

        cursor = m.favorites.aggregate([
            {'$match': {
                'type': 'rent_ticket',
                'status': {'$ne': 'deleted'},
                'user_id': {'$exists': True},
                'ticket_id': {'$exists': True}
            }},
            {'$group': {
                '_id': '$ticket_id',
                'count': {'$sum': 1}
            }},
            {'$group': {
                '_id': '$count',
                'count': {'$sum': 1}
            }}
        ])

        result = {}
        for single in cursor:
            if 1 < single['_id'] <= 3:
                result.update({'1~3': result.get('1~3', 0) + single['count']})
            elif 3 < single['_id'] <= 8:
                result.update({'3~8': result.get('3~8', 0) + single['count']})
            elif 8 < single['_id']:
                result.update({'8+': result.get('8+', 0) + single['count']})
        # print "房源被收藏的次数:"
        # print json.dumps(result, indent=2)
        final_result.update({'rent_ticket_renting_favorite_times_distribution': result})

        cursor = m.tickets.aggregate([
            {'$match': {
                'type': 'rent_intention',
                'status': {'$in': ["requested", "assigned", "in_progress", "rejected", "confirmed_video", "booked", "holding_deposit_paid", "checked_in"]}
            }},
            {'$project': {
                'interested_rent_tickets': True
            }},
            {'$group': {
                '_id': '$interested_rent_tickets',
                'count': {'$sum': 1}
            }},
            {'$group': {
                '_id': '$count',
                'count': {'$sum': 1}
            }}
        ])
        result = {}
        for single in cursor:
            if 1 < single['_id'] <= 3:
                result.update({'1~3': result.get('1~3', 0) + single['count']})
            elif 3 < single['_id'] <= 8:
                result.update({'3~8': result.get('3~8', 0) + single['count']})
            elif 8 < single['_id']:
                result.update({'8+': result.get('8+', 0) + single['count']})
        # print "房源被咨询的次数:"
        # print json.dumps(result, indent=2)
        final_result.update({'rent_ticket_renting_requested_times_distribution': result})

        cursor = m.log.aggregate([
            {'$match': {
                'type': 'route',
                'route': {
                    '$exists': True,
                    '$ne': None
                },
                'route': {'$regex': '^\/api\/1\/rent_ticket'}
            }},
            {'$project': {
                '_id': False,
                'ticket_id': {'$substr': ['$route', 19, 24]},
                'operation': {'$substr': ['$route', 43, 8]}
            }},
            {'$match': {
                'operation': '/refresh',
            }},
            {'$group': {
                '_id': '$ticket_id',
                'count': {'$sum': 1}
            }},
            {'$group': {
                '_id': '$count',
                'count': {'$sum': 1}
            }}
        ])

        result = {}
        for single in cursor:
            if 1 < single['_id'] <= 3:
                result.update({'1~3': result.get('1~3', 0) + single['count']})
            elif 3 < single['_id'] <= 8:
                result.update({'3~8': result.get('3~8', 0) + single['count']})
            elif 8 < single['_id']:
                result.update({'8+': result.get('8+', 0) + single['count']})
        # print "刷新房源的次数:"
        # print json.dumps(result, indent=2)
        final_result.update({'rent_ticket_renting_refresh_times_distribution': result})

        cursor = m.tickets.aggregate([
            {'$match': {
                "user_id": {"$exists": True},
                "type": "rent",
                "status": "to rent",
                "property_id": {'$exists': True}
            }},
            {'$project': {
                '_id': '$property_id',
            }}
        ])
        search_condition = []
        for single in cursor:
            if single['_id'] is None:
                continue
            search_condition.append(single['_id'])
        cursor = m.propertys.aggregate([
            {'$match': {
                '_id': {
                    '$in': search_condition
                }
            }},
            {'$group': {
                '_id': '$maponics_neighborhood._id',
                'count': {'$sum': 1}
            }},
            {'$sort': {
                'count': -1
            }}
        ])
        source = {}
        result = {}
        index = 0
        for single in cursor:
            if single['_id'] is None:
                source.update({'unset': {
                    'count': single['count'],
                    'name': '[unset]'
                }})
                continue
            if index >= 10:
                source.update({'other': {
                    'count': source.get('other', {}).get('count', 0) + single['count'],
                    'name': '[other]'
                }})
            else:
                name = ''
                try:
                    name = f_app.doogal.station.get(single['_id']).get('name', '')
                except:
                    pass
                source.update({unicode(index + 1): {
                    'count': single['count'],
                    'name': name
                }})
                index += 1
        for index in range(1, 11):
            result.update({source[unicode(index)]['name']: source[unicode(index)]['count']})
        result.update({source['other']['name']: source['other']['count']})
        # print "房源位置(街区):"
        # print json.dumps(result, indent=2)
        final_result.update({'rent_ticket_renting_location_neighborhood': result})

        cursor = m.propertys.aggregate([
            {'$match': {
                '_id': {
                    '$in': search_condition
                }
            }},
            {'$project': {
                'station': '$featured_facility.doogal_station',
            }},
            {'$unwind': '$station'},
            {'$group': {
                '_id': '$station',
                'count': {'$sum': 1}
            }},
            {'$sort': {
                'count': -1
            }}
        ])

        source = {}
        result = {}
        index = 0
        for single in cursor:
            if index >= 10:
                source.update({'other': {
                    'count': source.get('other', {}).get('count', 0) + single['count'],
                    'name': '[other]'
                }})
            else:
                name = ''
                try:
                    name = f_app.doogal.station.get(single['_id']).get('name', '')
                except:
                    pass
                source.update({unicode(index + 1): {
                    'count': single['count'],
                    'name': name
                }})
                index += 1
        for index in range(1, 11):
            result.update({source[unicode(index)]['name']: source[unicode(index)]['count']})
        result.update({source['other']['name']: source['other']['count']})
        # print "房源位置(地铁):"
        # print json.dumps(result, indent=2)
        final_result.update({'rent_ticket_renting_location_metro': result})

        cursor = m.propertys.aggregate([
            {'$match': {
                '_id': {
                    '$in': search_condition
                }
            }},
            {'$project': {
                'university': '$featured_facility.hesa_university'
            }},
            {'$unwind': '$university'},
            {'$group': {
                '_id': '$university',
                'count': {'$sum': 1}
            }},
            {'$sort': {
                'count': -1
            }}
        ])

        source = {}
        result = {}
        index = 0
        for single in cursor:
            if index >= 10:
                source.update({'other': {
                    'count': source.get('other', {}).get('count', 0) + single['count'],
                    'name': '[other]'
                }})
            else:
                name = ''
                try:
                    name = f_app.hesa.university.get(single['_id']).get('name', '')
                except:
                    pass
                source.update({unicode(index + 1): {
                    'count': single['count'],
                    'name': name
                }})
                index += 1
        for index in range(1, 11):
            result.update({source[unicode(index)]['name']: source[unicode(index)]['count']})
        result.update({source['other']['name']: source['other']['count']})
        # print "房源位置(大学):"
        # print json.dumps(result, indent=2)
        final_result.update({'rent_ticket_renting_location_university': result})

        cursor = m.propertys.aggregate([
            {'$match': {
                '_id': {
                    '$in': search_condition
                }
            }},
            {'$project': {
                'city': '$city._id'
            }},
            {'$group': {
                '_id': '$city',
                'count': {'$sum': 1}
            }},
            {'$sort': {
                'count': -1
            }}
        ])

        source = {}
        result = {}
        index = 0
        for single in cursor:
            if index >= 10:
                source.update({'other': {
                    'count': source.get('other', {}).get('count', 0) + single['count'],
                    'name': '[other]'
                }})
            else:
                name = ''
                try:
                    name = f_app.geonames.gazetteer.get(single['_id']).get('name', '')
                except:
                    pass
                source.update({unicode(index + 1): {
                    'count': single['count'],
                    'name': name
                }})
                index += 1
        for index in range(1, 11):
            result.update({source[unicode(index)]['name']: source[unicode(index)]['count']})
        result.update({source['other']['name']: source['other']['count']})
        # print "房源位置(城市):"
        # print json.dumps(result, indent=2)
        final_result.update({'rent_ticket_renting_location_city': result})

    return final_result
