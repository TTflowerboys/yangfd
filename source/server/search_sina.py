# coding: utf-8
from __future__ import unicode_literals
import sys
from app import f_app
from datetime import datetime, timedelta
from datetime import date
# from datetime import timedelta
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import Levenshtein
import re
import json
from six.moves import urllib
from pyquery import PyQuery as pq
import random
from time import sleep
import requests
import time

day_shift = int(sys.argv[1]) if len(sys.argv) > 1 else 0  # the date how many days before will be loaded
time_start_hours = 10


def generate_keyword_list(filename):
    wb = load_workbook(filename=filename)
    ws = wb.active
    temp = []
    result = []
    for row in ws.rows:
        for cell in row:
            if cell.value is None:
                if len(temp):
                    result.append(' '.join(temp))
                temp = []
            else:
                if len(temp) >= 2:
                    result.append(' '.join(temp))
                    temp = []
                temp.append(cell.value)
    return result


def get_weibo_search_result(keywords_list):

    analyze_keyword_count_orign = {}
    analyze_keyword_count_date = {}
    analyze_keyword_count_final = {}

    def get_correct_col_index(num):
        if num > 26*26:
            return "ZZ"
        if num >= 26:
            return get_correct_col_index(num/26-1)+chr(num-26+65)
        else:
            return chr(num+65)

    def add_link(sheet, target, link=None):
        if target is None:
            return
        if f_app.util.batch_iterable(target):
            pass
        else:
            for index in range(2, len(sheet.rows)+1):
                cell = sheet[target + unicode(index)]
                if len(cell.value):
                    if link is None:
                        cell.hyperlink = cell.value
                    else:
                        cell.hyperlink = unicode(link)

    def format_fit(sheet):
        simsun_font = Font(name="SimSun")
        alignment_fit_shrink = Alignment(shrink_to_fit=True)
        alignment_fit_wrap = Alignment(wrap_text=True)
        for row in sheet.rows:
            for cell in row:
                cell.font = simsun_font
                cell.alignment = alignment_fit_shrink
        for num, col in enumerate(sheet.columns):
            lenmax = 0
            for cell in col:
                lencur = 0
                if cell.value is None:
                    cell.value = ''
                if isinstance(cell.value, int) or isinstance(cell.value, datetime):
                    lencur = len(unicode(cell.value).encode("GBK"))
                elif cell.value is not None:
                    lencur = len(cell.value.encode("GBK", "replace"))
                if lencur > 150:
                    cell.alignment = alignment_fit_wrap
                    lencur = 150
                if lencur > lenmax:
                    lenmax = lencur

            sheet.column_dimensions[get_correct_col_index(num)].width = lenmax*0.86
            # print "col "+get_correct_col_index(num)+" fit."

    def simplify(keywords):
        # get weibo into a list
        result_list = []
        search_count = 0
        search_times = 0
        keyword_total = len(keywords)
        username = "13545078924"
        password = "bbt12345678"
        with f_app.sina_search(username=username, password=password) as ss:
            for num, keyword in enumerate(keywords):
                print unicode((num, keyword_total)) + " " + keyword
                result_search = {}
                analyze_keyword_count_orign[keyword] = 0
                for page in range(1, 3):
                    try:
                        result_search = ss.search(keyword, page=page)
                        search_times += 1
                    except:
                        # print "page " + unicode(page) + " fail"
                        continue
                    else:
                        if 'ok' not in result_search or result_search['ok'] != 1:
                            # print "page " + unicode(page) + " ok " + unicode(result_search['ok'])
                            continue
                        count = len(result_search['mblogList'])
                        for index in range(count):
                            result_search['mblogList'][index].update({"keyword": keyword})
                        result_list.extend(result_search['mblogList'])
                        search_count += count
                        analyze_keyword_count_orign[keyword] += count
                        # print "page " + unicode(page) + " count " + unicode(count)
                print "count " + unicode(search_count) + " times " + unicode(search_times)
        return result_list

    def reduce_weibo(weibo_list):
        result = []
        today = date.today()
        time_start = datetime(today.year, today.month, today.day, time_start_hours) - timedelta(days=day_shift + 1)
        cleanr = re.compile('<.*?>')
        total = 0
        count = 0
        for single in weibo_list:
            total += 1
            user = single['user']['screen_name']
            text = unicode(re.sub(cleanr, '', single['text']))
            time = single['created_timestamp']
            link = "http://weibo.com/" + unicode(single['user']['id']) + "/" + single['bid']
            single['time'] = datetime.fromtimestamp(time)
            time = datetime.fromtimestamp(time)
            if time < time_start:
                continue
            count += 1
            if single['keyword'] in analyze_keyword_count_date:
                analyze_keyword_count_date[single['keyword']] += 1
            else:
                analyze_keyword_count_date[single['keyword']] = 1
            result.append({
                "name": user,
                "text": text,
                "time": time,
                "link": link,
                "keyword": single['keyword']
            })
        print "after date filter " + unicode(count) + '/' + unicode(total) + " left."
        return result

    def remove_overlap(weibo_list):
        result = []
        result_extra = []
        count = 0
        total = len(weibo_list)
        for single in weibo_list:
            while weibo_list.count(single) > 1:
                weibo_list.remove(single)
        for single_weibo in weibo_list:
            step = 0
            cur_keyword_list = [single_weibo['keyword']]
            if single_weibo in result_extra:
                continue
            for index in weibo_list:
                if index == single_weibo or index in result_extra:
                    continue
                step = Levenshtein.distance(single_weibo['text'], index['text'])
                if step*1.0/len(unicode(single_weibo['text'])) < 0.17 and step*1.0/len(unicode(index['text'])) < 0.17:
                    result_extra.append(index)
                    cur_keyword_list.append(index['keyword'])
            if single_weibo['keyword'] in analyze_keyword_count_final:
                analyze_keyword_count_final[single_weibo['keyword']] += 1
            else:
                analyze_keyword_count_final[single_weibo['keyword']] = 1
            count += 1
            result.append({
                "name": single_weibo['name'],
                "text": single_weibo['text'],
                "time": single_weibo['time'],
                "link": single_weibo['link'],
                "keyword": cur_keyword_list
            })
        print "after remove overlaping " + unicode(count) + '/' + unicode(total) + " left."
        for single in result_extra:
            print single['time']
            print single['text']
        return result

    def crawler_powerapple(forum_id=None):
        if forum_id is None:
            return []
        list_url = 'http://www.powerapple.com/bbs/forums/' + forum_id
        page = 1
        result_list = []
        today = date.today()
        time_start = datetime(today.year, today.month, today.day, time_start_hours) - timedelta(days=day_shift + 1)
        while(page <= 8):
            print 'page: '+unicode(page)
            target_url = list_url + '?page=' + unicode(page)
            page += 1
            max_time = datetime(1970, 1, 1)
            try:
                list_page = f_app.request.get(target_url)
            except:
                continue
            dom = pq(list_page.content)
            topics = dom('li.topic-li')
            for topic_dom in topics:
                try:
                    topic = pq(topic_dom)
                    result = {}
                    array = []
                    if topic.attr('data-stick') == 'false':
                        time = datetime.strptime(topic('div.authortime div.threadtime')[0].text.replace('\n', '').replace(' ', ''), '%Y-%m-%d')
                        if time < time_start:
                            continue
                        result.update({'text': topic('h4')('a span')[0].text})
                        result.update({'link': 'http://www.powerapple.com' + topic('h4')('a span').parent().attr('href')})
                        result.update({'name': topic('div.authortime div.username a')[0].text})
                        try:
                            topic_page = f_app.request.get(result['link'])
                        except:
                            continue
                        try:
                            topic_page_dom = pq(topic_page.content)
                            text_dom = pq(topic_page_dom('div.post-list li')[1])('div.post-main div.postbody')
                            text_dom.find('br').replaceWith('\n')
                            text = text_dom.text()
                        except:
                            continue
                        time = datetime.strptime(topic_page_dom('div.post-list li').eq(1)('div.post-main div.posttime').text(), '%Y-%m-%d %H:%M')
                        print time
                        result.update({'time': unicode(time)})
                        max_time = max(time, max_time)
                        for i in re.split('\.|。|\n', text):
                            if len(i) == 0:
                                continue
                            if len(i) > 35:
                                for pics in re.split(',|，', i):
                                    for example in contact_keyword:
                                        if pics.find(example) != -1:
                                            array.append(pics)
                                            break
                            else:
                                for example in contact_keyword:
                                    if i.find(example) != -1:
                                        array.append(i)
                                        break
                        result.update({'contact_raw': array})
                        result_list.append(result)
                except:
                    continue
            if max_time < time_start:
                break
        return result_list

    def crawler_douban_group(group_name=None):

        def ua_generator():
            ua_list = [
                'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)',
                'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.36',
                'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 7.0; InfoPath.3; .NET CLR 3.1.40767; Trident/6.0; en-IN)',
                'Mozilla/5.0 (Windows; U; MSIE 9.0; WIndows NT 9.0; en-US))',
                'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:40.0) Gecko/20100101 Firefox/40.1',
                'Mozilla/5.0 (Windows NT 6.3; rv:36.0) Gecko/20100101 Firefox/36.0',
            ]
            return random.choice(ua_list)

        if group_name is None:
            return []
        result = []
        if f_app.util.batch_iterable(group_name):
            for single_group_name in group_name:
                single_result = crawler_douban_group(single_group_name)
                result.extend(single_result)
            return result
        index = 0
        today = date.today()
        time_start = datetime(today.year, today.month, today.day, time_start_hours) - timedelta(days=day_shift + 1)
        while(index <= 250):
            max_time = datetime(1970, 1, 1)
            try:
                list_page = f_app.request.get(
                    'http://www.douban.com/group/' + group_name + '/discussion',
                    {'start': index},
                    headers={"User-Agent": ua_generator()}
                )
            except:
                return []
            index += 25
            topics_dom = pq(list_page.content)('div.article tr')
            for topic in topics_dom:
                try:
                    topic_dom = pq(topic)
                    topic_result = {}
                    if topic_dom.attr('class') != 'th':
                        topic_result.update({'title': topic_dom('td.title a').attr('title')})
                        topic_result.update({'link': topic_dom('td.title a').attr('href')})
                        topic_result.update({'author': topic_dom.children().eq(1).text()})
                        try:
                            topic_page = f_app.request.get(
                                topic_result['link'],
                                headers={"User-Agent": ua_generator()}
                            )
                        except:
                            continue
                        sleep(1)
                        topic_page_dom = pq(topic_page.content)('div.topic-doc')
                        create_time = datetime.strptime(topic_page_dom('h3 span.color-green').text(), '%Y-%m-%d %H:%M:%S')
                        max_time = max(max_time, create_time)
                        if create_time < time_start:
                            continue
                        topic_result.update({'time': create_time})
                        try:
                            text_dom = topic_page_dom('p')
                            text_dom.find('br').replaceWith('\n')
                            text = text_dom.text()
                        except:
                            continue
                        array = []
                        for i in re.split('\.|。|\n', text):
                            if len(i) == 0:
                                continue
                            if len(i) > 35:
                                for pics in re.split(',|，', i):
                                    for example in contact_keyword:
                                        if pics.find(example) != -1:
                                            array.append(pics)
                                            break
                            else:
                                for example in contact_keyword:
                                    if i.find(example) != -1:
                                        array.append(i)
                                        break
                        topic_result.update({'contact_raw': array})
                        result.append(topic_result)
                except:
                    continue
            if max_time < time_start:
                break
        return result

    def get_ybirds_city_list(session, select_city):
        now = datetime.now()
        timestamp = int(time.mktime(now.timetuple())*1e3 + now.microsecond/1e3)
        url = 'http://ybirds.com/Home-Class-getHotCity.html?country=1&_=' + unicode(timestamp)
        page = session.get(
            url,
            headers={
                "POST": "/sso/login HTTP/1.1",
                "Host": "ybirds.com",
                "User-Agent": "Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:45.0) Gecko/20100101 Firefox/45.0",
                "Accept": "application/json, text/javascript, */*; q=0.01",
                "Accept-Language": "en-US,en;q=0.5",
                "Accept-Encoding": "gzip, deflate",
                "X-Requested-With": "XMLHttpRequest",
                "Referer": "http://ybirds.com/Home-Class-ChangeCity.html",
                "Connection": "keep-alive",
            }
        )
        match = re.compile('href=\"\\\\(.*?)\">(.*?)<')
        source = match.findall(unicode(page.content, 'unicode-escape'))
        city_list = {}
        for single in source:
            source = single[1].rstrip(' ')
            name = source.split(' ')[-1]
            key = source.replace(' ' + name, '')
            city_list.update({
                key: {
                    'link': 'http://ybirds.com' + single[0],
                    'name': source
                }
            })

        now = datetime.now()
        timestamp = int(time.mktime(now.timetuple())*1e3 + now.microsecond/1e3)
        url = 'http://www.ybirds.com/Home-Class-ajaxGetCity.html?countyID=1&_=' + unicode(timestamp)
        page = session.get(
            url,
            headers={
                "POST": "/sso/login HTTP/1.1",
                "Host": "ybirds.com",
                "User-Agent": "Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:45.0) Gecko/20100101 Firefox/45.0",
                "Accept": "application/json, text/javascript, */*; q=0.01",
                "Accept-Language": "en-US,en;q=0.5",
                "Accept-Encoding": "gzip, deflate",
                "X-Requested-With": "XMLHttpRequest",
                "Referer": "http://ybirds.com/Home-Class-ChangeCity.html",
                "Connection": "keep-alive",
            }
        )
        match = re.compile('href=\"\\\\(.*?)\">[:]*(.*?)<')
        source = match.findall(unicode(page.content, 'unicode-escape'))
        for single in source:
            source = single[1].rstrip(' ')
            name = source.split(' ')[-1]
            key = source.replace(' ' + name, '')
            if key not in select_city:
                continue
            city_list.update({
                key: {
                    'link': 'http://ybirds.com' + single[0],
                    'name': source
                }
            })
        return city_list

    def crawler_ybirds(session, city_name, city_switch_url, page_id=None):

        def ua_generator():
            ua_list = [
                'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)',
                'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.36',
                'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 7.0; InfoPath.3; .NET CLR 3.1.40767; Trident/6.0; en-IN)',
                'Mozilla/5.0 (Windows; U; MSIE 9.0; WIndows NT 9.0; en-US))',
                'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:40.0) Gecko/20100101 Firefox/40.1',
                'Mozilla/5.0 (Windows NT 6.3; rv:36.0) Gecko/20100101 Firefox/36.0',
            ]
            return random.choice(ua_list)

        def key_translate(key):
            key_list = {
                "POSTCODE": "postcode",
                "地址": "location",
                "联系人": "contact",
                "电话": "phone",
                "QQ": "qq",
                "付租方式": "pay_type",
                "房源类型": "house_type",
                "租赁形式": "rent_type",
                "押金": "deposite",
                "租金": "price",
                "出租方": "landlord_type",
            }
            if key in key_list:
                return key_list[key]
            else:
                return key

        if page_id is None:
            return []
        result = []
        if f_app.util.batch_iterable(page_id):
            print "target to " + city_switch_url
            session.get(
                city_switch_url,
                headers={
                    "POST": "/sso/login HTTP/1.1",
                    "Host": "ybirds.com",
                    "User-Agent": "Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:45.0) Gecko/20100101 Firefox/45.0",
                    "Accept": "application/json, text/javascript, */*; q=0.01",
                    "Accept-Language": "en-US,en;q=0.5",
                    "Accept-Encoding": "gzip, deflate",
                    "X-Requested-With": "XMLHttpRequest",
                    "Referer": "http://ybirds.com/Home-Class-ChangeCity.html",
                    "Connection": "keep-alive",
                })
            print 'city_switched to ' + city_name
            print json.dumps(session.cookies.get_dict(), indent=2)
            for single_page_id in page_id:
                single_result = []
                single_result.extend(crawler_ybirds(session, city_name, city_switch_url, single_page_id))
                result.extend(single_result)
            return result
        index = 1
        today = date.today()
        time_start = datetime(today.year, today.month, today.day, time_start_hours) - timedelta(days=day_shift + 1)
        while(index <= 5):
            max_time = datetime(1970, 1, 1)
            try:
                print 'loading page...'
                list_page = session.get(
                    'http://www.ybirds.com/Home-ColumnInfo-entrance?ctgClassID=' + unicode(page_id) + '&p=' + unicode(index),
                    headers={"User-Agent": ua_generator()},
                    timeout=30
                )
                sleep(1)
            except:
                return []
            index += 1
            topics_dom = pq(list_page.content)('div.mainList div.row.hasBigImg')

            # print "."
            for topic in topics_dom:
                try:
                    topic_dom = pq(topic)
                    topic_result = {}
                    topic_result.update({'title': topic_dom('div.title').attr('title')})
                    topic_result.update({'author': topic_dom('div.count span').eq(0).text()})
                    topic_result.update({'link': 'http://www.ybirds.com' + topic_dom('div.title a').attr('href')})
                    create_time = datetime.strptime(topic_dom('div.time').text(), '%Y-%m-%d %H:%M:%S')
                    topic_result.update({'time': create_time})
                    print topic_result['author']
                    print topic_result['time']
                    max_time = max(max_time, create_time)
                    if create_time < time_start:
                        continue
                    try:
                        topic_page = session.get(
                            topic_result['link'],
                            headers={"User-Agent": ua_generator()},
                            timeout=30
                        )
                        sleep(1)
                    except:
                        continue
                    topic_page_dom = pq(topic_page.content)
                    simple_intro = topic_page_dom('div.simpleIntro div.introType')
                    for single_intro in simple_intro:
                        single_intro_dom = pq(single_intro)
                        key = key_translate(single_intro_dom.children().eq(0).text().replace("：", ''))
                        value = single_intro_dom.children().eq(1).text()
                        topic_result.update({key: value})

                    detail_intros = topic_page_dom('div.productDetail div.introType')
                    for detail_intro in detail_intros:
                        single_intro_dom = pq(detail_intro)
                        key = key_translate(single_intro_dom.children().eq(0).text().replace("：", ''))
                        value = single_intro_dom.children().eq(1).text()
                        topic_result.update({key: value})

                    try:
                        text_dom = topic_page_dom('div.textDetail p')
                        text_dom.find('br').replaceWith('\n')
                        text = text_dom.text()
                    except:
                        continue
                    array = []
                    for i in re.split('\.|。|\n', text):
                        if len(i) == 0:
                            continue
                        if len(i) > 35:
                            for pics in re.split(',|，', i):
                                for example in contact_keyword:
                                    if pics.find(example) != -1:
                                        array.append(pics)
                                        break
                        else:
                            for example in contact_keyword:
                                if i.find(example) != -1:
                                    array.append(i)
                                    break
                    topic_result.update({'contact_raw': array})

                    result.append(topic_result)
                except:
                    continue
            if max_time < time_start:
                break
        return result

    def load_block_list(filename=None):
        block_list = {}
        wb_block_list = load_workbook(filename)
        sheet = wb_block_list.active
        for col in sheet.columns:
            key = None
            for index, cell in enumerate(col):
                if index == 0:
                    key = cell.value
                    block_list.update({key: []})
                elif cell.value is not None:
                    if key not in block_list:
                        block_list.update({key: []})
                    block_list[key].append(cell.value)
        return block_list

    def reduce_overlap(old_result):
        new_result = []
        result = []
        for single in old_result:
            if 'title' in single and single.get('title', '') in new_result:
                print single['title']
                continue
            new_result.append(single['title'])
            result.append(single)
        return result
    contact_keyword = ['微信', '电话', '联系', '邮箱', '地址', 'qq', 'QQ', 'wechat', 'WECHAT', 'phone', 'Phone', 'email', 'Wechat']
    header = ['微薄帐号', '时间', '链接地址', '最新内容', '来源关键词', '长期/短期', '名字', '电话', '微信', '邮箱', 'qq', '地址']
    header_powerapple = ['论坛帐号', '时间', '链接地址', '标题', '联系方式', '长期/短期', '名字', '电话', '微信', '邮箱', 'qq', '地址']
    header_douban_group = ['豆瓣帐号', '时间', '链接地址', '标题', '联系方式', '长期/短期', '名字', '电话', '微信', '邮箱', 'qq', '地址']
    header_ybirds = ['发帖帐号', '时间', '链接地址', '标题', '联系人', '联系方式', '长期/短期', '名字', '电话', '微信', '邮箱', 'qq', 'POSTCODE', '地址', '付租方式', '房源类型', '租赁形式', '押金', '租金', '出租方']
    result_weibo = []
    result_powerapple = []
    result_douban_group = []
    result_ybirds = {}

    block_list = load_block_list('block_list.xlsx')
    session = requests.session()

    result_weibo = remove_overlap(reduce_weibo(simplify(keywords_list)))
    result_powerapple = crawler_powerapple('10141')
    result_douban_group = reduce_overlap(crawler_douban_group(['ukhome', '436707', '338873', 'LondonHome']))

    city_list = {}
    retry = 0
    while(retry < 5):
        try:
            retry += 1
            city_list = get_ybirds_city_list(session, ['Coventry', 'Newcastle', 'Liverpool', 'Cardiff'])
            print 'city list recived'
            retry = 5
        except:
            pass

    retry = 0
    while(retry < 5):
        try:
            retry += 1
            for single_city in city_list:
                print single_city + ' starting...'
                result = crawler_ybirds(session, city_list[single_city]['name'], city_list[single_city]['link'], ['9', '10'])
                result_ybirds.update({city_list[single_city]['name']: reduce_overlap(result)})
            retry = 5
        except:
            pass

    wb = Workbook()
    ws_weibo = wb.active
    ws_powerapple = wb.create_sheet()
    ws_douban_group = wb.create_sheet()
    ws_weibo_block_list = wb.create_sheet()
    ws_powerapple_block_list = wb.create_sheet()
    ws_douban_group_block_list = wb.create_sheet()

    ws_weibo.title = '微博'
    ws_powerapple.title = '超级苹果论坛'
    ws_douban_group.title = '豆瓣小组'
    ws_weibo_block_list.title = '微博_黑名单'
    ws_powerapple_block_list.title = '超级苹果论坛_黑名单'
    ws_douban_group_block_list.title = '豆瓣小组_黑名单'

    ws_weibo.append(header)
    ws_weibo_block_list.append(header)
    for single in result_weibo:
        if single['name'] in block_list['weibo']:
            ws_weibo_block_list.append([
                single['name'],
                single['time'],
                single['link'],
                single['text'],
                ' & '.join(single['keyword'])
            ])
        else:
            ws_weibo.append([
                single['name'],
                single['time'],
                single['link'],
                single['text'],
                ' & '.join(single['keyword'])
            ])
    add_link(ws_weibo, 'C')
    format_fit(ws_weibo)
    add_link(ws_weibo_block_list, 'C')
    format_fit(ws_weibo_block_list)

    ws_powerapple.append(header_powerapple)
    ws_powerapple_block_list.append(header_powerapple)
    for single in result_powerapple:
        if single['name'] in block_list['powerapple']:
            ws_powerapple_block_list.append([
                single['name'],
                single['time'],
                single['link'],
                single['text'],
                '\n'.join(single['contact_raw'])
            ])
        else:
            ws_powerapple.append([
                single['name'],
                single['time'],
                single['link'],
                single['text'],
                '\n'.join(single['contact_raw'])
            ])
    add_link(ws_powerapple, 'C')
    format_fit(ws_powerapple)
    add_link(ws_powerapple_block_list, 'C')
    format_fit(ws_powerapple_block_list)

    ws_douban_group.append(header_douban_group)
    ws_douban_group_block_list.append(header_douban_group)
    for single in result_douban_group:
        if single['author'] in block_list['douban']:
            ws_douban_group_block_list.append([
                single['author'],
                single['time'],
                single['link'],
                single['title'],
                '\n'.join(single['contact_raw'])
            ])
        else:
            ws_douban_group.append([
                single['author'],
                single['time'],
                single['link'],
                single['title'],
                '\n'.join(single['contact_raw'])
            ])
    add_link(ws_douban_group, 'C')
    format_fit(ws_douban_group)
    add_link(ws_douban_group_block_list, 'C')
    format_fit(ws_douban_group_block_list)

    ws_ybirds = {}
    # ws_ybirds_block_list = {}
    for single_city in result_ybirds:

        ws_ybirds[single_city] = wb.create_sheet()
        ws_ybirds[single_city].title = single_city
        ws_ybirds[single_city].append(header_ybirds)

        for single in result_ybirds[single_city]:
            if single['author'] not in block_list['ybirds']:
                ws_ybirds[single_city].append([
                    single['author'],
                    single['time'],
                    single['link'],
                    single['title'],
                    single.get('contact', ''),
                    '\n'.join(single['contact_raw']),
                    '',
                    '',
                    single.get('phone', ''),
                    '',
                    '',
                    single.get('qq', ''),
                    single.get('postcode', ''),
                    single.get('location'),
                    single.get("pay_type"),
                    single.get("house_type"),
                    single.get("rent_type", ''),
                    single.get("deposite", ''),
                    single.get("price", ''),
                    single.get("landlord_type", '')
                ])
        add_link(ws_ybirds[single_city], 'C')
        format_fit(ws_ybirds[single_city])

    today = date.today()
    if day_shift:
        wb.save('weibo_search' + unicode(today - timedelta(days=day_shift)) + '~' + unicode(today) + '.xlsx')
    else:
        wb.save('weibo_search ' + unicode(today) + '.xlsx')

    # wb = Workbook()
    # ws = wb.active
    #
    # ws.append(["keyword", "源数据量", "按日期筛选后剩余", "去重以后剩余"])
    #
    # for single in analyze_keyword_count_orign:
    #     ws.append([
    #         single,
    #         unicode(analyze_keyword_count_orign[single]) if single in analyze_keyword_count_orign else unicode(0),
    #         unicode(analyze_keyword_count_date[single]) if single in analyze_keyword_count_date else unicode(0),
    #         unicode(analyze_keyword_count_final[single]) if single in analyze_keyword_count_final else unicode(0)
    #     ])
    #
    # format_fit(ws)
    # if day_shift:
    #     wb.save('weibo_search_keyword_analyze' + unicode(today - timedelta(days=day_shift)) + '~' + unicode(today)+'.xlsx')
    # else:
    #     wb.save('weibo_search_keyword_analyze' + unicode(today)+'.xlsx')


list_keyw = generate_keyword_list("keywords_search_weibo.xlsx")
get_weibo_search_result(list_keyw)
